# -*- coding: utf-8 -*-
from six import iteritems
from odoo import models, fields, api, _, tools, exceptions
import openpyxl as px
import io
import base64
from odoo.addons.robo.wizard.robo_company_settings import get_all_values, convert_to_string, get_mapped, ImportRecord
from dateutil.relativedelta import relativedelta

asset_mapping = {
    u'Paskutinio nudėvėjimo data': 'date_last_depreciation',
    u'Duomenų data': 'date',
    u'Nudėvėta suma': 'depreciation_amount',
    u'Pavadinimas': 'name',
    u'Dėvimas': 'running',
    u'Įsigijimo vertė': 'original_value',
    u'Skaičiavimo metodika': 'method',
    u'Likvidacinė vertė': 'salvage_value',
    u'Turto tipas': 'category',
    u'Įsigijimo data': 'pirkimo_data',
    u'Unikalus kodas': 'code',
    u'Analitinis kodas': 'analytic_code',
    u'Liko mėnesių nudėvėjimo': 'method_number',
    u'Įsigijimo kiekis': 'original_quantity',
    u'Likutinis kiekis': 'quantity_residual'
}


def import_assets(self, import_file):
    env = self.sudo().env
    asset_obj = env['account.asset.asset']
    category_obj = env['account.asset.category']
    AnalyticAccount = env['account.analytic.account']
    imported_assets = asset_obj
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Ilgalaikis turtas')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, asset_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        test_vals = record.get_dict()
        if test_vals['name']:
            for key, val in iteritems(test_vals):
                if val is None and key not in ['pirkimo_data', 'unknown', 'method_number', 'analytic_code']:
                    raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        else:
            continue
        name = convert_to_string(record.name)
        date = record.date
        date_last_depreciation = record.date_last_depreciation
        try:
            date_next_depreciation = date_last_depreciation + relativedelta(months=1, day=1)
        except:
            raise exceptions.UserError(
                _('Neteisingas datos formatas %s eilutėje. Palaikomas formatas: YYYY-MM-DD') % (i + 1))
        depreciation_amount = record.depreciation_amount
        running = record.running
        original_value = record.original_value
        original_quantity = record.original_quantity
        quantity_residual = record.quantity_residual
        if tools.float_compare(original_quantity, 0.0, precision_digits=3) < 0:
            raise exceptions.UserError(_('Eilutėje {0} įvestas neigiamas pradinis turto kiekis').format(i + 1))
        if tools.float_compare(quantity_residual, 0.0, precision_digits=3) < 0:
            raise exceptions.UserError(_('Eilutėje {0} įvestas neigiamas likutinis turto kiekis').format(i + 1))
        method = convert_to_string(record.method)
        used_method = 'linear'
        if method and isinstance(method, tuple([str, unicode])):
            if method == 'Dvigubo balanso':
                used_method = 'degressive'
        salvage_value = record.salvage_value
        category = record.category
        code = convert_to_string(record.code)
        analytic_code = convert_to_string(record.analytic_code)
        analytic_account = AnalyticAccount
        if analytic_code:
            analytic_account = AnalyticAccount.search([('code', '=', analytic_code)], limit=1)
            if not analytic_account:
                raise exceptions.UserError(_('Neteisingai nurodytas analitinis kodas %s eilutėje.') % (i + 1))
        try:
            value = original_value - depreciation_amount
        except:
            raise exceptions.UserError(_('Nepavyksta apskaičiuoti ilgalaikio turto vertės %s eilutėje. '
                                         'Įvestos įsigijimo ir nusidėvėjimo vertės turi būti skaitinės.') % (i + 1))
        category_id = False
        if category and isinstance(category, tuple([str, unicode])):
            category_id = category_obj.search([('name', 'like', category)], limit=1)
        if not category_id and category and isinstance(category, tuple([str, unicode])):
            category_id = category_obj.search([('name', 'like', category.split(' ')[0])], limit=1)
        if not category_id:
            raise exceptions.UserError(_('Neteisingai nurodyta ilgalaikio turto kategorija %s eilutėje.') % (i + 1))
        try:
            method_number = int(record.method_number)
        except:
            method_number = False
        validate = True
        if running and isinstance(running, tuple([str, unicode])):
            if running in ['Ne', 'NE', 'ne']:
                validate = False
        asset_values = {
            'name': name,
            'date': date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT),
            'date_first_depreciation': date_next_depreciation.strftime(tools.DEFAULT_SERVER_DATE_FORMAT),
            'value': value,
            'original_value': original_value,
            'method': used_method,
            'salvage_value': salvage_value,
            'category_id': category_id.id,
            'method_number': method_number or category_id.method_number,
            'method_period': category_id.method_period,
            'pirkimo_data': record.pirkimo_data or False,
            'code': code,
            'account_analytic_id': analytic_account.id,
            'original_quantity': original_quantity,
            'quantity': quantity_residual,
        }
        asset_id = asset_obj.with_context(is_being_imported=True).create(asset_values)
        if validate:
            asset_id.validate()
        imported_assets += asset_id
    imported_assets.create_starting_accounting_values()


class RoboCompanySettings(models.TransientModel):
    _inherit = 'robo.company.settings'

    import_assets = fields.Binary()
    validate_assets_automatically = fields.Boolean(string='Validate assets automatically when confirming invoice')

    @api.model
    def default_get(self, field_list):
        res = super(RoboCompanySettings, self).default_get(field_list)
        company = self.env.user.sudo().company_id
        res['validate_assets_automatically'] = company.validate_assets_automatically
        return res

    @api.model
    def _get_company_policy_field_list(self):
        res = super(RoboCompanySettings, self)._get_company_policy_field_list()
        res.append('validate_assets_automatically')
        return res

    @api.multi
    def set_default_import(self):
        """
        Calls threaded import preparation method
        for stock data.
        :return: None
        """
        super(RoboCompanySettings, self).set_default_import()
        self.threaded_import_prep(
            'import_assets',
            function=import_assets,
            imported_file=self.import_assets
        )

    @api.multi
    def _set_asset_number(self):
        if not self.env.user.has_group('robo_basic.group_robo_premium_accountant'):
            raise exceptions.ValidationError(_('Neturite teisės keisti ilgalaikio turto numerį'))
        if self.asset_padding < 3:
            raise exceptions.UserError(_('Ilgalaikio turto numeruotės dydis turi būti bent 3 skaitmenys'))
        if self.asset_next_number < 0:
            raise exceptions.UserError(_('Ilgalaikio turto numeruotės sekantis numeris turi būti didesnis už 0'))
        if len(str(self.asset_next_number)) > self.asset_padding:
            raise exceptions.UserError(
                _('Ilgalaikio turto numeruotės sekantis numeris negali būti didesnis nei sekos dydis'))

        next_number = str(self.asset_prefix or '') + str(self.asset_next_number or '').zfill(self.asset_padding) + str(
            self.asset_suffix)
        asset_exists = self.env['account.asset.asset'].search_count([('code', '=', next_number)])
        if asset_exists:
            raise exceptions.UserError(_('Ilgalaikis turtas su šiuo kodu jau egzistuoja'))
        sequence = self.env['ir.sequence'].search([('code', '=', 'ASSETS')], limit=1)
        if not sequence:
            raise exceptions.ValidationError(
                _('Ilgalaikio turto seka nerasta, susisiekite su sistemos administratoriumi'))
        sequence.sudo().write({
            'prefix': self.asset_prefix,
            'suffix': self.asset_suffix,
            'padding': self.asset_padding,
            'number_next_actual': self.asset_next_number
        })

    @api.multi
    def save_numberings(self):
        super(RoboCompanySettings, self).save_numberings()
        if self._context.get('assets'):
            self._set_asset_number()


RoboCompanySettings()
