# -*- coding: utf-8 -*-
from odoo import models, api, tools, _
import openpyxl as px
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill
import openpyxl.utils.cell
import cStringIO as StringIO
from datetime import datetime

XLS_EXT = 'xlsx'
class HrPayslip(models.Model):

    _inherit = 'hr.payslip'

    @api.multi
    def export_payslips(self):
        self.filtered(lambda r: r.state == 'draft').compute_sheet()
        headers1_sort = [
            _(u'Pagrindinis atlyginimas'),
            _(u'Darbas poilsio dienomis'),
            _(u'Viršvalandžiai darbo dienomis'),
            _(u'Viršvalandžiai naktį'),
            _(u'Naktinė pamaina'),
            _(u'Šventinės naktinės valandos'),
            _(u'Viršvalandžiai šventiniai / savaitgaliais'),
            _(u'Prastova'),
            _(u'Priemokos'),
            _(u'Išeitinė kompensacija'),
            _(u'Priedas'),
            _(u'Premijos'),
            _(u'Kompensacijos'),
            _(u'Dienpinigiai'),
            _(u'Nedarbingumo išmoka'),
            _(u'Apmokėjimas tėvams (motinoms)'),
            _(u'Kasmetinės atostogos'),
            _(u'Kompensacija už sukauptas atostogas'),
            _(u'Avansas'),
            _(u'Invalidumo išmoka'),
            _(u'Pajamos natūra'),
            _(u'Papildomos polsio dienos už darbą poilsio dienomis'),
            _(u'Apmokamas poilsis'),
            _(u'Neįvykdyta darbo laiko norma'),
            _(u'Viršyta darbo laiko norma'),
        ]
        headers2_sort = [
            _(u'Gyventojų pajamų mokestis (GPM)'),
            _(u'Darbuotojo socialinis draudimas (išskyrus papildomą pensijų kaupimą)'),
            _(u'Papildomas pensijų kaupimas'),
            _(u'Išmokėti atostoginiai'),
            _(u'Pajamos natūra'),
            _(u'Kitos išskaitos'),
        ]
        headers3_sort = [
            _(u'Atlyginimas (NETO)'),
            _(u'Išmokėtas Avansas'),
            _(u'Mokėtinas atlyginimas'),
            _(u'Bendras mokėtinas atlyginimas'),
            _(u'Darbdavio sodra'),
        ]
        headers0 = [
            _(u'Eil. nr.'),
            _(u'Vardas, pavardė'),
            _(u'Profesija'),
            _(u'Skyrius'),
            _(u'Data'),
            _(u'Sutarties nr.'),
            _(u'Tabelio nr.'),
        ]
        headers1 = {
            _(u'Pagrindinis atlyginimas'): ['BV', 'BM', 'BUD'],
            _(u'Darbas poilsio dienomis'): ['DP'],
            _(u'Viršvalandžiai darbo dienomis'): ['VD'],
            _(u'Viršvalandžiai naktį'): ['VDN'],
            _(u'Naktinė pamaina'): ['DN'],
            _(u'Šventinės naktinės valandos'): ['SNV'],
            _(u'Neįvykdyta darbo laiko norma'): ['NDL'],
            _(u'Viršyta darbo laiko norma'): ['VDL'],
            _(u'Viršvalandžiai šventiniai / savaitgaliais'): ['VSS'],
            _(u'Prastova'): ['PN'],
            _(u'Priemokos'): ['P', 'PNVDU', 'KR', 'MA'],
            _(u'Išeitinė kompensacija'): ['IST'],
            _(u'Priedas'): ['PD', 'PDN', 'PDNM'],
            _(u'Premijos'): ['PR'],
            _(u'Kompensacijos'): ['KOMP', 'KKPD'],
            _(u'Dienpinigiai'): ['KM'],
            _(u'Nedarbingumo išmoka'): ['L'],
            _(u'Apmokėjimas tėvams (motinoms)'): ['T'],
            _(u'Papildomos polsio dienos už darbą poilsio dienomis'): ['V'],
            _(u'Kasmetinės atostogos'): ['A'],
            _(u'Kompensacija už sukauptas atostogas'): ['AK'],
            _(u'Avansas'): ['AVN'],
            _(u'Invalidumo išmoka'): ['INV'],
            _(u'Pajamos natūra'): ['NTR']
        }
        headers2 = {
            _(u'Gyventojų pajamų mokestis (GPM)'): ['GPM'],
            _(u'Darbuotojo socialinis draudimas (išskyrus papildomą pensijų kaupimą)'): ['SDB'],
            _(u'Papildomas pensijų kaupimas'): ['SDP'],
            _(u'Pajamos natūra'): ['NTR'],
            _(u'Išmokėti atostoginiai'): ['AM'],
            _(u'Kitos išskaitos'): ['IŠSK'],
        }
        headers3 = {
            _(u'Atlyginimas (NETO)'): ['NET'],
            _(u'Išmokėtas Avansas'): ['AVN'],
            _(u'Mokėtinas atlyginimas'): ['M'],
            _(u'Bendras mokėtinas atlyginimas'): ['BENDM'],
            _(u'Darbdavio sodra'): ['SDD'],
        }
        index = 1
        all_data = {}
        filled_headers1 = []
        filled_headers2 = []
        filled_headers3 = []
        filled_headers_other_additions = []
        other_payments = self.env['hr.employee.payment'].search([
            ('type', '=', 'other'),
            ('date', '<=', max(self.mapped('date_to'))),
            ('date', '>=', min(self.mapped('date_from'))),
            ('state', '=', 'done')
        ])
        for payslip in self:
            leaves_accumulation_type = payslip.employee_id.sudo().with_context(
                date=payslip.date_to).leaves_accumulation_type
            if isinstance(payslip.employee_id.id, (int, long)):
                remaining_leaves = payslip.employee_id.sudo()._get_remaining_days(
                    date=payslip.date_to, accumulation_type=leaves_accumulation_type)
            else:
                remaining_leaves = 0
            label = _(' k.d.') if leaves_accumulation_type == 'calendar_days' else _(' d.d.')
            remaining_leaves_label = '{0:.2f}'.format(remaining_leaves).replace('.', ',') + label

            if payslip not in all_data:
                all_data[payslip] = {}
            viso_priskaityta = 0.0
            kiti_priskaitymai = 0.0
            viso_atskaityta = 0.0
            for key, codes in headers1.items():
                for code in codes:
                    amount = sum(payslip.line_ids.filtered(lambda r: r.code.strip() == code).mapped('amount'))
                    amount = tools.float_round(amount, precision_digits=2)
                    if key != _(u'Avansas'):
                        viso_priskaityta += amount
                    if not tools.float_is_zero(amount, precision_digits=2) and key not in filled_headers1:
                        filled_headers1.append(key)
                    if key not in all_data[payslip]:
                        all_data[payslip][key] = amount
                    else:
                        all_data[payslip][key] += amount
            for key, codes in headers2.items():
                for code in codes:
                    amount = sum(payslip.line_ids.filtered(lambda r: r.code.strip() == code).mapped('amount'))
                    amount = tools.float_round(amount, precision_digits=2)
                    viso_atskaityta += amount
                    if not tools.float_is_zero(amount, precision_digits=2) and key not in filled_headers2:
                        filled_headers2.append(key)
                    if key == _(u'Pajamos natūra'):
                        continue
                    if key not in all_data[payslip]:
                        all_data[payslip][key] = amount
                    else:
                        all_data[payslip][key] += amount
            for key, codes in headers3.items():
                for code in codes:
                    amount = sum(payslip.line_ids.filtered(lambda r: r.code.strip() == code).mapped('amount'))
                    amount = tools.float_round(amount, precision_digits=2)
                    if not tools.float_is_zero(amount, precision_digits=2) and key not in filled_headers3:
                        filled_headers3.append(key)
                    if key not in all_data[payslip]:
                        all_data[payslip][key] = amount
                    else:
                        all_data[payslip][key] += amount
                if key == _(u'Mokėtinas atlyginimas'):
                    amount_issk = sum(payslip.line_ids.filtered(lambda r: r.code.strip() == 'IŠSK').mapped('amount'))
                    if key not in all_data[payslip]:
                        all_data[payslip][key] = -amount_issk
                    else:
                        all_data[payslip][key] -= amount_issk
            for other_line in payslip.other_line_ids:
                amount = other_line.amount
                amount = tools.float_round(amount, precision_digits=2)
                selection_dict = dict(self.env['hr.payslip.other.line'].fields_get(allfields=['type'])['type']['selection'])
                type_str = selection_dict.get(other_line.type, '')
                key = '%s (%s)' % (other_line.name, type_str)
                if key not in all_data[payslip]:
                    all_data[payslip][key] = amount
                else:
                    all_data[payslip][key] += amount
                if not tools.float_is_zero(amount, precision_digits=2):
                    if other_line.type == 'priskaitymai':
                        kiti_priskaitymai += amount
                        viso_atskaityta += amount
                        if key not in filled_headers_other_additions:
                            filled_headers_other_additions.append(key)
                    elif other_line.type in ['gpm', 'sdb']:
                        viso_atskaityta -= amount
                        if key not in filled_headers2:
                            headers2_sort.append(key)
                            filled_headers2.append(key)
                    elif other_line.type == 'sdd' and key not in filled_headers3:
                        if key not in filled_headers3:
                            headers3_sort.append(key)
                            filled_headers3.append(key)

            date_from = max(payslip.date_from, payslip.contract_id.date_start)
            date_to = min(payslip.date_to, payslip.contract_id.date_end or payslip.date_to)
            payslip_payments = other_payments.filtered(
                lambda p: p.employee_id == payslip.employee_id and date_from <= p.date <= date_to
            )
            for payment in payslip_payments:
                amount = payment.amount_bruto
                amount = tools.float_round(amount, precision_digits=2)
                selection_dict = dict(self.env['hr.employee.payment'].fields_get(allfields=['type'])['type']['selection'])
                type_str = selection_dict.get(payment.type, '')
                key = '%s (%s)' % (payment.description, type_str)
                if key not in all_data[payslip]:
                    all_data[payslip][key] = amount
                else:
                    all_data[payslip][key] += amount
                if not tools.float_is_zero(amount, precision_digits=2):
                    kiti_priskaitymai += amount
                    viso_atskaityta += payment.amount_paid
                    if key not in filled_headers_other_additions:
                        filled_headers_other_additions.append(key)
                    key = '%s (%s)' % (payment.description, 'GPM')
                    all_data[payslip][key] = payment.amount_gpm
                    if key not in filled_headers2:
                        headers2_sort.append(key)
                        filled_headers2.append(key)
                    key = 'Darbuotojo socialinis draudimas (išskyrus papildomą pensijų kaupimą)'
                    all_data[payslip][key] = payment.amount_sdb
                    if key not in filled_headers2:
                        headers2_sort.append(key)
                        filled_headers2.append(key)
                    key = 'Darbdavio sodra'
                    all_data[payslip][key] = payment.amount_sdd
                    if key not in filled_headers3:
                        headers3_sort.append(key)
                        filled_headers3.append(key)

            all_data[payslip]['priskaitymai'] = viso_priskaityta
            all_data[payslip]['kiti_priskaitymai'] = kiti_priskaitymai
            all_data[payslip]['viso_atskaityta'] = viso_atskaityta
            all_data[payslip]['remaining_leaves_label'] = remaining_leaves_label
        # Header
        lines = []
        line_header = list(headers0)
        line_index = [i for i in xrange(len(headers0))]
        line_header += filled_headers1
        for header in filled_headers1:
            line_index.append(101 + headers1_sort.index(header))
        line_header += [_(u'Viso priskaityta')]
        line_index.append(150)
        line_header += filled_headers_other_additions
        for header in filled_headers_other_additions:
            line_index.append(151 + filled_headers_other_additions.index(header))
        line_header += [_(u'Kiti priskaitymai')]
        line_index.append(200)
        line_header += [_(u'Bendrai priskaityta')]
        line_index.append(201)
        line_header += filled_headers2
        for header in filled_headers2:
            line_index.append(301 + headers2_sort.index(header))
        line_header += [_(u'Viso atskaityta')]
        line_index.append(400)
        line_header += filled_headers3
        for header in filled_headers3:
            line_index.append(401 + headers3_sort.index(header))
        line_header += [_(u'Sukauptų atostogų likutis')]
        line_index.append(405)
        line_header += [_(u'Atlyginimas nurodytas sutartyje')]
        line_index.append(410)
        line_header += [_(u'Darbo vietos kaštai')]
        line_index.append(415)
        line_zipped = zip(line_index, line_header)
        line_zipped.sort()
        line_sorted = [y for x, y in line_zipped]
        lines.append(line_sorted)
        # Lines
        for payslip, vals in all_data.items():
            line = [index, payslip.employee_id.name or '', payslip.employee_id.job_id.name or '',
                    payslip.contract_id.with_context(date=payslip.date_from).appointment_id.department_id.name or
                    payslip.employee_id.department_id.name or '',
                    datetime.strptime(payslip.date_from, tools.DEFAULT_SERVER_DATE_FORMAT).strftime('%Y%m'),
                    payslip.contract_id.name or '', payslip.employee_id.tabelio_numeris or '']
            for header in filled_headers1:
                if header not in vals:
                    line.append(0.0)
                else:
                    line.append(vals[header])
            line.append(vals['priskaitymai'])
            for header in filled_headers_other_additions:
                if header not in vals:
                    line.append(0.0)
                else:
                    line.append(vals[header])
            line.append(vals['kiti_priskaitymai'])
            line.append(vals['priskaitymai'] + vals['kiti_priskaitymai'])
            for header in filled_headers2:
                if header not in vals:
                    line.append(0.0)
                else:
                    line.append(vals[header])
            line.append(vals['viso_atskaityta'])
            for header in filled_headers3:
                if header not in vals:
                    line.append(0.0)
                else:
                    line.append(vals[header])
            line.append(vals['remaining_leaves_label'])

            apps = payslip.contract_id.appointment_ids.filtered(lambda a: a.date_start <= payslip.date_to and
                                                                         (not a.date_end or a.date_end >= payslip.date_from))
            app = apps[0] if apps else False
            if app:
                line.append(app.wage)
            else:
                line.append(_(u'Nepavyko nustatyti'))

            # workplace costs (Darbo vietos kaštai)
            line.append(vals.get('priskaitymai', 0.0) + vals.get('Darbdavio sodra', 0.0))
            zipped_line = zip(line_index, line)
            zipped_line.sort()
            sorted_line = [y for x, y in zipped_line]
            lines.append(sorted_line)
            index += 1

        wb = px.Workbook()
        ws = wb.active
        ws.title = _('Algalapiai')
        thin = Side(style="thin")
        bold = Font(bold=True)
        border = Border(bottom=thin)
        c_style = NamedStyle(name='style', font=bold, border=border)
        # Sort lines by date field excluding the header line
        if len(lines) > 2:
            header_line = lines[0]
            lines = lines[1:len(lines)]
            lines.sort(key=lambda x: x[4])
            lines.insert(0, header_line)
        for r, rows in enumerate(lines):
            for c, col in enumerate(rows):
                if r == 0:
                    cell = ws.cell(r + 1, c + 1, col)
                    cell.style = c_style
                else:
                    ws.cell(r + 1, c + 1).value = col

        ws.freeze_panes = 'C2'

        # Set column width
        for i, col in enumerate(lines[0]):
            size = len(str(col))
            for l in lines:
                col_size = len(str(l[i]))
                if col_size > size:
                    size = col_size
            col_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = size

        f = StringIO.StringIO()
        wb.save(f)
        base64_file = f.getvalue().encode('base64')
        if len(self.mapped('payslip_run_id')) == 1:
            date_dt = datetime.strptime(self.mapped('payslip_run_id').date_start, tools.DEFAULT_SERVER_DATE_FORMAT)
            filename = 'Algalapiai %s-%s.%s' % (date_dt.year, date_dt.month, XLS_EXT)
        else:
            filename = 'Algalapiai.%s' % XLS_EXT
        if self._context.get('archive') or self._context.get('threaded_report'):
            return base64_file
        attach_id = self.env['ir.attachment'].create({
            'res_model': 'hr.payslip',
            'res_id': self.ids[0],
            'type': 'binary',
            'name': filename,
            'datas_fname': filename,
            'datas': base64_file,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/download?res_model=hr.payslip&res_id=%s&attach_id=%s' % (self.ids[0], attach_id.id),
            'target': 'self',
        }
