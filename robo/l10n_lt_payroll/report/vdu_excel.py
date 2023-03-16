# -*- coding: utf-8 -*-
from odoo import tools, _, exceptions
from datetime import datetime
import os
from sys import platform
import cStringIO as StringIO
import openpyxl as px
from openpyxl.styles import Font, Border, Side, NamedStyle, colors, Alignment

XLS_EXT = '.xlsx'
class VduExcel:
    def __init__(self, date_dt=datetime.utcnow()):
        self.row_num = 1
        self.sheet = False
        self.cell_style = False
        self.set_cell_style()
        self.load_top(date_dt)

    def set_cell_style(self):
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        thin = Side(style="thin", color=colors.BLACK)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font = Font(size=8, color=colors.BLACK)
        
        self.cell_style = NamedStyle(
            name="Cell Style", alignment=alignment,
            border=border, font=font
        )

    def load_top(self, date_dt):
        """
        loads top of document, mainly used to load legend of the doc and headers
        """
        if platform == 'win32':
            xls_flocation = '\\static\\src\\excel\\vdu_top' + XLS_EXT
        else:
            xls_flocation = '/static/src/excel/vdu_top' + XLS_EXT
        file_loc = os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + xls_flocation
        
        self.wb = px.load_workbook(file_loc)
        self.sheet = self.wb.active
        self.sheet.cell(3, 5).value = date_dt.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
        self.row_num += 5

    def write_line(self, data):
        """
        Writes data of a single employee to the xls spreadsheet.
        """
        s_row_num = self.row_num
        sheet = self.sheet
        vdu_record_data = data.get('salary_info', {}).get('vdu_record_data', [])
        vdu_record_data += [{}] * (3 - len(vdu_record_data))  # Add empty rows

        # Iterate over vdu records and write lines, bump margin accordingly
        for record in vdu_record_data:
            cell = sheet.cell(self.row_num, 2)
            cell.value = record.get('date', '')
            cell.style = self.cell_style

            cell = sheet.cell(self.row_num, 3)
            cell.value = record.get('amount', '')
            cell.style = self.cell_style

            cell = sheet.cell(self.row_num, 4)
            cell.value = record.get('days', '')
            cell.style = self.cell_style

            cell = sheet.cell(self.row_num, 5)
            cell.value = record.get('hours', '')
            cell.style = self.cell_style
            self.row_num += 1
        self.row_num -= 1

        # Merge col 2:5 if less than 2 vdu entries, because minimum
        # height of an employees entry is 2 lines, and theres only 1 vdu entry
        if len(vdu_record_data) < 2:
            for col_num in range(2, 6):
                sheet.merge_cells(
                    start_row=s_row_num, start_column=col_num, 
                    end_row=s_row_num + 1, end_column=col_num
                )
                sheet.cell(s_row_num, col_num).style = self.cell_style
                sheet.cell(s_row_num + 1, col_num).style = self.cell_style
            
        # Merge cells on columns 6:9 if there are 3 or more lines,
        # due to columns 6:9 having max 2 line entries, one for VDU data
        # and another for formula that was used
        if self.row_num - s_row_num > 1:
            start_row = s_row_num
            end_row = self.row_num
            sheet.merge_cells(
                start_row=start_row, start_column=1,
                end_row=end_row, end_column=1
            )
            for row_num in range(start_row, end_row+1):
                sheet.cell(row_num, 1).style = self.cell_style
            
            for col_num in range(6, 10):
                start_row = s_row_num + 1
                end_row = self.row_num
                sheet.merge_cells(
                    start_row=start_row, start_column=col_num,
                    end_row=end_row, end_column=col_num
                )
                for row_num in range(start_row, end_row+1):
                    sheet.cell(row_num, col_num).style = self.cell_style
        else:
            sheet.merge_cells(
                start_row=s_row_num, start_column=1,
                end_row=self.row_num + 1, end_column=1
            )
            for row_num in range(s_row_num, self.row_num + 2):
                sheet.cell(row_num, 1).style = self.cell_style

        cell = sheet.cell(s_row_num, 1)
        cell.value = data.get('employee').name
        cell.style = self.cell_style

        cell = sheet.cell(s_row_num, 6)
        cell.value = self.get_float_display_value(data.get('calculation_info_d', {}).get('amount', 0))
        cell.style = self.cell_style

        cell = sheet.cell(s_row_num + 1, 6)
        cell.value = data.get('calculation_info_d', {}).get('formula', '')
        cell.style = self.cell_style

        cell = sheet.cell(s_row_num, 7)
        cell.value = self.get_float_display_value(data.get('calculation_info_h', {}).get('amount', 0))
        cell.style = self.cell_style

        cell = sheet.cell(s_row_num + 1, 7)
        cell.value = data.get('calculation_info_h', {}).get('formula', '')
        cell.style = self.cell_style

        minimum_daily_wage_adjustment = data.get('minimum_wage_adjustment_d', dict())
        if minimum_daily_wage_adjustment:
            cell = sheet.cell(s_row_num, 8)
            cell.value = self.get_float_display_value(data.get('vdu_d', ''))
            cell.style = self.cell_style

            cell = sheet.cell(s_row_num + 1, 8)
            cell.value = minimum_daily_wage_adjustment.get('formula', '')
            cell.style = self.cell_style

        minimum_hourly_wage_adjustment = data.get('minimum_wage_adjustment_h', dict())
        if minimum_hourly_wage_adjustment:
            cell = sheet.cell(s_row_num, 9)
            cell.value = self.get_float_display_value(data.get('vdu_h', ''))
            cell.style = self.cell_style

            cell = sheet.cell(s_row_num + 1, 9)
            cell.value = minimum_hourly_wage_adjustment.get('formula', '')
            cell.style = self.cell_style

        sheet.row_dimensions[s_row_num].height = 15

        # Increment margin to make space for next record. +2 if
        # margin has not been bumped earlier and +1 otherwise
        if self.row_num == s_row_num:
            self.row_num += 2
        else:
            self.row_num += 1


    @staticmethod
    def get_float_display_value(float_value):
        """
        rounds a float if its a float, if its Falsy returns empty string,
        returns Exception if a supplied argument is not a float and is Truthy,
        this should never happen.
        """
        if float_value and isinstance(float_value, float):
            return round(float_value, 2)
        if not float_value:
            return '0.0'
        raise exceptions.UserError("Klaida eksportuojant xls dokumentą")

    def export(self):
        f = StringIO.StringIO()
        self.wb.save(f)
        return f.getvalue().encode('base64')
