# -*- coding: utf-8 -*-
from __future__ import division
from six import iteritems

import base64
import hashlib
import io
import re
import sys
import types
from datetime import datetime
from pytz import timezone

import openpyxl as px
from dateutil.relativedelta import relativedelta
from odoo.addons.l10n_lt_payroll.model.schedule_template import get_attendances
from odoo.addons.robo.models.robo_tools import sanitize_account_number
from odoo.addons.robo_basic.models.utils import validate_email

from odoo import _, api, exceptions, fields, models, tools
import threading

# ROBO: XLS IMPORT EXAMPLES PROTECTED CELLS PASSWORD: robolabs_xls

SAVIVALDYBES = [
    ('32', 'Akmenės skyriui'),
    ('11', 'Alytaus skyriui'),
    ('34', 'Anykščių skyriui'),
    ('12', 'Birštono skyriui'),
    ('36', 'Biržų skyriui'),
    ('15', 'Druskininkų skyriui'),
    ('42', 'Elektrėnų skyriui'),
    ('45', 'Ignalinos skyriui'),
    ('46', 'Jonavos skyriui'),
    ('47', 'Joniškio skyriui'),
    ('94', 'Jurbarko skyriui'),
    ('49', 'Kaišiadorių skyriui'),
    ('48', 'Kalvarijos skyriui'),
    ('nera', 'Karinių ir joms prilygintų struktūrų skyriui'),
    ('19', 'Kauno skyriui'),
    ('58', 'Kazlų Rūdos skyriui'),
    ('53', 'Kėdainių skyriui'),
    ('54', 'Kelmės skyriui'),
    ('21', 'Klaipėdos skyriui'),
    ('56', 'Kretingos skyriui'),
    ('57', 'Kupiškio skyriui'),
    ('59', 'Lazdijų skyriui'),
    ('18', 'Marijampolės skyriui'),
    ('61', 'Mažeikių skyriui'),
    ('62', 'Molėtų skyriui'),
    ('23', 'Neringos skyriui'),
    ('63', 'Pagėgių skyriui'),
    ('65', 'Pakruojo skyriui'),
    ('25', 'Palangos skyriui'),
    ('27', 'Panevėžio skyriui'),
    ('67', 'Pasvalio skyriui'),
    ('68', 'Plungės skyriui'),
    ('69', 'Prienų skyriui'),
    ('71', 'Radviliškio skyriui'),
    ('72', 'Raseinių skyriui'),
    ('74', 'Rietavo skyriui'),
    ('73', 'Rokiškio skyriui'),
    ('75', 'Skuodo skyriui'),
    ('84', 'Šakių skyriui'),
    ('85', 'Šalčininkų skyriui'),
    ('29', 'Šiaulių skyriui'),
    ('87', 'Šilalės skyriui'),
    ('88', 'Šilutės skyriui'),
    ('89', 'Širvintų skyriui'),
    ('86', 'Švenčionių skyriui'),
    ('77', 'Tauragės skyriui'),
    ('78', 'Telšių skyriui'),
    ('79', 'Trakų skyriui'),
    ('81', 'Ukmergės skyriui'),
    ('82', 'Utenos skyriui'),
    ('38', 'Varėnos skyriui'),
    ('39', 'Vilkaviškio skyriui'),
    ('13', 'Vilniaus skyriui'),
    ('30', 'Visagino skyriui'),
    ('43', 'Zarasų skyriui'),
    ('41', 'Vilniaus r.'),
]

DATE_FORMATS = ['%Y-%m-%d',
                '%d.%m.%Y',
                '%d/%m/%Y',
                '%y-%m-%d',
                '%d/%m/%y',
                '%d.%m.%y',
                '%Y/%m/%d',
                '%y/%m/%d',
                '%Y.%m.%d',
                '%y.%m.%d',
                ]


class RoboImportError(Exception):
    def __init__(self, name, value=None):
        self.name = name
        self.value = value
        self.args = (name, value)


class ImportRecord(object):

    def __init__(self, vals=None, header=None):
        if vals is None:
            vals = []
        if header is None:
            header = []
        self.vals = vals
        self.header = header

    def __getattr__(self, attr):
        res = False
        if not self.vals or not self.header:
            return False
        try:
            index = self.header.index(attr)
        except ValueError:
            return False
        if len(self.vals) >= index:
            res = self.vals[index]
            try:
                res = res.strip()
            except:
                pass
            # Don't try to convert everything to date
            if 'date' in attr or 'data' in attr:
                for fmt in DATE_FORMATS:
                    try:
                        date = datetime.strptime(res, fmt)
                    except (ValueError, TypeError):
                        pass
                    else:
                        res = date
                        break
        return res

    def get_dict(self, attrs=None):
        if attrs is None:
            attrs = []
        result = {}
        if attrs:
            for attr in attrs:
                result[attr] = self.__getattr__(attr)
        else:
            for attr in self.header:
                result[attr] = self.__getattr__(attr)
        return result


def get_all_values(row):
    return [cell.value for cell in row]


def get_mapped(values, mapping):
    result = []
    for val in values:
        mapped = mapping[val] if val in mapping else 'unknown'
        result.append(mapped)
    return result


def convert_to_string(a):
    res = a
    if isinstance(a, types.FloatType):
        a_rounded = tools.float_round(a, precision_digits=2)
        if tools.float_compare(a, a_rounded, precision_digits=2) != 0:
            raise exceptions.UserError(_('%s nėra sveikasis ir negali būti konvertuotas į tekstą'))
        res = str(int(round(a_rounded)))
    if isinstance(a, (types.IntType, types.LongType)):
        res = str(a)
    if not isinstance(res, basestring):
        if not a:
            return ''
        else:
            return a
    else:
        res = res.strip()
    return res


def check_number(num):
    if num and isinstance(num, tuple([int, float, long])):
        return True
    else:
        return False


partner_mapping = {
    u'Partnerio pavadinimas': 'name',
    u'Juridinis': 'is_company',
    u'Įmonės kodas': 'kodas',
    u'PVM kodas': 'vat',
    u'Telefonas': 'phone',
    u'El. paštas': 'email',
    u'Pastabos': 'comment',
    u'Sąskaitos apmokėjimo terminas (dienos) - klientams': 'payment_term',
    u'Sąskaitos apmokėjimo terminas (dienos) - tiekėjams': 'payment_term_supplier',
    u'Korespondecija. Šalies kodas': 'country_id',
    u'Korespondecija. Miestas': 'city',
    u'Korespondecija. Adresas': 'street',
    u'Korespondecija. Pašto kodas': 'zip',
    u'Banko sąskaita': 'bank_account',
    u'Banko pavadinimas': 'bank_name',
    u'Banko kodas': 'bank_code',
    u'Banko "SWIFT" kodas': 'bank_bic',
    u'Klientas': 'customer',
    u'Tiekėjas': 'supplier',
}


def import_partners(self, import_file):
    env = self.sudo().env
    partner_obj = env['res.partner']
    bank_bank_obj = env['res.bank']
    bank_obj = env['res.partner.bank']
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Partneriai')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    errors_general = []
    errors_system = []
    for i, row in enumerate(iter_rows):
        try:
            if not header:
                header = get_all_values(row)
                header_mapped = get_mapped(header, partner_mapping)
                continue
            values = get_all_values(row)
            if len(set(values)) == 1:
                break
            record = ImportRecord(values, header_mapped)
            name = convert_to_string(record.name)
            kodas = convert_to_string(record.kodas)
            vat = convert_to_string(record.vat)
            partner_id = False
            # check if we can update
            if kodas:
                partner_id = partner_obj.search([('kodas', '=', kodas)], limit=1)
            if not partner_id and vat:
                partner_id = partner_obj.search([('vat', '=', vat)], limit=1)
            # Per daug netikslu - neieškom
            # if not partner_id and name:
            #     partner_id = partner_obj.search([('name', 'like', name)], limit=1)
            is_company = isinstance(record.is_company, basestring) and record.is_company.lower() == 'taip'
            if not name and not kodas:
                continue
            if not name or not record.is_company:
                raise exceptions.UserError(
                    _('Nepavyko importuoti failo. Ne visi privalomi laukeliai supildyti %s eilutėje.' % (i + 1)))
            partner_vals = record.get_dict(['kodas', 'vat', 'phone', 'email', 'comment', 'city', 'street', 'zip'])
            partner_vals['name'] = name
            partner_vals['is_company'] = is_company
            partner_vals['kodas'] = kodas
            partner_vals['vat'] = vat
            partner_vals['customer'] = not record.customer or not isinstance(record.customer, basestring) or \
                                       record.customer.lower() != 'ne'
            partner_vals['supplier'] = not record.supplier or not isinstance(record.supplier, basestring) or \
                                       record.supplier.lower() != 'ne'
            payment_term = record.payment_term
            payment_term_supplier = record.payment_term_supplier
            if payment_term:
                try:
                    payment_term = int(payment_term)
                except:
                    raise exceptions.UserError(
                        _('Neteisingai nurodytas mokėjimo terminas klientui (%s) %s eilutėje.') % (payment_term, i + 1))
                payment_term_line_id = env['account.payment.term.line'].search(
                    [('value', '=', 'balance'), ('days', '=', payment_term)], limit=1)
                if payment_term_line_id:
                    partner_vals['property_payment_term_id'] = payment_term_line_id.payment_id.id
                else:
                    payment_term_id = env['account.payment.term'].create({
                        'name': str(payment_term) + ' d.',
                        'line_ids': [(0, 0, {
                            'value': 'balance',
                            'days': payment_term,
                            'option': 'day_after_invoice_date',
                        })]
                    })
                    partner_vals['property_payment_term_id'] = payment_term_id.id
            if payment_term_supplier:
                try:
                    payment_term_supplier = int(payment_term_supplier)
                except:
                    raise exceptions.UserError(
                        _('Neteisingai nurodytas mokėjimo terminas tiekėjui (%s) %s eilutėje.') % (payment_term, i + 1))
                payment_term_line_id = env['account.payment.term.line'].search(
                    [('value', '=', 'balance'), ('days', '=', payment_term_supplier)], limit=1)
                if payment_term_line_id:
                    partner_vals['property_supplier_payment_term_id'] = payment_term_line_id.payment_id.id
                else:
                    payment_term_id = env['account.payment.term'].create({
                        'name': str(payment_term) + ' d.',
                        'line_ids': [(0, 0, {
                            'value': 'balance',
                            'days': payment_term_supplier,
                            'option': 'day_after_invoice_date',
                        })]
                    })
                    partner_vals['property_supplier_payment_term_id'] = payment_term_id.id
            if partner_vals['email']:
                for partner_email in partner_vals['email'].split(';'):
                    if not partner_email:  # if field ends with ;
                        continue
                    if not validate_email(partner_email, verify=False):
                        raise exceptions.UserError(
                            _('Nepavyko importuoti failo. Neteisingai nurodytas el. pašto adresas %s eilutėje.' % (
                                    i + 1)))
            country = convert_to_string(record.country_id)
            if country and isinstance(country, tuple([str, unicode])) and len(country) > 2:
                raise exceptions.UserError(_('Neteisingai nurodytas šalies kodas (%s) %s eilutėje.') % (country, i + 1))
            if country:
                country_id = env['res.country'].search([('code', '=', country)], limit=1)
                if not country_id:
                    raise exceptions.UserError(
                        _('Neteisingai nurodytas šalies kodas (%s) %s eilutėje.') % (country, i + 1))
                partner_vals['country_id'] = country_id.id
                if country_id.code == 'LT' and is_company and not record.kodas:
                    raise exceptions.UserError(
                        _('Lietuviškoms bendrovėms privaloma nurodyti įmonės kodą (%s eilutė).') % (i + 1))
            if not partner_id:
                partner_id = partner_obj.create(partner_vals)
            else:
                partner_id.write(partner_vals)
            bank_bic = convert_to_string(record.bank_bic)
            bank_name = convert_to_string(record.bank_name)
            bank_code = convert_to_string(record.bank_code)
            bank_account = sanitize_account_number(record.bank_account)
            bank_id = False
            if bank_account and len(bank_account) >= 9:
                bank_code = bank_account[4:9]
                bank_id = bank_bank_obj.search([('kodas', '=', bank_code)], limit=1)
            if not bank_id and bank_bic and bank_bank_obj.search([('bic', '=', bank_bic)]):
                bank_id = bank_bank_obj.search([('bic', '=', bank_bic)], limit=1)
            if not bank_id and bank_code and bank_bank_obj.search([('kodas', '=', bank_code)]):
                bank_id = bank_bank_obj.search([('kodas', '=', bank_code)], limit=1)
            if not bank_id and bank_name and bank_bank_obj.search([('name', 'like', bank_name)]):
                bank_id = bank_bank_obj.search([('name', 'like', bank_name)], limit=1)
            if not bank_id and bank_name:
                bank_vals = {
                    'name': bank_name,
                }
                if bank_bic:
                    bank_vals['bic'] = bank_bic
                if bank_code:
                    bank_vals['kodas'] = bank_code
                bank_id = bank_bank_obj.create(bank_vals)
            if bank_account:
                bank_account_id = bank_obj.search([('acc_number', '=', bank_account)], limit=1)
                if not bank_account_id:
                    partner_bank_id = bank_obj.create({
                        'bank_id': bank_id.id if bank_id else False,
                        'partner_id': partner_id.id,
                        'acc_number': bank_account,
                    })
                    partner_bank_id.onchange_acc_number()
                else:
                    bank_account_id.update({
                        'acc_number': bank_account,
                    })
        except exceptions.UserError as exc:
            errors_general.append(_('%s eilutė %s') % (exc.name, i + 1))
        except exceptions.ValidationError as exc:
            errors_general.append(_('%s eilutė %s') % (exc.name, i + 1))
        except Exception as e:
            errors_system.append(_('%s eilutė %s') % (e, i + 1))
            env.cr.rollback()
    if errors_general:
        raise exceptions.UserError('\n'.join(errors_general))
    if errors_system:
        raise RoboImportError('\n'.join(errors_system))


customer_invoices_mapping = {
    u'Pard. Data': 'date_invoice',
    u'Numeris': 'number',
    u'Klientas': 'partner_id',
    u'Valiuta': 'currency_id',
    u'Suma EUR su PVM': 'amount_eur',
    # u'Kaina su PVM EUR': 'amount_eur_vat',
    u'Apmokėti iki': 'date_due',
    u'Suma valiuta su PVM': 'amount_currency',
    # u'Kaina su PVM': 'amount_currency_vat',
}


def import_customer_invoices(self, import_file):
    env = self.sudo().env
    partner_obj = env['res.partner']
    product_obj = env['product.product']
    currency_obj = env['res.currency']
    tax_obj = env['account.tax']
    invoice_obj = env['account.invoice']
    account_move_obj = env['account.move']
    account_obj = env['account.account']
    journal_obj = env['account.journal']
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Pardavimai')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    off_balance_id = account_obj.search([('code', '=', '999999')])
    base_country = env['res.country'].sudo().search([('code', '=', 'LT')], limit=1)
    if not off_balance_id:
        off_balance_id = account_obj.create({
            'name': _('Užbalansė'),
            'code': '999999',
            'user_type_id': env.ref('l10n_lt.account_type_uzbalanse').id,
        })
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, customer_invoices_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        invoice_vals = record.get_dict(['date_invoice', 'number', 'date_due'])
        partner = convert_to_string(record.partner_id)
        currency = convert_to_string(record.currency_id)
        try:
            amount_eur = float(record.amount_eur)
            amount_currency = float(record.amount_currency or 0.0)
        except ValueError:
            raise exceptions.UserError(_('Neteisinga skaitinė reikšmė %s eilutėje.') % (i + 1))

        # amount_eur_vat = record.amount_eur_vat
        if currency and amount_currency:
            if len(currency) != 3:
                raise exceptions.UserError(_('Valiutos kodas privalo būti iš 3 raidžių %s eilutėje.') % (i + 1))
            # if amount_currency < 0.0:
            #     raise exceptions.UserError(_('Suma valiuta turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
            currency_id = currency_obj.search([('name', '=', currency)], limit=1)
        else:
            currency_id = False
        inv_date = invoice_vals['date_invoice']
        if not inv_date:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        elif isinstance(inv_date, datetime):
            inv_date = inv_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
        if currency_id and tools.float_compare(amount_eur, 0, precision_digits=2) == 0:
            company_curr = self.company_id.currency_id
            if not self.env['res.currency.rate'].search([('name', '>=', inv_date + ' 00:00:00'),
                                                         ('name', '<=', inv_date + ' 23:59:59'),
                                                         ('currency_id', '=', currency_id.id)]):
                raise exceptions.UserError(_('Nėra valiutos %s kurso datai %s') % (currency_id.name, inv_date))
            amount_eur = currency_id.with_context(date=inv_date).compute(amount_currency, company_curr)
        # amount_currency_vat = record.amount_currency_vat
        if not invoice_vals['number'] or not partner or not amount_eur:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        # if amount_eur < 0.0:
        #     raise exceptions.UserError(_('Suma turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
        partner_id = partner_obj.search([('kodas', '=', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.search([('vat', '=', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.search([('name', 'like', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.create({
                'name': partner,
                'is_company': True,
                'country_id': base_country.id,
            })
        if 'date_due' in invoice_vals and not invoice_vals['date_due']:
            invoice_vals['date_due'] = invoice_vals['date_invoice']
        invoice_vals['partner_id'] = partner_id.id
        if amount_eur < 0.0:
            invoice_vals['type'] = 'out_refund'
            amount_eur = abs(amount_eur)
            amount_currency = abs(amount_currency)
        else:
            invoice_vals['type'] = 'out_invoice'
        invoice_vals['operacijos_data'] = invoice_vals['date_invoice']
        invoice_vals['imported'] = True
        invoice_vals['partner_country_id'] = False
        invoice_vals['move_name'] = invoice_vals['number']
        invoice_vals['journal_id'] = journal_obj.search([('type', '=', 'sale')], limit=1).id
        product_id = product_obj.search([('default_code', '=', 'importas')], limit=1)
        if not product_id:
            product_id = product_obj.create({
                'name': _('Importas'),
                'default_code': 'importas',
                'type': 'service',
                'acc_product_type': 'service',
                'categ_id': self.env.ref('l10n_lt.product_category_23').id,
            })
        # Limit = 1 because 2 Ne PVM exists (included price and not), in this case it does not matter
        # which one we choose because all of the taxes will be of the same inclusion type
        tax_id = tax_obj.search(
            [('code', '=', 'Ne PVM'), ('type_tax_use', '=', 'sale'), ('price_include', '=', False)], limit=1)
        if not tax_id:
            tax_id = tax_obj.create({
                'name': _('NE PVM objektas'),
                'code': 'Ne PVM',
                'description': '0%',
                'amount': 0.0,
            })
        invoice_line_vals = {
            'name': product_id.name,
            'product_id': product_id.id,
            'account_id': off_balance_id.id,
            'quantity': 1.0,
            'invoice_line_tax_ids': [(4, tax_id.id)],
        }
        if currency and amount_currency and currency_id:
            if len(currency) != 3:
                raise exceptions.UserError(_('Valiutos kodas privalo būti iš 3 raidžių %s eilutėje.') % (i + 1))
            # if amount_currency < 0.0:
            #     raise exceptions.UserError(_('Suma valiuta turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
            invoice_vals['currency_id'] = currency_id.id
            invoice_line_vals['price_unit'] = amount_currency
        else:
            invoice_line_vals['price_unit'] = amount_eur
        invoice_vals['invoice_line_ids'] = [(0, 0, invoice_line_vals)]
        inv = invoice_obj.with_context(journal_type='sale').create(invoice_vals)
        # Create special accounting entry
        line = []
        name = inv.name or '/'
        line_vals = {
            'type': 'dest',
            'name': name,
            'price': amount_eur,
            'account_id': inv.account_id.id,
            'date_maturity': inv.date_due,
            'invoice_id': inv.id,
            'partner_id': partner_id.id,
        }
        if invoice_vals['type'] == 'out_invoice':
            sign = 1.0
            line_vals['debit'] = amount_eur
            line_vals['credit'] = 0.0
        else:
            sign = -1.0
            line_vals['debit'] = 0.0
            line_vals['credit'] = amount_eur
        if 'currency_id' in invoice_vals and amount_currency:
            line_vals['currency_id'] = invoice_vals['currency_id']
            line_vals['amount_currency'] = sign * amount_currency
        line.append((0, 0, line_vals))
        line_vals2 = {
            'type': 'dest',
            'name': name,
            'price': amount_eur,
            'debit': 0.0,
            'credit': amount_eur,
            'account_id': off_balance_id.id,
            'date_maturity': inv.date_due,
            'invoice_id': inv.id
        }
        if invoice_vals['type'] == 'out_invoice':
            sign = -1.0
            line_vals2['debit'] = 0.0
            line_vals2['credit'] = amount_eur
        else:
            sign = 1.0
            line_vals2['debit'] = amount_eur
            line_vals2['credit'] = 0.0
        if 'currency_id' in invoice_vals and amount_currency:
            line_vals2['currency_id'] = invoice_vals['currency_id']
            line_vals2['amount_currency'] = sign * amount_currency
        line.append((0, 0, line_vals2))
        ctx = dict(self._context)
        move_vals = {
            'ref': inv.reference,
            'line_ids': line,
            'journal_id': inv.journal_id.id,
            'date': inv.date_invoice,
            'narration': inv.comment,
        }
        ctx['company_id'] = inv.company_id.id
        ctx['invoice'] = inv
        ctx_nolang = ctx.copy()
        ctx_nolang.pop('lang', None)
        move = account_move_obj.with_context(ctx_nolang).create(move_vals)
        # Pass invoice in context in method post: used if you want to get the same
        # account move reference when creating the same invoice after a cancelled one:
        move.post()
        # make the invoice point to that move
        vals = {
            'move_id': move.id,
            'state': 'open',
        }
        inv.with_context(ctx).write(vals)


supplier_invoices_mapping = {
    u'Pirk. Data': 'date_invoice',
    u'Numeris': 'reference',
    u'Tiekėjas': 'partner_id',
    u'Valiuta': 'currency_id',
    u'Suma EUR su PVM': 'amount_eur',
    # u'Kaina su PVM EUR': 'amount_eur_vat',
    u'Apmokėti iki': 'date_due',
    u'Suma valiuta su PVM': 'amount_currency',
    # u'Kaina su PVM': 'amount_currency_vat',
}


def import_supplier_invoices(self, import_file):
    env = self.sudo().env
    partner_obj = env['res.partner']
    product_obj = env['product.product']
    currency_obj = env['res.currency']
    tax_obj = env['account.tax']
    invoice_obj = env['account.invoice']
    account_move_obj = env['account.move']
    account_obj = env['account.account']
    journal_obj = env['account.journal']
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Pirkimai')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    off_balance_id = account_obj.search([('code', '=', '999999')])
    base_country = env['res.country'].sudo().search([('code', '=', 'LT')], limit=1)
    if not off_balance_id:
        off_balance_id = account_obj.create({
            'name': _('Užbalansė'),
            'code': '999999',
            'user_type_id': env.ref('l10n_lt.account_type_uzbalanse').id,
        })
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, supplier_invoices_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        invoice_vals = record.get_dict(['date_invoice', 'reference', 'date_due'])
        partner = convert_to_string(record.partner_id)
        currency = convert_to_string(record.currency_id)
        try:
            amount_eur = float(record.amount_eur)
            amount_currency = float(record.amount_currency or 0.0)
        except ValueError:
            raise exceptions.UserError(_('Neteisinga skaitinė reikšmė %s eilutėje.') % (i + 1))
        # amount_currency_vat = record.amount_currency_vat
        if currency and amount_currency:
            if len(currency) != 3:
                raise exceptions.UserError(_('Valiutos kodas privalo būti iš 3 raidžių %s eilutėje.') % (i + 1))
            # if amount_currency < 0.0:
            #     raise exceptions.UserError(_('Suma valiuta turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
            currency_id = currency_obj.search([('name', '=', currency)], limit=1)
            if not currency_id:
                raise exceptions.UserError(_('Nerasta valiuta %s %s eilutėje') % (currency, i + 1))
        else:
            currency_id = False
        inv_date = invoice_vals['date_invoice']
        if not inv_date:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        elif isinstance(inv_date, datetime):
            inv_date = inv_date.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
        if not inv_date:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        if currency_id and tools.float_compare(amount_eur, 0, precision_digits=2) == 0:
            company_curr = self.company_id.currency_id
            if not self.env['res.currency.rate'].search([('name', '>=', inv_date + ' 00:00:00'),
                                                         ('name', '<=', inv_date + ' 23:59:59'),
                                                         ('currency_id', '=', currency_id.id)]):
                raise exceptions.UserError(_('Nėra valiutos %s kurso datai %s') % (currency_id.name, inv_date))
            amount_eur = currency_id.with_context(date=inv_date).compute(amount_currency, company_curr)
        if not invoice_vals['date_invoice'] or not invoice_vals['reference'] or not partner or not amount_eur:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        # if amount_eur < 0.0:
        #     raise exceptions.UserError(_('Suma turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
        partner_id = partner_obj.search([('kodas', '=', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.search([('vat', '=', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.search([('name', 'like', partner)], limit=1)
        if not partner_id:
            partner_id = partner_obj.create({
                'name': partner,
                'is_company': True,
                'country_id': base_country.id,
            })
        if 'date_due' in invoice_vals and not invoice_vals['date_due']:
            invoice_vals['date_due'] = invoice_vals['date_invoice']
        invoice_vals['partner_id'] = partner_id.id
        if amount_eur < 0.0:
            invoice_vals['type'] = 'in_refund'
            amount_eur = abs(amount_eur)
            amount_currency = abs(amount_currency)
        else:
            invoice_vals['type'] = 'in_invoice'
        invoice_vals['operacijos_data'] = invoice_vals['date_invoice']
        invoice_vals['imported'] = True
        invoice_vals['journal_id'] = journal_obj.search([('type', '=', 'purchase')], limit=1).id
        product_id = product_obj.search([('default_code', '=', 'importas')], limit=1)
        if not product_id:
            product_id = product_obj.create({
                'name': _('Importas'),
                'default_code': 'importas',
                'type': 'service',
                'acc_product_type': 'service',
                'categ_id': self.env.ref('l10n_lt.product_category_23').id,
            })
        tax_id = tax_obj.search(
            [('code', '=', 'Ne PVM'), ('type_tax_use', '=', 'purchase'), ('price_include', '=', False)], limit=1)
        if not tax_id:
            tax_id = tax_obj.create({
                'name': _('NE PVM objektas'),
                'code': 'Ne PVM',
                'description': '0%',
                'amount': 0.0,
                'type_tax_use': 'purchase',
            })
        invoice_line_vals = {
            'name': product_id.name,
            'product_id': product_id.id,
            'account_id': off_balance_id.id,
            'quantity': 1.0,
            'invoice_line_tax_ids': [(4, tax_id.id)],
        }
        if currency and amount_currency:
            if len(currency) != 3:
                raise exceptions.UserError(_('Valiutos kodas privalo būti iš 3 raidžių %s eilutėje.') % (i + 1))
            # if amount_currency < 0.0:
            #     raise exceptions.UserError(_('Suma valiuta turi būti teigiamas skaičius %s eilutėje.') % (i + 1))
            currency_id = currency_obj.search([('name', '=', currency)], limit=1)
            invoice_vals['currency_id'] = currency_id.id
            invoice_line_vals['price_unit'] = amount_currency
        else:
            invoice_line_vals['price_unit'] = amount_eur
        invoice_vals['invoice_line_ids'] = [(0, 0, invoice_line_vals)]
        inv = invoice_obj.create(invoice_vals)
        # Create special accounting entry
        line = []
        name = inv.name or '/'
        line_vals = {
            'type': 'dest',
            'name': name,
            'price': amount_eur,
            'account_id': inv.account_id.id,
            'date_maturity': inv.date_due,
            'invoice_id': inv.id,
            'partner_id': partner_id.id,
        }
        if invoice_vals['type'] == 'in_invoice':
            sign = -1.0
            line_vals['debit'] = 0.0
            line_vals['credit'] = amount_eur
        else:
            sign = 1.0
            line_vals['debit'] = amount_eur
            line_vals['credit'] = 0.0
        if 'currency_id' in invoice_vals and amount_currency:
            line_vals['currency_id'] = invoice_vals['currency_id']
            line_vals['amount_currency'] = sign * amount_currency
        line.append((0, 0, line_vals))
        line_vals2 = {
            'type': 'dest',
            'name': name,
            'price': amount_eur,
            'credit': 0.0,
            'debit': amount_eur,
            'account_id': off_balance_id.id,
            'date_maturity': inv.date_due,
            'invoice_id': inv.id
        }
        if invoice_vals['type'] == 'in_invoice':
            sign = 1.0
            line_vals2['debit'] = amount_eur
            line_vals2['credit'] = 0.0
        else:
            sign = -1.0
            line_vals2['debit'] = 0.0
            line_vals2['credit'] = amount_eur
        if 'currency_id' in invoice_vals and amount_currency:
            line_vals2['currency_id'] = invoice_vals['currency_id']
            line_vals2['amount_currency'] = sign * amount_currency
        line.append((0, 0, line_vals2))
        ctx = dict(self._context)
        move_vals = {
            'ref': inv.reference,
            'line_ids': line,
            'journal_id': inv.journal_id.id,
            'date': inv.date_invoice,
            'narration': inv.comment,
        }
        ctx['company_id'] = inv.company_id.id
        ctx['invoice'] = inv
        ctx_nolang = ctx.copy()
        ctx_nolang.pop('lang', None)
        move = account_move_obj.with_context(ctx_nolang).create(move_vals)
        # Pass invoice in context in method post: used if you want to get the same
        # account move reference when creating the same invoice after a cancelled one:
        move.post()
        # make the invoice point to that move
        vals = {
            'move_id': move.id,
            'state': 'open',
        }
        inv.with_context(ctx).write(vals)


product_mapping = {
    u'Pavadinimas': 'name',
    u'Kodas': 'default_code',
    u'Barkodas': 'barcode',
    u'Ar Paslauga?': 'service',
    u'Pirkimo kaina be PVM EUR': 'standard_price',
    u'Pardavimo kaina be PVM EUR': 'list_price',
    u'Prekės grupė': 'category',
    u'Pardavimo kor. sąskaita': 'income_account',
    u'Sanaudų kor. sąskaita': 'expense_account',
    u'Prekių vertės kor. sąskaita': 'stock_account',
    u'Prekės svoris, kg': 'kg',
    u'Prekės tūris, m3': 'volume',
    u'Tara(popierius), g': 'tara_popierius',
    u'Tara(plastm.), g': 'tara_plastmase',
    u'Tara(kita), g': 'tara_kita',
    u'Tara(PET), g': 'tara_pet',
    u'Tara(metalas), g': 'tara_metalas',
    u'Tara(medinė), g': 'tara_medine',
    u'Tara(komb.), g': 'tara_kombinuota',
    u'Tara(stiklas), g': 'tara_stiklas',
    u'Intrastat kodas': 'intrastat',
    u'Intrastat pavadinimas': 'intrastat_name',
    u'Produkto intrastat aprašymas': 'intrastat_description',
    u'Kilmės šalis': 'kilmes_salis',
}


def import_products(self, import_file):
    env = self.sudo().env
    product_obj = env['product.product']
    category_obj = env['product.category']
    account_obj = env['account.account']
    intrastat_obj = env['report.intrastat.code']
    country_obj = env['res.country']

    robo_package_installed = env['ir.module.module'].search([('name', '=', 'robo_package'),
                                                             ('state', 'in', ['installed', 'to upgrade'])], count=True)

    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Prekės (paslaugos)')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, product_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        product_vals = record.get_dict(['name', 'default_code'])
        if not product_vals['name'] and not product_vals['default_code']:
            raise exceptions.UserError(_('Neužpildyti visi privalomi laukai %s eilutėje.') % (i + 1))
        product_id = product_obj
        product_name = convert_to_string(product_vals['name'])
        product_code = convert_to_string(product_vals['default_code'])
        if product_code:
            product_id = product_obj.search([('default_code', '=', product_code)])
        if product_name:
            if len(product_id) > 1 or (not product_id and not product_code) or \
                    (product_id and product_id.with_context(lang='lt_LT').name != product_name and
                     product_id.with_context(lang='en_US').name != product_name):
                product_id = product_obj.with_context(lang='lt_LT').search([('name', '=', product_name)])
                if not product_id:
                    product_id = product_obj.with_context(lang='en_US').search([('name', '=', product_name)])
        if len(product_id) > 1:
            raise exceptions.UserError(_('Rasti keli produkto %s atitikmenys.') % product_code or product_name)
        # search by name is too risky
        # if not product_id and product_vals['name']:
        #     product_id = product_obj.search([('name', 'like', product_vals['name'])], limit=1)
        category_name = convert_to_string(record.category)
        income_account = convert_to_string(record.income_account)
        expense_account = convert_to_string(record.expense_account)
        stock_account = convert_to_string(record.stock_account)
        service = record.service
        update_product_type = service is not None
        is_service = service in ['Taip', 'TAIP'] if update_product_type else \
            (product_id.type == 'service' if product_id else True)
        income_account_id = account_obj.search([('code', '=', income_account)], limit=1)
        expense_account_id = account_obj.search([('code', '=', expense_account)], limit=1)
        stock_account_id = account_obj.search([('code', '=', stock_account)], limit=1)
        category_id = False
        if category_name and category_name != 'False':
            category_id = category_obj.search([('name', '=', category_name)], limit=1)
            if not category_id:
                category_id = category_obj.search([('full_name', '=', category_name)], limit=1)
            if not category_id:
                parent_categ = env.ref('l10n_lt.product_category_1') if not is_service else env.ref(
                    'l10n_lt.product_category_2')
                category_vals = {
                    'name': category_name,
                    'parent_id': parent_categ.id,
                    'property_stock_valuation_account_id': (
                                                               stock_account_id.id if stock_account_id and not is_service else False) or parent_categ.property_stock_valuation_account_id.id,
                    'property_account_income_categ_id': (
                                                            income_account_id.id if income_account_id else False) or parent_categ.property_account_income_categ_id.id,
                }
                if not is_service and expense_account_id:
                    category_vals[
                        'property_stock_account_output_categ_id'] = expense_account_id.id or parent_categ.property_stock_account_output_categ_id.id
                elif is_service and expense_account_id:
                    category_vals[
                        'property_account_expense_categ_id'] = expense_account_id.id or parent_categ.property_account_expense_categ_id.id
                if not is_service:
                    category_vals['property_cost_method'] = 'real'
                    category_vals['property_valuation'] = 'real_time'
                category_id = category_obj.create(category_vals)
                category_id._onchange_parent_id()
                category_id.onchange_category_type()
        # if not category_id:
        #     category_id = env.ref('l10n_lt.product_category_2')
        # if category_id:
        #     product_vals['categ_id'] = category_id.id
        if income_account_id:
            product_vals['property_account_income_id'] = income_account_id.id
        if is_service and expense_account_id:
            product_vals['property_account_expense_id'] = expense_account_id.id
        if is_service:
            product_vals['type'] = 'service'
            product_vals['acc_product_type'] = 'service'
        else:
            product_vals['type'] = 'product'
            product_vals['acc_product_type'] = 'product'

        # Additional product info
        kg = record.kg
        if check_number(kg):
            product_vals['weight'] = kg
        volume = record.volume
        if check_number(volume):
            product_vals['volume'] = volume
        list_price = record.list_price
        if check_number(list_price):
            product_vals['list_price'] = list_price
        standard_price = record.standard_price
        if check_number(standard_price):
            product_vals['standard_price'] = standard_price
        barcode = convert_to_string(record.barcode)
        if barcode and barcode != 'False':
            product_vals['barcode'] = barcode

        intrastat_code = convert_to_string(record.intrastat)
        intrastat_name = convert_to_string(record.intrastat_name)
        intrastat_description = convert_to_string(record.intrastat_description)
        if intrastat_code and intrastat_code != 'False':
            intrastat_id = intrastat_obj.search([('name', '=', intrastat_code)], limit=1)
            if intrastat_id:
                product_vals['intrastat_id'] = intrastat_id.id
            else:
                intrastat_vals = {'name': intrastat_code}
                if intrastat_name and intrastat_name != 'False':
                    intrastat_vals['description'] = intrastat_name
                intrastat_id = intrastat_obj.create(intrastat_vals)
                product_vals['intrastat_id'] = intrastat_id.id
            if intrastat_description and intrastat_description != 'False':
                product_vals['intrastat_description'] = intrastat_description
        kilmes_salis = convert_to_string(record.kilmes_salis).upper()
        if kilmes_salis:
            country_id = country_obj.search([('code', '=', kilmes_salis)], limit=1)
            if country_id:
                product_vals['kilmes_salis'] = country_id.id

        if product_id:
            if category_id:
                product_vals['categ_id'] = category_id.id
            product_id.write(product_vals)
        else:
            if not category_id:
                if not is_service:
                    category_id = env.ref('l10n_lt.product_category_1')
                else:
                    category_id = env.ref('l10n_lt.product_category_2')
            product_vals['categ_id'] = category_id.id
            product_id = product_obj.with_context(skip_constraints=True).create(product_vals)

        if robo_package_installed and not product_id.product_tmpl_id.product_package_default_ids:
            package_obj = env['product.package']
            package_default_obj = env['product.package.default']
            tara_popierius = record.tara_popierius
            if check_number(tara_popierius):
                # P3:DivOK
                tara_popierius /= 1000.0
                package_id = package_obj.create({
                    'name': 'Popierius',
                    'package_category': 'pirmine',
                    'material_type': 'popierius',
                    'weight': tara_popierius,
                    'use_type': 'vienkartine',
                    'recycling_type': 'perdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_plastmase = record.tara_plastmase
            if check_number(tara_plastmase):
                # P3:DivOK
                tara_plastmase /= 1000.0
                package_id = package_obj.create({
                    'name': 'Plastmasė',
                    'package_category': 'pirmine',
                    'material_type': 'plastikas',
                    'weight': tara_plastmase,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_kita = record.tara_kita
            if check_number(tara_kita):
                # P3:DivOK
                tara_kita /= 1000.0
                package_id = package_obj.create({
                    'name': 'Kita',
                    'package_category': 'pirmine',
                    'material_type': 'kita',
                    'weight': tara_kita,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_pet = record.tara_pet
            if check_number(tara_pet):
                # P3:DivOK
                tara_pet /= 1000.0
                package_id = package_obj.create({
                    'name': 'PET',
                    'package_category': 'pirmine',
                    'material_type': 'pet',
                    'weight': tara_pet,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_metalas = record.tara_metalas
            if check_number(tara_metalas):
                # P3:DivOK
                tara_metalas /= 1000.0
                package_id = package_obj.create({
                    'name': 'Metalas',
                    'package_category': 'pirmine',
                    'material_type': 'metalas',
                    'weight': tara_metalas,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_medine = record.tara_medine
            if check_number(tara_medine):
                # P3:DivOK
                tara_medine /= 1000.0
                package_id = package_obj.create({
                    'name': 'Medis',
                    'package_category': 'pirmine',
                    'material_type': 'medis',
                    'weight': tara_medine,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_kombinuota = record.tara_kombinuota
            if check_number(tara_kombinuota):
                # P3:DivOK
                tara_kombinuota /= 1000.0
                package_id = package_obj.create({
                    'name': 'Kombinuota',
                    'package_category': 'pirmine',
                    'material_type': 'kombinuota',
                    'weight': tara_kombinuota,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })

            tara_stiklas = record.tara_stiklas
            if check_number(tara_stiklas):
                # P3:DivOK
                tara_stiklas /= 1000.0
                package_id = package_obj.create({
                    'name': 'Stiklas',
                    'package_category': 'pirmine',
                    'material_type': 'stiklas',
                    'weight': tara_stiklas,
                    'use_type': 'vienkartine',
                    'recycling_type': 'neperdirbama',
                })
                package_default_obj.create({
                    'package_id': package_id.id,
                    'product_tmpl_id': product_id.product_tmpl_id.id,
                    'qty_in_pack': 1.0,
                })


financials_mapping = {
    u'Kodas': 'code',
    u'Pavadinimas': 'name',
    u'D': 'debit',
    u'K': 'credit',
}


def import_financials(self, import_file):
    env = self.sudo().env
    account_obj = env['account.account']
    account_move_obj = env['account.move']
    off_books_account_id = self.env['account.account'].search([('code', '=', '999999')])
    if not off_books_account_id:
        raise exceptions.UserError(_('Nepavyko surasti atitinkamos apskaitos informacijos.'))
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name=u'Likučiai')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    journal_id = self.env['account.journal'].search([('code', '=', 'START')], limit=1)
    if not journal_id:
        journal_id = self.env['account.journal'].create({
            'name': 'Pradiniai likučiai',
            'code': 'START',
            'type': 'general',
            'update_posted': True,
        })
    lines = []
    move = {
        'ref': u'Pradiniai buhalteriniai likučiai',
        'date': (env.user.company_id.compute_fiscalyear_dates()['date_from']).strftime(
            tools.DEFAULT_SERVER_DATE_FORMAT),
        'journal_id': journal_id.id,
        'line_ids': lines,
    }
    total_debit = 0.0
    total_credit = 0.0
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, financials_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        code = str(convert_to_string(record.code) or '')
        debit = record.debit
        credit = record.credit
        if not code or (not debit and not credit):
            continue
        if isinstance(debit, tuple([str, unicode])) and isinstance(credit, tuple([str, unicode])):
            if '=' in debit or '=' in credit:
                continue
            else:
                try:
                    debit = float(debit)
                    credit = float(credit)
                except:
                    raise exceptions.UserError(_('Neteisingas duomenų formatas %s eilutėje.') % (i + 1))
        if (debit and not isinstance(debit, tuple([int, float, long]))) or (
                credit and not isinstance(credit, tuple([int, float, long]))):
            raise exceptions.UserError(_('Neteisingas duomenų formatas %s eilutėje.') % (i + 1))
        if not debit:
            debit = 0.0
        if not credit:
            credit = 0.0
        debit = tools.float_round(debit, precision_digits=2)
        credit = tools.float_round(credit, precision_digits=2)
        if tools.float_is_zero(debit, precision_digits=2) and tools.float_is_zero(credit, precision_digits=2):
            continue
        account_id = account_obj.search([('code', '=', code)], limit=1)
        if not account_id and not code:
            raise exceptions.UserError(_('Nepavyko rasti atitinkamos DK sąskaitos %s eilutėje.') % (i + 1))
        elif not account_id and code and record.name:
            account_vals = {
                'code': code,
                'name': convert_to_string(record.name),
                'company_id': self.env.user.company_id.id,
            }
            code_class = str(code[0])
            if code_class == '1':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_fixed_assets').id
            elif code_class == '2':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_current_assets').id
            elif code_class == '3':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_equity').id
            elif code_class == '4':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_current_liabilities').id
            elif code_class == '5':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_revenue').id
            elif code_class == '6':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_expenses').id
            else:
                account_vals['user_type_id'] = self.env.ref('l10n_lt.account_type_uzbalanse').id
            try:
                account_id = account_obj.create(account_vals)
            except:
                raise exceptions.UserError(_('Nepavyko pridėti naujos DK sąskaitos.'))
        if not account_id:
            raise exceptions.UserError(_('Nepavyko rasti atitinkamos DK sąskaitos %s eilutėje.') % (i + 1))
        if debit > 0.0:
            total_debit += debit
            lines.append((0, 0, {
                'name': convert_to_string(record.name) or account_id.name or code,
                'account_id': account_id.id,
                'debit': debit,
                'credit': 0.0,
                'journal_id': journal_id.id,
            }))
        if credit > 0.0:
            total_credit += credit
            lines.append((0, 0, {
                'name': convert_to_string(record.name) or account_id.name or code,
                'account_id': account_id.id,
                'debit': 0.0,
                'credit': credit,
                'journal_id': journal_id.id,
            }))
    if tools.float_compare(total_credit, total_debit, precision_digits=2) != 0:
        if total_debit > total_credit:
            last_credit = total_debit - total_credit
            last_debit = 0.0
        else:
            last_debit = total_credit - total_debit
            last_credit = 0.0
        lines.append((0, 0, {
            'name': u'Balansuojantis įrašas',
            'account_id': off_books_account_id.id,
            'debit': last_debit,
            'credit': last_credit,
            'journal_id': journal_id.id,
        }))
    if lines:
        account_move_obj.create(move)


employees_mapping = {
    u'Darbuotojo info. Vardas Pavardė': 'employee_name',
    u'Darbuotojo info. Lytis': 'gender',
    u'Darbuotojo info. Asmens kodas': 'identification_no',
    u'Darbuotojo info. Soc. Draudimo serija': 'sodra_serija',
    u'Darbuotojo info. Soc. Draudimo numeris': 'sodra_numeris',
    u'Darbuotojo info. SODRA Papildomai': 'sodra_papildomai',
    u'Darbuotojo info. Papildomos SODRA tipas': 'sodra_papildomai_type',
    u'Darbuotojo info. Sąskaita': 'bank_account',
    u'Darbuotojo info. Pilietybė': 'country',
    u'Darbuotojo info. Pareigos': 'job_name',
    u'Darbuotojo info. Skyrius': 'department_name',
    u'Darbuotojo info. Šeimyninė padėtis': 'seimynine_padetis',
    u'Darbuotojo info. Rezidentas': 'is_resident',
    u'Darbuotojo info. Adresas': 'address',
    u'Darbuotojo info. Miestas': 'city',
    u'Darbuotojo info. Tel. Nr.': 'phone',
    u'Darbuotojo info. El. paštas': 'email',
    u'Darbuotojo info. Savivaldybė': 'savivaldybe',
    u'Paskutinio mėn. bruto VDU išskaičiavimui': 'vdu_1',
    u'Paskutinį mėnesį dirbtų dienų skaičius': 'dd_1',
    u'Paskutinį mėnesį dirbtų valandų skaičius': 'vv_1',
    u'Priešpaskutinio mėn. bruto VDU išskaičiavimui': 'vdu_2',
    u'Priešpaskutinį mėnesį dirbtų dienų skaičius': 'dd_2',
    u'Priešpaskutinį mėnesį dirbtų valandų skaičius': 'vv_2',
    u'Priešpriešpaskutinio mėn. bruto VDU išskaičiavimui': 'vdu_3',
    u'Priešpriešpaskutinį mėnesį dirbtų dienų skaičius': 'dd_3',
    u'Priešpriešpaskutinį mėnesį dirbtų valandų skaičius': 'vv_3',
    u'Darbuotojo sutartis. Sutarties numeris': 'contract_name',
    u'Darbuotojo sutartis. Dirbo nuo': 'contract_date_start',
    u'Darbuotojo sutartis. Atlyginimo tipas': 'contract_wage_type',
    u'Darbuotojo sutartis. Etatas': 'contract_etatas',
    u'Darbuotojo sutartis. Darbo valandos per mėn.': 'contract_monthly_hours',
    u'Darbuotojo sutartis. Mėn. Atlyginimas': 'contract_monthly_wage',
    u'Darbuotojo sutartis. Val. kaina': 'contract_hourly_rate',
    u'Darbuotojo sutartis. Darbo laiko grafiko tipas': 'contract_schedule_type',
    u'Darbuotojo sutartis. Trumpinti prieš šventes': 'contract_shorter_before_holidays',
    u'Darbuotojo sutartis. Dirbo iki': 'contract_date_end',
    u'Darbuotojo sutartis. Taikyti NPD': 'contract_apply_npd',
    u'Darbuotojo sutartis. NPD reikšmė': 'contract_npd',
    u'Darbuotojo sutartis. Darbo norma': 'work_norm',
    u'Darbuotojo sutartis. Vaikų skaičius': 'contract_num_children',
    u'Darbuotojo sutartis. Taikyti Papild. NPD': 'contract_pnpd',
    u'Duomenų paėmimo data': 'date_data',
    u'Priklausančių atostogų dienos. Kaupimas darbo arba kalendorinėmis dienomis': 'holidays_num_days_type',
    u'Priklausančių atostogų dienos. Priklausančių dienų per metus': 'holidays_num_days_yearly',
    u'Priklausančių atostogų dienos. Pagrindinių dienų likutis': 'holidays_num_days',
    u'Koeficientas taikomas apskaičiuojant atostoginius': 'holiday_coefficient',
    u'Duomenys apie vaikus. Vardas Pavardė': 'child_name',
    u'Duomenys apie vaikus. Gimimo data': 'child_date_born',
    u'Duomenys apie vaikus. Mokyklos baigimo data': 'child_date_school_end',
}


def import_employees(self, import_file):
    env = self.sudo().env
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name=u'Personalas')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    num_rows = sheet.max_row
    row_iter_obj = enumerate(iter_rows)
    i = 0
    try:
        while i < num_rows:
            i, row = row_iter_obj.next()
            if not header:
                header = get_all_values(row)
                header_mapped = get_mapped(header, employees_mapping)
                continue
            values = get_all_values(row)
            if len(set(values)) == 1:
                break
            record = ImportRecord(values, header_mapped)
            name = convert_to_string(record.employee_name)
            identification = convert_to_string(record.identification_no)
            employee_gender = convert_to_string(record.gender)
            department = env['hr.department']
            department_name = convert_to_string(record.department_name)
            if department_name:
                department = env['hr.department'].search([('name', '=', department_name)], limit=1)
                if not department:
                    department = env['hr.department'].create({'name': department_name})

            if identification:
                employee = env['hr.employee'].search(
                    [('identification_id', '=', identification), '|', ('active', '=', True), ('active', '=', False)],
                    limit=1)
            else:
                raise exceptions.Warning(_('Nenurodytas darbuotojo %s asmens kodas') % name)
            if not employee:
                employee = env['hr.employee'].search([('name', '=', name), ('identification_id', '=', False)], limit=1)
            savivaldybe = str()
            is_non_resident = True if record.is_resident and str(record.is_resident).lower().strip() == 'ne' else False
            if record.savivaldybe:
                if str(record.savivaldybe) in map(lambda r: r[0], SAVIVALDYBES):
                    savivaldybe = str(record.savivaldybe)
            if not is_non_resident and not savivaldybe:
                savivaldybe = '13'
            if not employee:
                if not department:
                    department = env['hr.department'].search([], limit=1)
                    if not department:
                        raise exceptions.UserError(_('Sistemoje nenurodytas nė vienas padalinys'))
                # Gender is required for newly created employees
                if not employee_gender:
                    raise exceptions.UserError(_('Nenurodyta darbuotojo %s lytis') % name)
                employee_vals = {'name': name,
                                 'identification_id': identification,
                                 'time_efficiency': 1,
                                 'department_id': department.id,
                                 'savivaldybe': savivaldybe,
                                 'is_non_resident': is_non_resident,
                                 }
                employee = self.env['hr.employee'].create(employee_vals)
                employee._onchange_identification_id()
            if employee and not employee.active:
                employee.write({
                    'active': True,
                })
            if record.savivaldybe:
                if str(record.savivaldybe) in map(lambda r: r[0], SAVIVALDYBES):
                    employee.savivaldybe = str(record.savivaldybe)
            if employee_gender:
                gender = 'female' if employee_gender.lower().strip() == 'moteris' else 'male'
                employee.gender = gender

            holiday_coefficient = record.holiday_coefficient or 1.0
            try:
                holiday_coefficient = float(holiday_coefficient)
            except:
                raise exceptions.UserError(_('Nurodytas koeficientas taikomas apskaičiuojant atostoginius nėra '
                                             'realusis skaičius'))
            if tools.float_compare(holiday_coefficient, 0.0, precision_digits=2) < 0 or \
                    tools.float_compare(holiday_coefficient, 2.0, precision_digits=2) > 0:
                # No specific reason for 2.0 being max value but the coefficient correlates to employee post and
                # there should not be cases where employee works 80 hour weeks
                raise exceptions.UserError(_('Koeficientas taikomas apskaičiuojant atostoginius privalo būti '
                                             'tarp 0 ir 2'))
            employee.holiday_coefficient = holiday_coefficient

            sodra_nr = (convert_to_string(record.sodra_serija or '') or u'') + (
                    convert_to_string(record.sodra_numeris or '') or u'')
            if sodra_nr:
                employee.sodra_id = sodra_nr

            if record.country:
                country = env['res.country'].search([('code', 'ilike', convert_to_string(record.country))], limit=1)
                if country:
                    employee.country_id = country.id
            if record.seimynine_padetis:
                if record.seimynine_padetis == 'Nevedęs/Netekėjusi':
                    employee.seimynine_padetis = env.ref('l10n_lt_payroll.nesusituokes').id
                elif record.seimynine_padetis == 'Vedęs/Ištekėjusi':
                    employee.seimynine_padetis = env.ref('l10n_lt_payroll.susituokes').id
                elif record.seimynine_padetis == 'Išsiskyręs/Išsiskyrusi':
                    employee.seimynine_padetis = env.ref('l10n_lt_payroll.issiskyres').id
                elif record.seimynine_padetis == 'Našlys/Našlė':
                    employee.seimynine_padetis = env.ref('l10n_lt_payroll.naslys').id
            if record.address:
                employee.address_home_id.street = record.address
            if record.city:
                employee.address_home_id.city = record.city
            if record.phone:
                employee.mobile_phone = record.phone
            if record.email and record.email != employee.work_email:
                employee.work_email = record.email
            max_monthly_hours = False
            men_constraint = 'etatas'

            vdu_type = 'val' if record.contract_wage_type and record.contract_wage_type.lower().strip() == 'valandinis' else 'men'

            if record.contract_name or \
                    ((record.dd_1 is not None or record.vv_1 is not None) and record.vdu_1 is not None) or \
                    (record.dd_2 is not None or record.vv_2 is not None) and record.vdu_2 is not None or \
                    (record.dd_3 is not None or record.vv_3 is not None) and record.vdu_3 is not None:
                try:
                    date_data = record.date_data.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
                except:
                    raise exceptions.Warning(
                        _('Nurodytas neteisingas duomenų nuėmimo datos formatas : %s') % record.date_data)
                date_data_dt = datetime.strptime(date_data, tools.DEFAULT_SERVER_DATE_FORMAT)

            importing_vdu = False
            if (record.dd_1 is not None or record.vv_1 is not None) and record.vdu_1 is not None:
                try:
                    dd_1 = int(round(float(record.dd_1)))
                except:
                    dd_1 = False
                try:
                    vv_1 = float(record.vv_1)
                except:
                    vv_1 = False
                try:
                    vdu_1 = float(record.vdu_1)
                except:
                    vdu_1 = False
                if not vdu_1:
                    raise exceptions.UserError(_('Neteisingai nurodytas %s darbuotojui VDU') % employee.name)
                if not dd_1 and not vv_1:
                    raise exceptions.UserError(_(
                        'Neteisingai nurodytas %s darbuotojui dirbtų dienų arba valandų skaičius') % employee.name)
                if vdu_type == 'men' and not dd_1:
                    raise exceptions.UserError(_(
                        'Pateikiant mėnesinį VDU būtina nurodyti dirbtas dienas. Darbuotojas: %s') % employee.name)
                elif vdu_type != 'men' and not vv_1:
                    raise exceptions.UserError(_(
                        'Pateikiant valandinį VDU būtina nurodyti dirbtas valandas. Darbuotojas: %s') % employee.name)
                date_vdu = date_data_dt
                year = date_vdu.year
                month = date_vdu.month
                year_id = env['years'].search([('code', '=', year)], limit=1).id
                if not year_id:
                    raise exceptions.Warning(_('Nerasti metai sistemoje: %s') % year)
                # TODO set contract on employee vdu records
                vdu_record = self.env['employee.vdu'].search([('employee_id', '=', employee.id),
                                                              ('year_id', '=', year_id),
                                                              ('month', '=', month)], limit=1)
                if not vdu_record:
                    vdu_record = self.env['employee.vdu'].create({'employee_id': employee.id,
                                                                  'year_id': year_id,
                                                                  'month': month})
                time_worked_1 = dd_1 if vdu_type == 'men' else vv_1
                # P3:DivOK -- vdu_1 is always float
                vdu_record.write({'vdu': vdu_1 / time_worked_1,
                                  'hours_worked': vv_1,
                                  'days_worked': dd_1,
                                  'amount': vdu_1,
                                  'type': vdu_type})
                importing_vdu = True
            if (record.dd_2 is not None or record.vv_2 is not None) and record.vdu_2 is not None:
                try:
                    dd_2 = int(round(float(record.dd_2)))
                except:
                    dd_2 = False
                try:
                    vv_2 = float(record.vv_2)
                except:
                    vv_2 = False
                try:
                    vdu_2 = float(record.vdu_2)
                except:
                    vdu_2 = False
                if not vdu_2:
                    raise exceptions.UserError(_('Neteisingai nurodytas %s darbuotojui VDU') % employee.name)
                if not dd_2 and not vv_2:
                    raise exceptions.UserError(_(
                        'Neteisingai nurodytas %s darbuotojui dirbtų dienų arba valandų skaičius') % employee.name)
                if vdu_type == 'men' and not dd_2:
                    raise exceptions.UserError(_(
                        'Pateikiant mėnesinį VDU būtina nurodyti dirbtas dienas. Darbuotojas: %s') % employee.name)
                elif vdu_type != 'men' and not vv_2:
                    raise exceptions.UserError(_(
                        'Pateikiant valandinį VDU būtina nurodyti dirbtas valandas. Darbuotojas: %s') % employee.name)
                date_vdu = date_data_dt - relativedelta(months=1)
                year = date_vdu.year
                month = date_vdu.month
                year_id = env['years'].search([('code', '=', year)], limit=1).id
                if not year_id:
                    raise exceptions.Warning(_('Nerasti metai sistemoje: %s') % year)
                vdu_record = self.env['employee.vdu'].search([('employee_id', '=', employee.id),
                                                              ('year_id', '=', year_id),
                                                              ('month', '=', month)], limit=1)
                if not vdu_record:
                    vdu_record = self.env['employee.vdu'].create({'employee_id': employee.id,
                                                                  'year_id': year_id,
                                                                  'month': month})
                time_worked_2 = dd_2 if vdu_type == 'men' else vv_2
                # P3:DivOK -- vdu_2 is always float
                vdu_record.write({'vdu': vdu_2 / time_worked_2,
                                  'hours_worked': vv_2,
                                  'days_worked': dd_2,
                                  'amount': vdu_2,
                                  'type': vdu_type})
                importing_vdu = True
            if (record.dd_3 is not None or record.vv_3 is not None) and record.vdu_3 is not None:
                try:
                    dd_3 = int(round(float(record.dd_3)))
                except:
                    dd_3 = False
                try:
                    vv_3 = float(record.vv_3)
                except:
                    vv_3 = False
                try:
                    vdu_3 = float(record.vdu_3)
                except:
                    vdu_3 = False
                if not vdu_3:
                    raise exceptions.UserError(_('Neteisingai nurodytas %s darbuotojui VDU') % employee.name)
                if not dd_3 and not vv_3:
                    raise exceptions.UserError(_(
                        'Neteisingai nurodytas %s darbuotojui dirbtų dienų arba valandų skaičius') % employee.name)
                if vdu_type == 'men' and not dd_3:
                    raise exceptions.UserError(_(
                        'Pateikiant mėnesinį VDU būtina nurodyti dirbtas dienas. Darbuotojas: %s') % employee.name)
                elif vdu_type != 'men' and not vv_3:
                    raise exceptions.UserError(_(
                        'Pateikiant valandinį VDU būtina nurodyti dirbtas valandas. Darbuotojas: %s') % employee.name)
                date_vdu = date_data_dt - relativedelta(months=2)
                year = date_vdu.year
                month = date_vdu.month
                year_id = env['years'].search([('code', '=', year)], limit=1).id
                if not year_id:
                    raise exceptions.Warning(_('Nerasti metai sistemoje: %s') % year)
                vdu_record = self.env['employee.vdu'].search([('employee_id', '=', employee.id),
                                                              ('year_id', '=', year_id),
                                                              ('month', '=', month)], limit=1)
                if not vdu_record:
                    vdu_record = self.env['employee.vdu'].create({'employee_id': employee.id,
                                                                  'year_id': year_id,
                                                                  'month': month})
                time_worked_3 = dd_3 if vdu_type == 'men' else vv_3
                # P3:DivOK -- vdu_3 is always float
                vdu_record.write({'vdu': vdu_3 / time_worked_3,
                                  'amount': vdu_3,
                                  'hours_worked': vv_3,
                                  'days_worked': dd_3,
                                  'type': vdu_type})
                importing_vdu = True

            job = env['hr.job']
            if record.job_name:
                job = env['hr.job'].search([('name', '=', record.job_name)], limit=1)
                if not job:
                    job = env['hr.job'].create({'name': record.job_name})

            if job:
                employee.sudo().write({'job_id': job.id})

            employee_job = job or employee.job_id

            if record.contract_name:
                schedule_types = {
                    'fiksuotas': 'fixed',
                    'lankstus': 'lankstus',
                    'suskaidytas': 'suskaidytos',
                    'individualus': 'individualus',
                }
                schedule_type = convert_to_string(record.contract_schedule_type).lower().strip()
                if schedule_type == 'suminė apskaita':
                    schedule_type = 'sumine'
                elif not schedule_types.get(schedule_type, False):
                    raise exceptions.UserError(_('Nenumatytas grafiko tipas'))
                else:
                    schedule_type = schedule_types[schedule_type]
                if not record.date_data:
                    raise exceptions.Warning(_('Nenurodyta, kuriam laikui nuimti duomenys'))

                if not record.contract_etatas and vdu_type == 'val':
                    etatas = 1.0
                elif not record.contract_etatas and not record.contract_monthly_hours:
                    raise exceptions.Warning(
                        _('Nenurodyta nei etato dalis, nei valandų skaičius per mėnesį. Eilutė: %s') % (i + 1))
                elif not record.contract_etatas and record.contract_monthly_hours:
                    etatas = 1.0
                    max_monthly_hours = round(record.contract_monthly_hours, 2)
                    men_constraint = 'val_per_men'
                else:
                    try:
                        etatas = float(record.contract_etatas)
                    except ValueError:
                        raise exceptions.UserError(
                            _('Neteisingai nurodyta etato dalis. Etato reikšmė turi būti skaitinė. Eilutė: %s') % (
                                    i + 1))
                    if not 0 < etatas <= 1.5:
                        raise exceptions.Warning(
                            _('Neteisingai nurodyta etato dalis. Etatas turi būti tarp 0 ir 1.5. Eilutė: %s') % (
                                    i + 1))
                etatas = round(etatas, 2)

                if not record.contract_date_start:
                    raise exceptions.Warning(_('Darbuotojo %s sutarties %s pradžios data nenurodyta') % (
                        employee.name, record.contract_name))
                try:
                    date_start = record.contract_date_start.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
                except:
                    raise exceptions.Warning(_('Darbuotojo %s sutarties %s pradžios data neteisingai suformatuota') % (
                        employee.name, record.contract_name))
                struct = env.ref(
                    'l10n_lt_payroll.hr_payroll_salary_structure_valandinis') if record.contract_wage_type and record.contract_wage_type.lower().strip() == 'valandinis' else env.ref(
                    'l10n_lt_payroll.hr_payroll_salary_structure_menesinis')
                contract = env['hr.contract'].search(
                    [('name', '=', record.contract_name), ('employee_id', '=', employee.id)], limit=1)
                if contract and importing_vdu:
                    Holidays = env['hr.holidays']
                    contract_holidays = Holidays.search([('contract_id', '=', contract.id)])
                    contract_holidays.write({'recalculate': True})
                    Holidays.cron_recalculate()
                elif not contract:
                    contract_vals = {'name': record.contract_name,
                                     'employee_id': employee.id,
                                     'date_start': date_start,
                                     # 'struct_id': struct.id
                                     }
                    contract = env['hr.contract'].create(contract_vals)
                wage = record.contract_hourly_rate if record.contract_wage_type and record.contract_wage_type.lower().strip() == 'valandinis' else record.contract_monthly_wage
                appointment = env['hr.contract.appointment'].search(
                    [('date_start', '=', date_start), ('contract_id', '=', contract.id)])
                # schedule
                sched_shorter_before_holidays = True if record.contract_shorter_before_holidays in ['Taip',
                                                                                                    'taip',
                                                                                                    'TAIP'] else False
                if not appointment:
                    schedule_template = self.env['schedule.template'].create({
                        'template_type': schedule_type,
                        'fixed_attendance_ids': get_attendances(etatas),
                        'etatas_stored': etatas,
                        'wage_calculated_in_days': True if schedule_type == 'fixed' and struct.code == 'MEN' else False,
                        'shorter_before_holidays': sched_shorter_before_holidays,
                    })
                    schedule_template.set_default_fixed_attendances()

                    if not employee_job:
                        raise exceptions.UserError(_('Nenurodytas darbuotojo %s pareigų pavadinimas') % employee.name)
                    employee_department = department or employee.department_id

                    appointment_values = {'name': record.contract_name,
                                          'contract_id': contract.id,
                                          'date_start': date_start,
                                          'wage': wage,
                                          'struct_id': struct.id if struct else False,
                                          'schedule_template_id': schedule_template.id,
                                          'job_id': employee_job.id,
                                          'department_id': employee_department.id,
                                          }
                    appointment = env['hr.contract.appointment'].create(appointment_values)
                date_end = record.date_end.strftime(tools.DEFAULT_SERVER_DATE_FORMAT) if record.date_end else False
                # Accumulate vals and do a single write to each record instead of lots of
                # assignments, so constraints are not triggered correctly without any
                # missing data that would be added later in the code
                template_vals = {'men_constraint': men_constraint}
                contract_vals = {'date_end': date_end, 'override_taxes': False}

                sodra_papild = record.sodra_papildomai
                sodra_papildomai = True if sodra_papild and sodra_papild.lower().strip() == 'taip' else False
                employee.sodra_papildomai = sodra_papildomai

                sodra_papild_type = record.sodra_papildomai_type.lower().strip() if record.sodra_papildomai_type else False
                sodra_papildomai_type = 'exponential' if sodra_papild_type and sodra_papild_type == 'palaipsniui' else 'full'

                appointment_vals = {
                    'wage': wage,
                    'sodra_papildomai': sodra_papildomai,
                    'sodra_papildomai_type': sodra_papildomai_type,
                    'shorter_before_holidays': sched_shorter_before_holidays
                }

                # Gather up the values
                if men_constraint == 'val_per_men':
                    template_vals['max_monthly_hours'] = max_monthly_hours
                if record.holidays_num_days_type and 'kalend' in record.holidays_num_days_type.lower().strip():
                    appointment_vals['leaves_accumulation_type'] = 'calendar_days'
                else:
                    appointment_vals['leaves_accumulation_type'] = 'work_days'
                if record.holidays_num_days_yearly:
                    appointment_vals['num_leaves_per_year'] = record.holidays_num_days_yearly or 20.0
                if record.work_norm:
                    template_vals['work_norm'] = float(record.work_norm)

                if len(appointment.contract_id.appointment_ids) == 1:
                    appointment_vals['date_end'] = date_end

                if record.contract_apply_npd and record.contract_apply_npd.lower().strip() == 'taip':
                    appointment_vals['use_npd'] = True
                    if record.contract_npd is not None:
                        contract_vals.update({'npd': record.contract_npd, 'override_taxes': True})
                else:
                    appointment_vals['use_npd'] = False
                # if record.contract_pnpd == 'Ne':
                #     contract.pnpd = 0
                #     contract.use_pnpd = True
                #     contract.override_taxes = True
                # else:
                #     contract.use_pnpd = False

                # Make a single write to each record
                appointment.schedule_template_id.write(template_vals)
                appointment.write(appointment_vals)
                contract.write(contract_vals)

                if not record.date_data:
                    raise exceptions.Warning(_('Nenurodyta, kuriam laikui nuimtas atostogų likutis'))

                if record.holidays_num_days:
                    try:
                        holiday_days = float(record.holidays_num_days.replace(',', '.'))
                    except AttributeError:
                        holiday_days = float(record.holidays_num_days)
                    employee.with_context(date=date_data).remaining_leaves = holiday_days
            bank_account_num = str(record.bank_account or '')
            if bank_account_num:
                partner_bank_rec = env['res.partner.bank'].search([('acc_number', '=', bank_account_num),
                                                                   ('partner_id', '=', employee.address_home_id.id)],
                                                                  limit=1)
                if not partner_bank_rec:
                    partner_bank_rec = env['res.partner.bank'].create({'acc_number': bank_account_num,
                                                                       'partner_id': employee.address_home_id.id})
                if not partner_bank_rec.bank_id:
                    partner_bank_rec.onchange_acc_number()
                employee.bank_account_id = partner_bank_rec.id
            if record.contract_num_children is not None:
                employee.seimos_nariai.filtered(lambda r: r.seimos_narys == 'vaikas').unlink()
                num_children = int(round(record.contract_num_children or 0))
                if num_children > 0:
                    child_name = record.child_name
                    if not record.child_date_born:
                        continue
                        # raise exceptions.UserError(_('Nenurodyda vaiko gimimo data'))
                    date_born = record.child_date_born.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
                    date_school_end = record.child_date_school_end and record.child_date_school_end.strftime(
                        tools.DEFAULT_SERVER_DATE_FORMAT) or ''
                    vaikas_val = {'name': child_name,
                                  'birthday': date_born,
                                  'mokyklos_baigimas': date_school_end,
                                  'seimos_narys': 'vaikas',
                                  'employee_id': employee.id}
                    env['seimos.nariai'].create(vaikas_val)
                for k in xrange(1, num_children):
                    i, row = row_iter_obj.next()
                    values = get_all_values(row)
                    record = ImportRecord(values, header_mapped)
                    child_name = record.child_name
                    if not record.child_date_born:
                        continue
                        # raise exceptions.UserError(_('Nenurodyda vaiko gimimo data'))
                    date_born = record.child_date_born.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
                    date_school_end = record.child_date_school_end and record.child_date_school_end.strftime(
                        tools.DEFAULT_SERVER_DATE_FORMAT) or ''
                    vaikas_val = {'name': child_name,
                                  'birthday': date_born,
                                  'mokyklos_baigimas': date_school_end,
                                  'seimos_narys': 'vaikas',
                                  'employee_id': employee.id}
                    env['seimos.nariai'].create(vaikas_val)

    except StopIteration:
        pass

    except Exception as exc:
        raise exceptions.UserError(
            _('Klaida importuojant duomenis, eilutė %d: %s') % (i + 1, str(exc.args[0] if exc.args else exc.message)))


aml_mapping = {
    u'Žurnalo kodas': 'journal_code',
    u'Buh. sąskaitos kodas': 'account_code',
    u'Buh. sąskaitos pavadinimas': 'account_name',
    u'Pavadinimas': 'name',
    u'D': 'debit',
    u'K': 'credit',
    u'Data': 'date',
    u'Terminas': 'date_maturity',
    u'Partnerio pavadinimas': 'partner_name',
    u'Partnerio kodas': 'partner_code',
    u'Valiuta': 'currency',
    u'Suma valiuta': 'amount_currency',
    u'Analitinės sąskaitos kodas': 'account_analytic_code',
    u'Analitinės sąskaitos pavadinimas': 'account_analytic_name',
    u'A klasės kodas': 'a_klase_kodas_id',
    u'B klasės kodas': 'b_klase_kodas_id',
    u'Grupavimo klasifikatorius': 'group_num',
}


def import_aml(self, import_file):
    def convert_date(date_val, line):
        """Checks whether passed date matches the date format"""
        try:
            if isinstance(date_val, basestring):
                datetime.strptime(date_val, tools.DEFAULT_SERVER_DATE_FORMAT)
            else:
                date_val = date_val.strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
        except:
            raise exceptions.UserError(_('Blogas datos formatas %s eilutėje %s') % (date_val, line + 1))
        return date_val

    env = self.sudo().env
    account_obj = env['account.account']
    account_move_obj = env['account.move']
    off_books_account_id = self.env['account.account'].search([('code', '=', '999999')])
    if not off_books_account_id:
        raise exceptions.UserError(_('Nepavyko surasti atitinkamos apskaitos informacijos.'))
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name=u'Likučiai')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    init_balance_journal = self.env['account.journal'].search([('code', '=', 'START')], limit=1)
    if not init_balance_journal:
        init_balance_journal = self.env['account.journal'].create({
            'name': 'Pradiniai likučiai',
            'code': 'START',
            'type': 'general',
            'update_posted': True,
        })
    partner_obj = env['res.partner']

    # Lines can either be general move lines,
    # or initial balance lines
    grouped_init_lines = {}
    grouped_general_lines = {}
    for i, row in enumerate(iter_rows):
        if not header:
            header = get_all_values(row)
            header_mapped = get_mapped(header, aml_mapping)
            continue
        values = get_all_values(row)
        if len(set(values)) == 1:
            break
        record = ImportRecord(values, header_mapped)
        code = str(record.account_code or '')
        debit = record.debit
        credit = record.credit
        date = record.date
        date_maturity = record.date_maturity

        journal_to_use = init_balance_journal
        general_line = False
        # Check for journal code
        if record.journal_code:
            general_line = True
            journal_code = str(record.journal_code)
            journal = env['account.journal'].search([('code', '=', journal_code)], limit=1)
            if not journal:
                raise exceptions.UserError(_('Nerastas žurnalas pagal paduotą kodą eilutėje %s') % str(i + 1))
            journal_to_use = journal

        if not date:
            raise exceptions.UserError(_('Nenurodyta data eilutėje %s') % str(i + 1))

        # Try to convert both of the dates
        date = convert_date(date, i)
        if date_maturity:
            date_maturity = convert_date(date_maturity, i)

        if not code or (not debit and not credit):
            continue
        if isinstance(debit, tuple([str, unicode])) and isinstance(credit, tuple([str, unicode])):
            if '=' in debit or '=' in credit:
                continue
            else:
                try:
                    debit = float(debit)
                    credit = float(credit)
                except:
                    raise exceptions.UserError(_('Neteisingas duomenų formatas %s eilutėje.') % (i + 1))
        if (debit and not isinstance(debit, tuple([int, float, long]))) or (
                credit and not isinstance(credit, tuple([int, float, long]))):
            raise exceptions.UserError(_('Neteisingas duomenų formatas %s eilutėje.') % (i + 1))
        if not debit:
            debit = 0.0
        if not credit:
            credit = 0.0
        debit = tools.float_round(debit, precision_digits=2)
        credit = tools.float_round(credit, precision_digits=2)
        if tools.float_is_zero(debit, precision_digits=2) and tools.float_is_zero(credit, precision_digits=2):
            continue
        account_id = account_obj.with_context(show_views=True).search([('code', '=', code)], limit=1)
        if not account_id and not code:
            raise exceptions.UserError(_('Nepavyko rasti atitinkamos DK sąskaitos %s eilutėje.') % (i + 1))
        elif not account_id and code and record.account_name:
            account_vals = {
                'code': code,
                'name': record.account_name,
                'company_id': self.env.user.company_id.id,
            }
            code_class = str(code[0])
            if code_class == '1':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_fixed_assets').id
            elif code_class == '2':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_current_assets').id
            elif code_class == '3':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_equity').id
            elif code_class == '4':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_current_liabilities').id
            elif code_class == '5':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_revenue').id
            elif code_class == '6':
                account_vals['user_type_id'] = self.env.ref('account.data_account_type_expenses').id
            else:
                account_vals['user_type_id'] = self.env.ref('l10n_lt.account_type_uzbalanse').id
            try:
                account_id = account_obj.create(account_vals)
            except:
                raise exceptions.UserError(_('Nepavyko pridėti naujos DK sąskaitos numerio %s.') % str(code))
        if not account_id:
            raise exceptions.UserError(_('Nepavyko rasti atitinkamos DK sąskaitos %s eilutėje.') % (i + 1))
        partner_name = convert_to_string(record.partner_name)
        partner_code = convert_to_string(record.partner_code)
        partner_id = False
        if partner_name and isinstance(partner_name, basestring):
            partner = partner_obj.search([('name', 'like', partner_name)], limit=1)
            is_company = True
            if partner_code:
                if not partner:
                    partner = partner_obj.search([('kodas', '=', partner_code)], limit=1)

            if not partner_code or len(partner_code) == 11:
                is_company = False

            if not partner:
                partner = partner_obj.create({
                    'name': partner_name,
                    'kodas': partner_code,
                    'is_company': is_company,
                })
            partner_id = partner.id
        account_analytic_id = False
        if record.account_analytic_code:
            analytic_code = str(record.account_analytic_code)
            account_analytic = env['account.analytic.account'].search([('code', '=', analytic_code)], limit=1)
            if not account_analytic:
                if not record.account_analytic_name:
                    raise exceptions.UserError(
                        _('Nenurodytas analitinės sąskaitos %s pavadinimas eilutėje %s') % (analytic_code, str(i + 1)))
                account_analytic = self.env['account.analytic.account'].create(
                    {'name': str(record.account_analytic_name),
                     'code': analytic_code})
            account_analytic_id = account_analytic.id
        currency_id = False
        amount_currency = 0.0
        if record.currency:
            currency_name = str(record.currency).upper()
            env.cr.execute('SELECT id from res_currency where name = %s limit 1', (currency_name,))
            querry_res = env.cr.fetchall()
            if querry_res:
                currency_id = querry_res[0][0]
            else:
                raise exceptions.UserError(
                    _('Nerasta valiuta %s, valiutos kodas turi būti trys didžiosios raidės') % str(record.currency))
            currency = env['res.currency'].browse(currency_id)
            if currency_id != self.company_id.currency_id.id:
                currency.active = True
                amount_currency = float(record.amount_currency or 0.0)
            elif currency_id and currency_id == self.company_id.currency_id.id:
                currency_id = False

        a_klase_kodas, b_klase_kodas = False, False
        if record.a_klase_kodas_id:
            a_code = str(record.a_klase_kodas_id)
            a_klase_kodas = env['a.klase.kodas'].search([('code', '=', a_code)], limit=1)
            if not a_klase_kodas:
                raise exceptions.UserError(_('Nepavyko rasti atitinkamo A klasės kodo %s eilutėje.') % (i + 1))

        if record.b_klase_kodas_id:
            b_code = str(record.b_klase_kodas_id)
            b_klase_kodas = env['b.klase.kodas'].search([('code', '=', b_code)], limit=1)
            if not b_klase_kodas:
                raise exceptions.UserError(_('Nepavyko rasti atitinkamo B klasės kodo %s eilutėje.') % (i + 1))

        line_vals = {
            'name': record.name or '/',
            'account_id': account_id.id,
            'debit': 0.0,
            'credit': 0.0,
            'journal_id': journal_to_use.id,
            'partner_id': partner_id,
            'analytic_account_id': account_analytic_id,
            'currency_id': currency_id,
            'amount_currency': amount_currency,
            'date_maturity': date_maturity,
        }

        if debit > 0.0:
            line_vals['debit'] = debit
        if credit > 0.0:
            line_vals['credit'] = credit
        if a_klase_kodas:
            line_vals['a_klase_kodas_id'] = a_klase_kodas.id
        if b_klase_kodas:
            line_vals['b_klase_kodas_id'] = b_klase_kodas.id

        # Add either to general or init balance group
        if general_line:
            grouped_general_lines.setdefault(record.group_num, {})
            grouped_general_lines[record.group_num].setdefault(journal_to_use, {})
            grouped_general_lines[record.group_num][journal_to_use].setdefault(date, []).append(line_vals)
        else:
            grouped_init_lines.setdefault(record.group_num, {})
            grouped_init_lines[record.group_num].setdefault(date, []).append(line_vals)

    is_view_line_account_problems = ''
    # Create entries for for initial balance aml import
    for group_num, lines_by_date in iteritems(grouped_init_lines):
        for date, lines in iteritems(lines_by_date):
            total_credit = sum(ln['credit'] for ln in lines)
            total_debit = sum(ln['debit'] for ln in lines)
            if tools.float_compare(total_credit, total_debit, precision_digits=2):
                if total_debit > total_credit:
                    last_credit = total_debit - total_credit
                    last_debit = 0.0
                else:
                    last_debit = total_credit - total_debit
                    last_credit = 0.0
                last_line_vals = {
                    'name': u'Balansuojantis įrašas',
                    'account_id': off_books_account_id.id,
                    'debit': last_debit,
                    'credit': last_credit,
                    'journal_id': init_balance_journal.id,
                }
                lines_by_date[date].append(last_line_vals)
            date_line_vals = lines_by_date[date]
            if '--test-enable' not in sys.argv:  # checks should not be applied during tests
                acc_ids = [dl['account_id'] for dl in date_line_vals]
                accounts = env['account.account'].browse(acc_ids)
                view_acc_ids = accounts.filtered(lambda a: a.is_view)
                for view_acc in view_acc_ids:
                    is_view_line_account_problems += '%s - %s\n' % (view_acc.code, view_acc.name)
                if is_view_line_account_problems:
                    continue
            move_vals = {
                'ref': u'Pradiniai buhalteriniai likučiai',
                'date': date,
                'journal_id': init_balance_journal.id,
                'line_ids': [(0, 0, val) for val in date_line_vals],
            }
            move = account_move_obj.create(move_vals)
            move.post()

    if is_view_line_account_problems:
        msg = _('Klaida, negalite naudoti šių suminių sąskaitų: \n\n%s') % is_view_line_account_problems
        raise exceptions.ValidationError(msg)

    # Create entries for general aml import
    for group_num, lines_by_journal in grouped_general_lines.items():
        for journal, lines_by_date in lines_by_journal.items():
            for date, lines in lines_by_date.items():
                move_vals = {
                    'date': date,
                    'journal_id': journal.id,
                    'line_ids': [(0, 0, val) for val in lines],
                }
                move = account_move_obj.create(move_vals)
                move.post()


du_mapping = {
    u'Darbuotojo Vardas Pavardė': 'employee_name',
    u'Asmens kodas': 'identification_code',
    u'Sutarties numeris': 'contract_name',
    u'Data iki': 'date_to',
    u'Darbo dienų': 'days_worked',
    u'Darbo valandų': 'hours_worked',
    u'Pagrindinis atlyginimas': 'salary_main',
    u'Bruto': 'suma_bruto',
    u'Už darbą poilsio dienomis': 'suma_poilsio',
    u'Priedas už darbo rezultatus': 'priedas',
    u'Premija': 'premija',
    u'Ketvirtinė premija': 'premija',
    u'Nedarbingumo išmoka': 'nedarbingumo_išmoka',
    u'Apmokėjimas tėvams (motinoms)': 'suma_papild_tėv',
    u'Kasmetinės atostogos': 'suma_a',
    u'GPM': 'suma_gpm',
    u'Darbuotojo socialinis draudimas': 'suma_d_sd',
    u'Pajamos natūra': 'natura',
    u'Dienpinigiai': 'dienpinigiai',
    u'Mokėtinas atlyginimas': 'moketinas_atlyginimas',
    u'Darbdavio sodra': 'sodra_darbdavio',
    u'Avansas': 'avansas',
    u'Išeitinė kompensacija': 'iseitines',
    u'Kompensacija už nepanaudotas atostogas': 'nepanaudotos_atostogos',
    u'Išskaitos': u'isskaitos',
    u'Prastova ne dėl darbuotojo kaltės': u'prastova'
}

salary_code_mapping = {
    'salary_main': 'BM',
    'suma_poilsio': 'DP',
    'suma_bruto': 'MEN',
    'premija': 'PR',
    'priedas': 'PD',
    'nedarbingumo_išmoka': 'L',
    'suma_papild_tėv': 'T',
    'suma_a': 'A',
    'suma_gpm': 'GPM',
    'suma_d_sd': 'SDB',
    'natura': 'NTR',
    'dienpinigiai': 'KM',
    'moketinas_atlyginimas': 'M',
    'sodra_darbdavio': 'SDD',
    'avansas': 'AVN',
    'iseitines': 'IST',
    'nepanaudotos_atostogos': 'AK',
    'isskaitos': u'IŠSK',
    'prastova': 'PN',
}


def import_du(self, import_file):
    def get_payslip_line_vals(env, code, amount):
        rule = env['hr.salary.rule'].search([('code', '=', code)], limit=1)
        qty = 1
        rate = 100
        if not rule:
            raise exceptions.Warning(_('Nerasta atlyginimo taisyklė %s') % code)
        payslip_line_vals = {
            'salary_rule_id': rule.id,
            'contract_id': contract.id,
            'name': rule.name,
            'code': rule.code,
            'category_id': rule.category_id.id,
            'sequence': rule.sequence,
            'appears_on_payslip': rule.appears_on_payslip,
            'condition_select': rule.condition_select,
            'condition_python': rule.condition_python,
            'condition_range': rule.condition_range,
            'condition_range_min': rule.condition_range_min,
            'condition_range_max': rule.condition_range_max,
            'amount_select': rule.amount_select,
            'amount_fix': rule.amount_fix,
            'amount_python_compute': rule.amount_python_compute,
            'amount_percentage': rule.amount_percentage,
            'amount_percentage_base': rule.amount_percentage_base,
            'register_id': rule.register_id.id,
            'amount': amount,
            'employee_id': contract.employee_id.id,
            'quantity': qty,
            'rate': rate,
        }
        return payslip_line_vals

    env = self.sudo().env
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_active_sheet()
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    num_rows = sheet.max_row
    row_iter_obj = enumerate(iter_rows)
    i = 0
    journal_id = env.user.company_id.salary_journal_id.id
    a_klase_kodai = []
    indices_with_a_klase = []
    cutoff_priskaitymai = 0
    if not journal_id:
        raise exceptions.Warning(_('Nesukonfigūruotas atlyginimų žurnalas'))
    try:
        while i < num_rows:
            i, row = row_iter_obj.next()
            if row and row[0].value == 'A klasė':
                a_klase_kodai = [r.value for r in row]
                indices_with_a_klase = [j for j, val in enumerate(a_klase_kodai) if val and j > 0]
                continue
            if not header:
                header = get_all_values(row)
                try:
                    cutoff_priskaitymai = [c.value for c in row].index('Bruto')
                except ValueError:
                    raise exceptions.Warning(_('Bruto reikšmė nerasta'))
                header_mapped = get_mapped(header, du_mapping)
                continue
            # if a_klase_kodai and header:
            #     a_klase_kodai_mapping = dict(zip(header, a_klase_kodai))
            values = get_all_values(row)
            if len(set(values)) == 1:
                break
            record = ImportRecord(values, header_mapped)
            employee_name = record.employee_name
            employee_identification_id = str(record.identification_code or '')
            if not employee_identification_id or len(employee_identification_id) == 0:
                raise exceptions.Warning(_('Nenurodytas darbuotojo %s asmens kodas') % employee_name)

            employee = env['hr.employee'].search([
                ('identification_id', '=', employee_identification_id),
                '|',
                ('active', '=', True),
                ('active', '=', False)
            ])
            if not employee:
                employee = env['hr.employee'].search(
                    [('name', '=', employee_name), '|', ('active', '=', True), ('active', '=', False)])
            if len(employee) > 1:
                raise exceptions.Warning(
                    _('Rasti keli darbuotojai tuo pačiu vardu %s. Pabandykite pirmiausia pakeisti darbuotojo vardą '
                      'pridedant unikalų prierašą ir bandykite dar kartą.') % employee_name)
            if not employee:
                raise exceptions.Warning(_('Nerastas darbuotojas %s.') % employee_name)
            if not employee.active:
                employee.write({
                    'active': True,
                })
            contract_name = str(record.contract_name)
            contract = env['hr.contract'].search([('name', '=', contract_name), ('employee_id', '=', employee.id)],
                                                 limit=1)
            if not contract:
                raise exceptions.Warning(_('Darbuotojo %s kontraktas %s nerastas') % (employee_name, contract_name))
            date = record.date_to
            if isinstance(date, basestring):
                try:
                    date = datetime.strptime(date, tools.DEFAULT_SERVER_DATE_FORMAT)
                except:
                    raise exceptions.Warning(_('Netinkamas datos formatas eilutėje %s') % (i + 1))
            if not isinstance(date, datetime):
                raise exceptions.Warning(_('Netinkamas datos formatas eilutėje %s') % (i+1))
            date_from = (date + relativedelta(day=1)).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
            date_to = (date + relativedelta(day=31)).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
            payslip = env['hr.payslip'].search([('date_from', '=', date_from),
                                                ('date_to', '=', date_to),
                                                ('contract_id', '=', contract.id),
                                                ('imported', '=', True)], limit=1)
            if not payslip:
                payslip = self.env['hr.payslip'].create({'date_from': date_from,
                                                         'date_to': date_to,
                                                         'contract_id': contract.id,
                                                         'employee_id': employee.id,
                                                         'imported': True})
            men_struct = self.env.ref('l10n_lt_payroll.hr_payroll_salary_structure_menesinis').id
            val_struct = self.env.ref('l10n_lt_payroll.hr_payroll_salary_structure_valandinis').id
            is_monthly_struct = payslip.contract_id.struct_id.code == 'MEN'
            struct_to_use = men_struct if is_monthly_struct else val_struct
            payslip.write({'state': 'done',
                           'struct_id': struct_to_use,
                           'name': 'Istorinis algalapis %s, %s - %s' % (employee.name, date_from, date_to)
                           })
            input_line_vals = {'name': 'Faktiškai dirbtas laikas',
                               'code': 'FD',
                               'sequence': 1,
                               'number_of_hours': record.hours_worked,
                               'number_of_days': record.days_worked,
                               'contract_id': contract.id}
            payslip.worked_days_line_ids = [(5,), (0, 0, input_line_vals)]
            payslip_lines = [(5,)]
            for var_name, salary_code in iteritems(salary_code_mapping):
                amount = record.__getattr__(var_name)
                if var_name == 'suma_bruto' and not is_monthly_struct:
                    new_line_vals = get_payslip_line_vals(env, 'VAL', amount)
                else:
                    new_line_vals = get_payslip_line_vals(env, salary_code, amount)
                payslip_lines.append((0, 0, new_line_vals))
                payslip_lines.append((0, 0, get_payslip_line_vals(env, 'BRUTON', amount)))
            payslip_other_lines = [(5,)]
            for j in indices_with_a_klase:
                a_klase_kodas = a_klase_kodai[j]
                a_klase_id = self.env['a.klase.kodas'].search([('code', '=', a_klase_kodas)], limit=1).id
                if not a_klase_id:
                    raise exceptions.Warning(_('Nerastas a klasės kodas %s') % a_klase_kodas)
                if j < cutoff_priskaitymai:
                    type = 'priskaitymai'
                else:
                    type = 'gpm'
                amount = row[j].value or 0.0
                if tools.float_compare(amount, 0, precision_digits=2) != 0:
                    l_vals = {'type': type,
                              'a_klase_kodas_id': a_klase_id,
                              'amount': row[j].value,
                              'name': header[j]}
                    payslip_other_lines.append((0, 0, l_vals))
            payslip.line_ids = payslip_lines
            payslip.other_line_ids = payslip_other_lines
            payslip.generate_vdu()

    except StopIteration:
        pass


class RoboCompanySettings(models.TransientModel):
    _name = 'robo.company.settings'

    analytic_account_id = fields.Many2one('account.analytic.account', string='Numatytoji analitinė sąskaita',
                                          required=False)
    company_id = fields.Many2one('res.company')
    company_logo = fields.Binary()
    company_name = fields.Char(required=True)
    company_ceo = fields.Many2one('hr.employee', required=True)
    company_kodas = fields.Char(readonly=True)
    company_vat = fields.Char()
    company_street = fields.Char()
    company_street2 = fields.Char()
    company_city = fields.Char()
    company_zip = fields.Char()
    company_country_id = fields.Many2one('res.country')

    # company_bank_account = fields.Char()
    # company_bank_name = fields.Char()
    # company_bank_code = fields.Char()
    # company_bank_bic = fields.Char()
    # company_bank_currency = fields.Many2one('res.currency')
    company_bank_accounts = fields.One2many('robo.company.settings.bank', 'settings_id', string='Banko sąskaitos')
    company_draudejo_kodas = fields.Char()
    company_savivaldybe = fields.Selection(SAVIVALDYBES)
    company_evrk = fields.Many2one('evrk.kodai')
    company_email = fields.Char()
    company_phone = fields.Char()
    company_fax = fields.Char()
    company_website = fields.Char()
    payroll_bank_journal_id = fields.Many2one('account.journal', string='Darbo užmokesčio banko žurnalas',
                                              domain="[('type', '=', 'bank')]",
                                              help='Banko žurnalas, naudojamas darbo užmokesčio mokėjimams')
    advance_payment_day = fields.Integer(string='Avanso mokėjimo diena')
    proforma_series = fields.Char(required=False)
    proforma_number = fields.Integer(required=True)
    proforma_length = fields.Integer(required=True)
    proforma_actual_number = fields.Char(compute='_proforma_actual_number')
    invoice_series = fields.Char(required=False)
    invoice_journal_id = fields.Many2one('account.journal', string='Žurnalas', domain="[('type', '=', 'sale')]",
                                         groups='robo_basic.group_robo_select_invoice_journal')
    invoice_number = fields.Integer(required=True)
    invoice_length = fields.Integer(required=True)
    invoice_actual_number = fields.Char(compute='_invoice_actual_number')
    asset_prefix = fields.Char(string='Priešdėlis', groups='robo_basic.group_robo_premium_accountant')
    asset_suffix = fields.Char(string='Sufiksas', groups='robo_basic.group_robo_premium_accountant')
    asset_padding = fields.Integer(string='Sekos dydis', groups='robo_basic.group_robo_premium_accountant')
    asset_next_number = fields.Integer(string='Sekantis numeris', groups='robo_basic.group_robo_premium_accountant')
    in_payment_series = fields.Char(required=False)
    in_cash_journal_id = fields.Many2one('account.journal', string='In cash journal', domain="[('type', '=', 'cash')]")
    in_payment_number = fields.Integer(required=True)
    in_payment_length = fields.Integer(required=True)
    in_payment_actual_number = fields.Char(compute='_in_payment_actual_number')
    out_payment_series = fields.Char(required=False)
    out_payment_number = fields.Integer(required=True)
    out_payment_length = fields.Integer(required=True)
    out_payment_actual_number = fields.Char(compute='_out_payment_actual_number')
    in_cash_receipt_series = fields.Char(required=False)
    in_cash_receipt_number = fields.Integer(required=True)
    in_cash_receipt_length = fields.Integer(required=True)
    in_cash_receipt_actual_number = fields.Char(compute='_in_cash_receipt_actual_number')
    out_cash_receipt_series = fields.Char(required=False)
    out_cash_receipt_number = fields.Integer(required=True)
    out_cash_receipt_length = fields.Integer(required=True)
    out_cash_receipt_actual_number = fields.Char(compute='_out_cash_receipt_actual_number')
    contract_series = fields.Char()
    contract_number = fields.Integer(required=True)
    contract_length = fields.Integer(required=True)
    contract_actual_number = fields.Char(compute='_contract_actual_number')
    credit_invoice_series = fields.Char(required=False)
    credit_invoice_number = fields.Integer(required=True)
    credit_invoice_length = fields.Integer(required=True)
    credit_invoice_actual_number = fields.Char(compute='_compute_credit_invoice_actual_number')
    import_partners = fields.Binary()
    import_customer_invoices = fields.Binary()
    import_supplier_invoices = fields.Binary()
    import_products = fields.Binary()
    import_financials = fields.Binary()
    import_employees = fields.Binary()
    import_aml = fields.Binary()
    import_du = fields.Binary()
    module_robo_stock = fields.Boolean(string='Aktyvuoti sandėlio apskaitą klientui',
                                       groups='robo_basic.group_robo_premium_accountant')
    module_work_schedule = fields.Boolean(string='Aktyvuoti darbo laiko apskaitos grafiką',
                                          groups='robo_basic.group_robo_premium_accountant')
    module_work_schedule_analytics = fields.Boolean(string='Aktyvuoti analitiką pagal darbo laiko apskaitos grafikus',
                                                    groups='robo_basic.group_robo_premium_accountant')
    module_robo_analytic = fields.Boolean(string='Aktyvuoti analitiką',
                                          groups='robo_basic.group_robo_premium_accountant')
    module_robo_mrp = fields.Boolean(string='Aktyvuoti gamybos apskaitą klientui',
                                     groups='robo_basic.group_robo_premium_accountant')
    module_robo_api = fields.Boolean(string='Aktyvuoti ROBO API',
                                     groups='robo_basic.group_robo_premium_accountant')
    module_mixed_vat_rate = fields.Boolean(string='Aktyvuoti mišrų PVM',
                                           groups='robo_basic.group_robo_premium_accountant')
    module_amazon_integration = fields.Boolean(string='Aktyvuoti Amazon integraciją',
                                               groups='robo_basic.group_robo_premium_accountant')
    module_neopay = fields.Boolean(string='Instaliuoti Neopay integraciją')
    politika_atostoginiai = fields.Selection(
        [('su_du', 'Visada su darbo užmokesčiu'),
         ('rinktis', 'Leisti rinktis')],
        string='Atostoginių politika')
    politika_atostogu_suteikimas = fields.Selection(
        [('ceo', 'Tvirtina vadovas'),
         ('department', 'Padalinio vadovas')],
        string='Dokumentų tvirtinimas')
    holiday_policy_inform_manager = fields.Boolean(string='Siųsti laukiančių prašymų pranešimus padalinio vadovui')
    default_payment_term_days = fields.Integer(string='Numatytasis kliento mokėjimo terminas')
    set_default_payment_term_to_all_customers = fields.Boolean(
        string='Nustatyti visų klientų mokėjimo terminą lygų numatytajam'
    )
    isaf_default_day = fields.Integer(string='iSAF pateikimo termino diena',
                                      groups='robo_basic.group_robo_premium_accountant')
    default_supplier_payment_term_days = fields.Integer(string='Numatytasis tiekėjo mokėjimo terminas')
    inv_due_date_edit = fields.Boolean(string='Keisti sąskaitose apmokėjimo datą',
                                       groups="robo_basic.group_robo_premium_accountant")
    default_msg_receivers = fields.Many2many('res.partner', string='Papildomi žinučių gavėjai',
                                             help='Gavėjai, kai SF kūrėjas buhalteris.',
                                             groups="robo_basic.group_robo_premium_accountant")
    proforma_show_price_vat_incl = fields.Boolean(string='Išankstinėse sąskaitose rodyti sumą su PVM',
                                                  groups="robo_basic.group_robo_premium_accountant")
    apr_send_reminders = fields.Boolean(string='Aktyvuoti automatinius priminimus',
                                        help='Jei užstatyta, klientai su įjungtu priminimų siuntimu gaus el. laiškus.'
                                             ' Jei neužstatyta, nei vienas klientas negaus automatinių priminimų.')
    apr_enabled_by_default = fields.Boolean(string='Aktyvuoti visiems naujai sukurtiems partneriams pagal nutylėjimą')
    apply_default_apr_settings_to_all = fields.Boolean(
        string='Taikyti numatytus nustatymus visiems esamiems partneriams',
        help='Aktyvuoti automatinius priminimus visiems partneriams pagal nutylėjimą. Jei įjungsite, mokėjimo priminimai aktyvuosis visiems esamiems partneriams!'
    )
    apr_send_before = fields.Boolean(string='Siųsti automatinius priminimus prieš mokėjimo terminą',
                                     help='Jei užstatyta, klientas gaus automatinius mokėjimo priminimus prieš mokėjimo terminą')
    apr_send_before_ndays = fields.Integer(string='Dienų skaičius iki mokėjimo termino',
                                           help='Priminimas bus siunčiamas ... dienų iki mokėjimo termino')
    apr_send_on_date = fields.Boolean(string='Siųsti priminimą termino dieną',
                                      help='Jei užstatyta, klientas gaus automatinius mokėjimo priminimus', )
    apr_send_after = fields.Boolean(string='Siųsti automatinius priminimus po mokėjimo termino',
                                    help='Jei užstatyta, klientas gaus automatinius mokėjimo priminimus')
    apr_send_after_ndays = fields.Integer(string='Dienų skaičius nuo mokėjimo termino')
    apr_min_amount_to_send = fields.Integer(string='Mažiausia suma, nuo kurios siųsti mokėjimo priminimus')
    apr_email_cc = fields.Text(string='Siųsti laiškų kopijas el. paštu')
    apr_email_reply_to = fields.Text(string='Priminimų laiškai atsakomi \'reply-to\'',
                                     help='Šis el. paštas bus naudojamas, kai klientai norės atsakyti į priminimo laišką')
    politika_neatvykimai = fields.Selection(
        [('own', 'Darbuotojas mato tik savo neatvykimus'),
         ('department', 'Darbuotojas mato tik savo skyriaus darbuotojų neatvykimus'),
         ('all', 'Darbuotojas mato visų darbuotojų neatvykimus')], string='Neatvykimų politika')
    accumulated_days_policy = fields.Selection(
        [('allow', 'Leisti'),
         ('deny', 'Drausti')], string='Atostogos su nepakankamu likučiu', default='deny')
    worker_policy = fields.Selection([
        ('enabled', 'Darbuotojas mato tik savo kortelę'),
        ('disabled', 'Darbuotojas mato visas darbuotojų korteles')], string='Darbuotojų kortelių politika')
    automatic_salary_reconciliation = fields.Boolean(
        string='Automatiškai dengti atlyginimus',
        help='Išjungus šį pasirinkimą atlyginimų įrašai nebebus automatiškai dengimai', default=True)
    automatic_bank_reconciliation = fields.Selection([  # TODO: Remove next week
        ('full_reconcile', 'Drausti dalinį sudengimą'),
        ('partial_reconcile', 'Įgalinti dalinį sudengimą')], string='Automatinis banko išrašų dengimas',
        help='Įjungus dalinio sudengimo draudimą importuojami banko išrašai automatiškai bus dengiami tik tada, '
             'jei apskaitoje randamas įrašas(-ai) su identiška suma.', default='full_reconcile')
    company_activity_form = fields.Selection(
        [('uab', 'Uždaroji akcinė bendrovė'),
         ('vsi', 'Viešoji įstaiga'),
         ('mb', 'Mažoji bendrija'),
         ('iv', 'Individuali veikla')], string='Įmonės veiklos forma', default='uab',
        groups='robo_basic.group_robo_premium_accountant')
    uab_report_size = fields.Selection([('max', 'Išplėstinis balansas'),
                                        ('mid', 'Sutrumpintas balansas'),
                                        ('min', 'Trumpas balansas')],
                                       string='UAB balanso dydis', default='max')
    required_du_analytic = fields.Boolean(string='Priverstinė DU analitika')
    change_analytic_on_accountant_validated = fields.Boolean(
        string='Leisti keisti analitiką buhalterio patvirtintoms sąskaitoms')
    additional_analytic_actions = fields.Boolean(string='Papildomi veiksmai keičiant analitiką')
    company_invoice_text = fields.Text(string='Tekstas spausdinamas sąskaitų faktūrų apačioje')
    company_proforma_invoice_text_different_from_regular = fields.Boolean(
        string='Išankstinėse sąskaitose rodyti kitokį tekstą')
    company_invoice_proforma_text = fields.Text(string='Tekstas spausdinamas išankstinių sąskaitų faktūrų apačioje')
    force_accounting_date = fields.Boolean(string='Nenaudoti realaus laiko apskaitos')
    invoice_vat_printing_method = fields.Selection(string='Sąskaitų su PVM spausdinimo metodas', required=True,
                                                   selection=[('B2B', 'B2B'), ('B2C', 'B2C')], default='B2B')
    longterm_assets_min_val = fields.Float(string='Laikyti ilgalaikiu materialiuoju turtu išlaidas nuo, EUR',
                                           default=300, required=True,
                                           groups='robo_basic.group_robo_premium_accountant')
    longterm_non_material_assets_min_val = fields.Float(
        string='Laikyti ilgalaikiu nematerialiuoju turtu išlaidas nuo, EUR',
        default=300, required=True,
        groups='robo_basic.group_robo_premium_accountant')
    e_documents_allow_historic_signing = fields.Boolean(string='Leisti formuoti el. dokumentus praeities data',
                                                        default=False,
                                                        groups='robo_basic.group_robo_premium_accountant')
    e_documents_allow_historic_signing_spec = fields.Boolean(
        string='Leisti formuoti darbo užmokesčio el. dokumentus praeities data',
        default=True,
        groups="robo_basic.group_robo_premium_accountant",
        help='Leidimas formuoti šiuos el. dokumentus praeities data:\n'
             'Įsakymas dėl atleidimo iš darbo \n '
             'Įsakymas dėl priėmimo į darbą \n '
             'Įsakymas dėl darbo užmokesčio pakeitimo')
    sumine_apskaita_period_amount = fields.Integer(string='Suminės darbo laiko apskaitos periodo mėnesių skaičius',
                                                   groups="robo_basic.group_robo_premium_accountant",
                                                   required=True,
                                                   default=3)
    sumine_apskaita_period_start = fields.Date(string='Suminės darbo laiko apskaitos skaičiavimo pradžia',
                                               groups="robo_basic.group_robo_premium_accountant",
                                               required=True,
                                               default=datetime(2017, 1, 1))
    print_invoice_presenter = fields.Boolean(string='Spausdinti sąskaitą išrašiusio asmens vardą sąskaitų apačioje')
    invoice_cc_emails = fields.Char(string='Invoice carbon copy (CC) receivers',
                                    help='Field consists of emails that are invoice carbon copy (CC) receivers. '
                                         'Multiple emails must be separated by a semicolon (";")')
    default_action_after_fixed_term_contract_end = fields.Selection([
        ('extend', 'Create document to extend fixed term contract'),
        ('change_type', 'Create document to change contract type to indefinite duration'),
        ('terminate', 'Create document to terminate the work relation'),
        ('nothing', 'Do not create any document'),
    ], default='change_type', string='Action after fixed term contract comes to an end')
    fixed_term_contract_extension_by_months = fields.Integer(string='Number of months to extend fixed term contract by',
                                                             default=1, )
    print_invoice_partner_balance = fields.Selection([
        ('disabled', 'Išjungta'),
        ('enabled_all', 'Taikoma visiems klientams'),
        ('enabled_partial', 'Taikoma pasirinktiems klientams')],
        string='Spausdinti partnerio skolą/permoką sąskaitoje faktūroje',
        help="Pasirinkus opciją 'Taikoma pasirinktiems klientams', opciją pažymėti galite kliento kortelėje")
    force_need_action_repr = fields.Boolean(
        string='Užduoti reprezentacinių išlaidų tvirtinimo klausimus '
               'vadovui/atsakingam asmeniui kai sąskaitą įveda vartotojas')
    prevent_duplicate_product_code = fields.Boolean(string='Neleisti produkto kodų dublikatų', default=False)
    prevent_empty_product_code = fields.Boolean(string='Neleisti tuščių produkto kodų', default=False)
    allow_zero_allowance_business_trip = fields.Boolean(string='Leisti įvesti nulinius '
                                                               'dienpinigius komandiruočių el. dokumentuose')
    form_business_trip_payments_immediately_after_signing = fields.Boolean(string='Formuoti dienpinigių mokėjimo '
                                                                                  'pavedimus iš karto po komandiruotės '
                                                                                  'įsakymo pasirašymo')
    automatically_send_business_trip_allowance_payments_to_bank = fields.Boolean(
        string='Automatically send business trip allowance payments to bank',
    )
    form_gpm_line_with_holiday_payout = fields.Boolean(string='Formuoti GPM mokėjimą kartu su atostoginių išmokėjimu, '
                                                              'kai išmokama prieš atostogas')
    enable_product_uom = fields.Boolean(string='Įgalinti matavimo vienetus')
    enable_periodic_invoices = fields.Boolean(string='Įgalinti periodines sąskaitas')
    enable_periodic_front_statements = fields.Boolean(string='Įgalinti periodinius mokėjimo ruošinius')
    enable_cash_registers = fields.Boolean(string='Įgalinti kasos aparatus')
    enable_invoice_journal_selection = fields.Boolean(string='Įgalinti skirtingas sąskaitų numeruotes')
    enable_invoice_reconciliation_on_private_consumption = fields.Boolean(
        string='Įgalinti automatinį sąskaitų sudengimą tenkinant privačius poreikius',
    )
    enable_employment_requests_on_order_sign = fields.Boolean(string='Įgalinti automatinį prašymo dėl priėmimo '
                                                                     'į darbą kūrimą', default=True)
    enable_paypal_integration = fields.Boolean(string='Įgalinti Paypal integraciją')
    enable_revolut_integration = fields.Boolean(string='Įgalinti Revolut integraciją')
    enable_paysera_integration = fields.Boolean(string='Įgalinti Paysera integraciją')
    enable_e_banking_integration = fields.Boolean(string='Activate Enable banking integration')
    enable_braintree_integration = fields.Boolean(string='Enable Braintree integration')
    enable_seb_integration = fields.Boolean(string='Įgalinti SEB integraciją')
    invoice_print_only_foreign_currency = fields.Boolean(string='Sąskaitoje spausdinti tik užsienio valiuta')
    invoice_print_discount_type = fields.Selection([('perc', 'Procentais'), ('currency', 'Pinigine verte')],
                                                   string='Spausdinant sąskaitą rodyti nuolaidą')
    invoice_default_proforma_print = fields.Selection([('proforma', 'Išankstinė sąskaita'),
                                                       ('commercial_offer', 'Komercinis pasiūlymas')],
                                                      string='Kaip spausdinti išankstines sąskaitas')
    e_documents_enable_advance_setup = fields.Boolean(
        string='Įgalinti avanso nustatymus el. dokumentuose', default=False)
    auto_form_employee_advance_balance_document = fields.Boolean(
        string='Automatiškai formuoti avansinės apyskaitos dokumentą',
        help='Pažymėjus, kiekvieno mėnesio 21 dieną darbuotojams bus suformuojami avansinės apyskaitos dokumentai')
    show_robo_api_settings = fields.Boolean(string='Rodyti ROBO API nustatymus')
    activate_threaded_front_reports = fields.Boolean(
        string='Eksportuoti ataskaitas kaip foninę užduotį',
        help='Įmonėms turinčioms daugiau duomenų rekomenduojama įgalinti šią opciją, ataskaitos bus generuojamos fone')
    embed_einvoice_xml = fields.Boolean(
        string='Įterpti e-sąskaitos duomenis sąskaitų PDF failuose',
        help='Įgalinus - sąskaitų PDF dokumentuose bus įterpiamos standartizuotos e-sąskaitos. Gavėjams e-sąskaitos'
             'suteikia galimybę sąskaitas apdoroti automatiškai, be papildomų žmogaus veiksmų.',
    )
    show_machine_readable = fields.Boolean(
        string='Rodyti "Machine Readable" (e-sąskaitos) logotipą sąskaitos poraštėje',
        help='"Machine Readable" logotipas sąskaitų poraštėse parodo, kad sąskaita gali būti lengvai apdorota kitų '
             'sistemų, nes joje yra įkoduotas XML su visa reikalinga sąskaitos informacija.',
        default=True
    )
    require_2fa = fields.Boolean(string='Require 2FA', help='When enabled, all users must use 2FA for connecting')
    show_2fa_session_kill_warning = fields.Boolean(compute='_compute_show_2fa_session_kill_warning')
    custom_cash_receipt_header_enabled = fields.Boolean(string='Allow setting custom cash receipt headers')
    custom_invoice_color_text = fields.Char(string='Custom system invoice template text color')
    custom_invoice_color_details = fields.Char(string='Custom system invoice template detail color')
    custom_invoice_footer_enabled = fields.Boolean('Custom invoice footer')
    custom_invoice_footer = fields.Text('Footer to be shown on invoices')
    invoice_footer_preview = fields.Html('Footer preview', compute='_compute_invoice_footer_preview')
    bank_integrations_to_activate_html = fields.Html(readonly=True)
    bank_integrations_to_activate = fields.Boolean()
    fiscalyear_last_month = fields.Selection([(1, 'Sausis'), (2, 'Vasaris'), (3, 'Kovas'), (4, 'Balandis'),
                                              (5, 'Gegužė'), (6, 'Birželis'), (7, 'Liepa'), (8, 'Rugpjūtis'),
                                              (9, 'Rugsėjis'), (10, 'Spalis'), (11, 'Lapkritis'), (12, 'Gruodis')],
                                             string='Paskutinis fiskalinių metu mėnuo', default=12)
    fiscalyear_last_day = fields.Integer('Paskutinė fiskalinių metų diena', default=31)

    substitute_report_partner = fields.Many2one(
        'res.partner', string='Substitute reports partner',
        domain="[('is_company', '=', True)]",
        groups="robo_basic.group_robo_premium_accountant",
        help="If set, this partners' info will be used when forming all but GPM reports instead of company information"
    )
    holiday_accumulation_usage_policy = fields.Boolean(
        string='Use holiday accumulation and usage records when calculating holiday payments'
    )
    holiday_accumulation_usage_start_date = fields.Date(
        string='Date from which the company uses holiday accumulation and usage records',
        groups="robo_basic.group_robo_premium_accountant"
    )
    allow_use_ne_pvm_objektas = fields.Boolean(string='Let users set non-VAT objects')
    allow_accumulative_work_time_accounting_net_bonus_orders = fields.Boolean(
        string='Allow bonus orders with NET amounts specified for employees working by accumulative work time '
               'accounting',
        groups="robo_basic.group_robo_premium_accountant"
    )
    use_latest_product_price = fields.Boolean(string='Get the most recent price of a product on sale invoices')
    use_children_records_for_parental_leave_documents = fields.Boolean(
        string='Use children\'s records for monthly parental leave documents'
    )
    show_paid_invoice_state_on_printing = fields.Boolean(string='Show paid invoice state on printing')
    use_last_unit_price_of_account_invoice_line = fields.Boolean(
        string='Use the unit price of the last saved account invoice line'
    )

    @api.onchange('holiday_accumulation_usage_policy')
    def _onchange_holiday_accumulation_usage_policy_set_usage_start_date(self):
        if self.holiday_accumulation_usage_policy and not self.holiday_accumulation_usage_start_date:
            self.holiday_accumulation_usage_start_date = datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATE_FORMAT)

    @api.multi
    @api.depends('custom_invoice_footer', 'custom_invoice_footer_enabled', 'show_machine_readable',
                 'embed_einvoice_xml')
    def _compute_invoice_footer_preview(self):
        template = self.env.ref('saskaitos.invoice_footer', False)
        if self.env.user.is_manager() and template:
            for rec in self:
                company = rec.sudo().company_id or rec.company_id.sudo()
                custom_footer_enabled = company.custom_invoice_footer_enabled
                custom_footer = company.custom_invoice_footer
                show_machine_readable = company.show_machine_readable
                embed_einvoice_xml = company.embed_einvoice_xml
                company.custom_invoice_footer_enabled = rec.custom_invoice_footer_enabled
                company.custom_invoice_footer = rec.custom_invoice_footer
                company.show_machine_readable = rec.show_machine_readable
                company.embed_einvoice_xml = rec.embed_einvoice_xml
                rec.invoice_footer_preview = template.render({'company': company})
                company.custom_invoice_footer_enabled = custom_footer_enabled
                company.custom_invoice_footer = custom_footer
                company.show_machine_readable = show_machine_readable
                company.embed_einvoice_xml = embed_einvoice_xml

    @api.multi
    @api.constrains('sumine_apskaita_period_amount')
    def _check_sumine_apskaita_period_amount_is_correct(self):
        lower_bound = 1
        upper_bound = 4
        for rec in self:
            if not lower_bound <= rec.sudo().sumine_apskaita_period_amount <= upper_bound:
                raise exceptions.ValidationError(
                    _('Suminės apskaitos periodų skaičius privalo būti nuo %s iki %s')
                    % (str(lower_bound), str(upper_bound + 1))
                )

    @api.multi
    @api.constrains('invoice_cc_emails')
    def _check_invoice_cc_emails(self):
        self.ensure_one()
        if not self.invoice_cc_emails:
            return
        email_regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        for email in self.invoice_cc_emails.split(';'):
            if not email:
                continue
            if not re.match(email_regex, email.strip()):
                raise exceptions.ValidationError(
                    _('Carbon copy (CC) invoice receiver emails are not correct. Invalid email address: {}.\n'
                      'In case of multiple emails, separate them by semicolon signs (";").').format(email))

    @api.multi
    def btn_open_paypal_settings(self):
        paypal_apis = self.env['paypal.api'].search([])
        if len(paypal_apis) > 1:
            action = self.env.ref('robo.paypal_api_action').read()[0]
            action['domain'] = [('id', 'in', paypal_apis.ids)]
        else:
            action = {
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'paypal.api',
                'view_id': self.env.ref('robo.paypal_api_view_form').id,
                'type': 'ir.actions.act_window',
                'target': 'current',
            }
            if paypal_apis:
                action['res_id'] = paypal_apis.id
        return action

    @api.multi
    def btn_open_revolut_settings(self):
        revolut_apis = self.env['revolut.api'].search([])
        if len(revolut_apis) > 1:
            action = self.env.ref('robo.revolut_api_action').read()[0]
        else:
            action = {
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'revolut.api',
                'view_id': self.env.ref('robo.revolut_api_view_form').id,
                'type': 'ir.actions.act_window',
                'target': 'current',
            }
            if revolut_apis:
                action['res_id'] = revolut_apis.id
        return action

    @api.multi
    def btn_open_paysera_settings(self):
        """
        Open Paysera settings window. If settings object is not yet created, create one.
        :return: None
        """
        paysera_configuration = self.env['paysera.configuration'].initiate_settings()
        action = {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'paysera.configuration',
            'res_id': paysera_configuration.id,
            'view_id': self.env.ref('sepa.form_paysera_configuration').id,
            'type': 'ir.actions.act_window',
            'target': 'current',
        }
        return action

    @api.multi
    def btn_open_enable_banking_settings(self):
        """
        Open Enable Banking settings window. If settings object is not yet created, create one.
        :return: JS action (dict)
        """
        enable_banking_configuration = self.env['enable.banking.configuration'].initiate_settings()
        action = {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'enable.banking.configuration',
            'res_id': enable_banking_configuration.id,
            'view_id': self.env.ref('sepa.form_enable_banking_configuration').id,
            'type': 'ir.actions.act_window',
            'target': 'current',
        }
        return action

    @api.multi
    def btn_open_braintree_settings(self):
        """
        Open Braintree settings window. If settings object is not yet created, create one.
        :return: JS action (dict)
        """
        braintree_configuration = self.env['braintree.configuration'].get_configuration()
        action = {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'braintree.configuration',
            'res_id': braintree_configuration.id,
            'view_id': self.env.ref('sepa.form_braintree_configuration').id,
            'type': 'ir.actions.act_window',
            'target': 'current',
        }
        return action

    @api.multi
    def btn_open_seb_settings(self):
        """
        Open SEB settings window. If settings object is not yet created, create one.
        :return: None
        """
        seb_configuration = self.env['seb.configuration'].initiate_settings()
        # Always renew journal data on opening
        seb_configuration.renew_journal_data()
        action = {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'seb.configuration',
            'res_id': seb_configuration.id,
            'view_id': self.env.ref('sepa.form_seb_configuration').id,
            'type': 'ir.actions.act_window',
            'target': 'current',
        }
        return action

    @api.multi
    def save_numberings(self):
        if not self.env.user.is_manager():
            raise exceptions.UserError(_('Tik vadovas gali pakeisti šį nustatymą'))
        if self._context.get('invoice'):
            self._set_invoice_number()
        if self._context.get('credit_invoice'):
            self._set_credit_invoice_number()
        if self._context.get('proforma'):
            self._set_proforma_number()
        if self._context.get('in_payment'):
            self._set_in_payment_number()
        if self._context.get('out_payment'):
            self._set_out_payment_number()
        if self._context.get('in_cash'):
            self._set_in_cash_receipt_number()
        if self._context.get('out_cash'):
            self._set_out_cash_receipt_number()
        if self._context.get('contract'):
            self._set_contract_number()

    @api.one
    @api.depends('invoice_series', 'invoice_number', 'invoice_length')
    def _invoice_actual_number(self):
        self.invoice_actual_number = str(self.invoice_series or '') + str(self.invoice_number or 1).zfill(
            self.invoice_length)

    @api.one
    @api.depends('credit_invoice_series', 'credit_invoice_number', 'credit_invoice_length')
    def _compute_credit_invoice_actual_number(self):
        self.credit_invoice_actual_number = str(self.credit_invoice_series or '') + str(
            self.credit_invoice_number or 1).zfill(
            self.credit_invoice_length)

    @api.one
    @api.depends('proforma_series', 'proforma_number', 'proforma_length')
    def _proforma_actual_number(self):
        self.proforma_actual_number = str(self.proforma_series or '') + str(self.proforma_number or 1).zfill(
            self.proforma_length)

    @api.one
    @api.depends('in_payment_series', 'in_payment_number', 'in_payment_length')
    def _in_payment_actual_number(self):
        self.in_payment_actual_number = str(self.in_payment_series or '') + str(self.in_payment_number or 1).zfill(
            self.in_payment_length)

    @api.one
    @api.depends('out_payment_series', 'out_payment_number', 'out_payment_length')
    def _out_payment_actual_number(self):
        self.out_payment_actual_number = str(self.out_payment_series or '') + str(self.out_payment_number or 1).zfill(
            self.out_payment_length)

    @api.one
    @api.depends('in_cash_receipt_series', 'in_cash_receipt_number', 'in_cash_receipt_length')
    def _in_cash_receipt_actual_number(self):
        self.in_cash_receipt_actual_number = str(self.in_cash_receipt_series or '') + str(
            self.in_cash_receipt_number or 1).zfill(
            self.in_cash_receipt_length)

    @api.one
    @api.depends('out_cash_receipt_series', 'out_cash_receipt_number', 'out_cash_receipt_length')
    def _out_cash_receipt_actual_number(self):
        self.out_cash_receipt_actual_number = str(self.out_cash_receipt_series or '') + str(
            self.out_cash_receipt_number or 1).zfill(
            self.out_cash_receipt_length)

    @api.one
    @api.depends('contract_series', 'contract_number', 'contract_length')
    def _contract_actual_number(self):
        self.contract_actual_number = str(self.contract_series or '') + str(self.contract_number or 1).zfill(
            self.contract_length)

    @api.one
    @api.depends('require_2fa')
    def _compute_show_2fa_session_kill_warning(self):
        self.show_2fa_session_kill_warning = self.require_2fa and not self.company_id.require_2fa

    @api.onchange('invoice_journal_id')
    def _onchange_invoice_journal_id(self):
        if self.invoice_journal_id and self.invoice_journal_id.sequence_id:
            sequence = self.invoice_journal_id.sequence_id
            self.invoice_series = sequence.prefix
            self.invoice_length = sequence.padding
            self.invoice_number = sequence.number_next_actual

    @api.onchange('in_cash_journal_id')
    def _onchange_in_cash_journal_id(self):
        if self.in_cash_journal_id and self.in_cash_journal_id.sequence_id:
            sequence = self.in_cash_journal_id.sequence_id
            self.in_cash_receipt_series = sequence.prefix
            self.in_cash_receipt_length = sequence.padding
            self.in_cash_receipt_number = sequence.number_next_actual

    @api.onchange('company_draudejo_kodas')
    def onchange_company_draudejo_kodas(self):
        if self.company_draudejo_kodas:
            self.company_draudejo_kodas = self.company_draudejo_kodas.replace(' ', '')

    @api.model
    def get_payment_data(self):
        res = {}
        in_payment_sequence_code = 'account.payment.customer.invoice'
        in_payment_sequence = self.env['ir.sequence'].search([('code', '=', in_payment_sequence_code)], limit=1)
        if in_payment_sequence:
            series, length, number = in_payment_sequence.get_prefix_size_number()
            res.update({'in_payment_series': series,
                        'in_payment_length': length,
                        'in_payment_number': number,
                        })

        out_payment_sequence_code = 'account.payment.supplier.invoice'
        out_payment_sequence = self.env['ir.sequence'].search([('code', '=', out_payment_sequence_code)], limit=1)
        if out_payment_sequence:
            series, length, number = out_payment_sequence.get_prefix_size_number()
            res.update({'out_payment_series': series,
                        'out_payment_length': length,
                        'out_payment_number': number,
                        })

        in_cash_receipt_sequence_code = 'cash.receipt.inbound'
        in_cash_receipt_sequence = self.env['ir.sequence'].search([('code', '=', in_cash_receipt_sequence_code)],
                                                                  limit=1)
        if in_cash_receipt_sequence:
            series, length, number = in_cash_receipt_sequence.get_prefix_size_number()
            res.update({'in_cash_receipt_series': series,
                        'in_cash_receipt_length': length,
                        'in_cash_receipt_number': number,
                        })

        out_cash_receipt_sequence_code = 'cash.receipt.outbound'
        out_cash_receipt_sequence = self.env['ir.sequence'].search([('code', '=', out_cash_receipt_sequence_code)],
                                                                   limit=1)
        if out_cash_receipt_sequence:
            series, length, number = out_cash_receipt_sequence.get_prefix_size_number()
            res.update({'out_cash_receipt_series': series,
                        'out_cash_receipt_length': length,
                        'out_cash_receipt_number': number,
                        })

        proforma_sequence_code = 'account.invoice.proforma'
        proforma_sequence = self.env['ir.sequence'].search([('code', '=', proforma_sequence_code)],
                                                           limit=1)
        if proforma_sequence:
            series, length, number = proforma_sequence.get_prefix_size_number()
            res.update({'proforma_series': series,
                        'proforma_length': length,
                        'proforma_number': number,
                        })

        credit_invoice_code = 'KR'
        credit_invoice_sequence = self.env['ir.sequence'].search([('code', '=', credit_invoice_code)],
                                                                 limit=1)
        if credit_invoice_sequence:
            series, length, number = credit_invoice_sequence.get_prefix_size_number()
            res.update({'credit_invoice_series': series,
                        'credit_invoice_length': length,
                        'credit_invoice_number': number,
                        })

        return res

    @api.model
    def get_du_data(self):
        res = {}
        sequence_code = 'DU'
        sequence = self.env['ir.sequence'].search([('code', '=', sequence_code)], limit=1)
        if sequence:
            series, length, number = sequence.get_prefix_size_number()
            res.update({'contract_series': series,
                        'contract_length': length,
                        'contract_number': number,
                        })
        return res

    @api.model
    def default_get(self, field_list):
        if not self.env.user.is_manager():
            return {}
        company_id = self.sudo().env.user.company_id
        # bank_journal = self.sudo().env['account.journal'].search([('type', '=', 'bank'),
        #                                                           ('display_on_footer', '=', True)], limit=1)
        # company_bank_account = bank_journal.bank_acc_number
        # company_bank_name = bank_journal.bank_id.name if bank_journal.bank_id else ''
        # company_bank_code = bank_journal.bank_id.kodas if bank_journal.bank_id else ''
        # company_bank_bic = bank_journal.bank_id.bic if bank_journal.bank_id else ''
        # company_bank_currency = bank_journal.currency_id.id if bank_journal.currency_id else bank_journal.company_id.currency_id.id
        company_banks = []
        bank_journal_ids = self.sudo().env['account.journal'].search(
            [('type', '=', 'bank'), ('show_on_dashboard', '=', True)]).filtered(
            lambda b: b.bank_acc_number)
        for bank_journal_id in bank_journal_ids:
            company_bank_account = bank_journal_id.bank_acc_number
            company_bank_name = bank_journal_id.bank_id.name if bank_journal_id.bank_id else ''
            company_bank_code = bank_journal_id.bank_id.kodas if bank_journal_id.bank_id else ''
            company_bank_bic = bank_journal_id.bank_id.bic if bank_journal_id.bank_id else ''
            company_bank_currency = bank_journal_id.currency_id.id if bank_journal_id.currency_id else bank_journal_id.company_id.currency_id.id
            show_footer = bank_journal_id.display_on_footer
            company_banks.append((0, 0, {
                'journal_id': bank_journal_id.id,
                'company_bank_account': company_bank_account,
                'company_bank_name': company_bank_name,
                'company_bank_code': company_bank_code,
                'company_bank_bic': company_bank_bic,
                'company_bank_currency': company_bank_currency,
                'show_footer': show_footer,
            }))
        customer_journal = self.sudo().env['account.journal'].search([('type', '=', 'sale')], limit=1)
        asset_sequence = self.sudo().env['ir.sequence'].search([('code', '=', 'ASSETS')], limit=1)
        in_cash_journal = self.sudo().env['account.journal'].search([('type', '=', 'cash')], limit=1)
        if asset_sequence:
            asset_prefix = asset_sequence.prefix
            asset_suffix = asset_sequence.suffix
            asset_padding = asset_sequence.padding
            asset_next_number = asset_sequence.number_next_actual
        else:
            asset_prefix = asset_suffix = ''
            asset_padding = asset_next_number = 0
        if customer_journal and customer_journal.sequence_id:
            invoice_prefix = customer_journal.sequence_id.prefix
            invoice_size = customer_journal.sequence_id.padding
            invoice_number_next = customer_journal.sequence_id.number_next_actual
            invoice_journal_id = customer_journal.id
        elif not customer_journal and not customer_journal.sequence_id:
            invoice_prefix = 'KL'
            invoice_size = 4
            invoice_number_next = 1
            invoice_journal_id = False
        if in_cash_journal and in_cash_journal.sequence_id:
            in_cash_prefix = in_cash_journal.sequence_id.prefix
            in_cash_size = in_cash_journal.sequence_id.padding
            in_cash_number_next = in_cash_journal.sequence_id.number_next_actual
            in_cash_journal_id = in_cash_journal.id
        elif not in_cash_journal and not in_cash_journal.sequence_id:
            in_cash_prefix = 'PPK'
            in_cash_size = 5
            in_cash_number_next = 1
            in_cash_journal_id = False
        module_obj = self.sudo().env['ir.module.module']
        robo_api_installed = bool(module_obj.search(
            [('name', '=', 'robo_api'), ('state', 'in', ['installed', 'to upgrade'])], count=True))
        neopay_installed = bool(module_obj.search(
            [('name', '=', 'neopay'), ('state', 'in', ['installed', 'to upgrade'])], count=True))
        bank_integration_label = self.env['api.bank.integrations'].check_for_integrations_to_activate()
        res = {
            'bank_integrations_to_activate': bank_integration_label != 'disabled',
            'bank_integrations_to_activate_html': bank_integration_label,
            'company_id': company_id.id,
            'company_logo': company_id.partner_id.image_medium,
            'company_name': company_id.name,
            'company_ceo': company_id.vadovas.id if company_id.vadovas else False,
            'company_kodas': company_id.company_registry,
            'company_vat': company_id.vat,
            'company_street': company_id.street,
            'company_street2': company_id.street2,
            'company_city': company_id.city,
            'invoice_vat_printing_method': company_id.invoice_vat_printing_method,
            'company_zip': company_id.zip,
            'company_draudejo_kodas': company_id.draudejo_kodas,
            'company_savivaldybe': company_id.savivaldybe,
            'company_evrk': company_id.evrk.id if company_id.evrk else False,
            'company_country_id': company_id.country_id.id if company_id.country_id else False,
            'default_payment_term_days': company_id.default_payment_term_id.num_days,
            'isaf_default_day': company_id.isaf_default_day,
            'default_supplier_payment_term_days': company_id.default_supplier_payment_term_id.num_days,
            # 'company_bank_account': company_bank_account,
            # 'company_bank_name': company_bank_name,
            # 'company_bank_code': company_bank_code,
            # 'company_bank_bic': company_bank_bic,
            # 'company_bank_currency': company_bank_currency,
            'company_bank_accounts': company_banks,
            'company_email': company_id.email,
            'company_phone': company_id.phone,
            'company_website': company_id.website,
            'company_fax': company_id.fax,
            'invoice_series': invoice_prefix,
            'invoice_length': invoice_size,
            'invoice_number': invoice_number_next,
            'in_cash_receipt_series': in_cash_prefix,
            'in_cash_receipt_length': in_cash_size,
            'in_cash_receipt_number': in_cash_number_next,
            'invoice_default_proforma_print': company_id.invoice_default_proforma_print,
            'asset_prefix': asset_prefix,
            'asset_suffix': asset_suffix,
            'asset_padding': asset_padding,
            'asset_next_number': asset_next_number,
            'invoice_journal_id': invoice_journal_id,
            'in_cash_journal_id': in_cash_journal_id,
            'default_msg_receivers': [(6, 0, company_id.default_msg_receivers.ids)],
            'payroll_bank_journal_id': company_id.payroll_bank_journal_id.id,
            'force_accounting_date': company_id.force_accounting_date,
            'e_documents_allow_historic_signing': company_id.e_documents_allow_historic_signing,
            'print_invoice_presenter': company_id.print_invoice_presenter,
            'print_invoice_partner_balance': company_id.print_invoice_partner_balance,
            'invoice_cc_emails': company_id.invoice_cc_emails,
            'force_need_action_repr': company_id.force_need_action_repr,
            'e_documents_allow_historic_signing_spec': company_id.e_documents_allow_historic_signing_spec,
            'sumine_apskaita_period_amount': company_id.sumine_apskaita_period_amount,
            'sumine_apskaita_period_start': company_id.sumine_apskaita_period_start,
            'substitute_report_partner': company_id.substitute_report_partner.id,
            'prevent_duplicate_product_code': company_id.prevent_duplicate_product_code,
            'prevent_empty_product_code': company_id.prevent_empty_product_code,
            'allow_zero_allowance_business_trip': company_id.allow_zero_allowance_business_trip,
            'form_business_trip_payments_immediately_after_signing': company_id.form_business_trip_payments_immediately_after_signing,
            'form_gpm_line_with_holiday_payout': company_id.form_gpm_line_with_holiday_payout,
            'enable_product_uom': company_id.enable_product_uom,
            'enable_periodic_invoices': company_id.enable_periodic_invoices,
            'enable_periodic_front_statements': company_id.enable_periodic_front_statements,
            'enable_cash_registers': company_id.enable_cash_registers,
            'enable_invoice_journal_selection': company_id.enable_invoice_journal_selection,
            'enable_invoice_reconciliation_on_private_consumption': company_id.enable_invoice_reconciliation_on_private_consumption,
            'enable_employment_requests_on_order_sign': company_id.enable_employment_requests_on_order_sign,
            'invoice_print_only_foreign_currency': company_id.invoice_print_only_foreign_currency,
            'invoice_print_discount_type': company_id.invoice_print_discount_type,
            'enable_paypal_integration': company_id.enable_paypal_integration,
            'enable_revolut_integration': company_id.enable_revolut_integration,
            'enable_paysera_integration': company_id.enable_paysera_integration,
            'enable_e_banking_integration': company_id.enable_e_banking_integration,
            'enable_braintree_integration': company_id.enable_braintree_integration,
            'enable_seb_integration': company_id.enable_seb_integration,
            'module_neopay': neopay_installed,
            'e_documents_enable_advance_setup': company_id.e_documents_enable_advance_setup,
            'auto_form_employee_advance_balance_document': company_id.auto_form_employee_advance_balance_document,
            'activate_threaded_front_reports': company_id.activate_threaded_front_reports,
            'company_invoice_text': company_id.company_invoice_text,
            'company_proforma_invoice_text_different_from_regular': company_id.company_proforma_invoice_text_different_from_regular,
            'company_invoice_proforma_text': company_id.company_invoice_proforma_text,
            'show_robo_api_settings': robo_api_installed,
            'embed_einvoice_xml': company_id.embed_einvoice_xml,
            'show_machine_readable': company_id.show_machine_readable,
            'custom_invoice_color_text': company_id.custom_invoice_color_text,
            'custom_invoice_color_details': company_id.custom_invoice_color_details,
            'custom_invoice_footer_enabled': company_id.custom_invoice_footer_enabled,
            'custom_invoice_footer': company_id.custom_invoice_footer,
            'require_2fa': company_id.require_2fa,
            'holiday_accumulation_usage_policy': company_id.holiday_accumulation_usage_policy,
            'holiday_accumulation_usage_start_date': company_id.holiday_accumulation_usage_start_date,
            'allow_accumulative_work_time_accounting_net_bonus_orders': company_id.allow_accumulative_work_time_accounting_net_bonus_orders,
            'use_latest_product_price': company_id.use_latest_product_price,
            'default_action_after_fixed_term_contract_end': company_id.default_action_after_fixed_term_contract_end,
            'fixed_term_contract_extension_by_months': company_id.fixed_term_contract_extension_by_months,
            'custom_cash_receipt_header_enabled': company_id.custom_cash_receipt_header_enabled,
            'automatically_send_business_trip_allowance_payments_to_bank': company_id.automatically_send_business_trip_allowance_payments_to_bank,
            'use_children_records_for_parental_leave_documents': company_id.use_children_records_for_parental_leave_documents,
            'show_paid_invoice_state_on_printing': company_id.show_paid_invoice_state_on_printing,
            'use_last_unit_price_of_account_invoice_line': company_id.use_last_unit_price_of_account_invoice_line,
        }
        res.update(self.get_payment_data())
        res.update(self.get_du_data())
        res['longterm_assets_min_val'] = company_id.sudo().longterm_assets_min_val
        res['longterm_non_material_assets_min_val'] = company_id.sudo().longterm_non_material_assets_min_val
        if self.env.user.is_accountant():
            res['module_robo_stock'] = bool(
                module_obj.search([('name', '=', 'robo_stock'), ('state', '=', 'installed')], count=True))
            res['module_work_schedule'] = company_id.module_work_schedule
            res['module_work_schedule_analytics'] = company_id.module_work_schedule_analytics
            res['module_robo_analytic'] = bool(
                module_obj.search([('name', '=', 'robo_analytic'), ('state', '=', 'installed')], count=True))
            res['module_robo_mrp'] = bool(
                module_obj.search([('name', '=', 'robo_mrp'), ('state', '=', 'installed')], count=True))
            res['module_robo_api'] = robo_api_installed
            res['module_mixed_vat_rate'] = bool(
                module_obj.search([('name', '=', 'mixed_vat_rate'), ('state', '=', 'installed')], count=True))
            res['module_amazon_integration'] = bool(
                module_obj.search([('name', '=', 'amazon_integration'), ('state', '=', 'installed')], count=True))
            res['politika_atostoginiai'] = company_id.politika_atostoginiai
            res['politika_neatvykimai'] = company_id.politika_neatvykimai
            res['accumulated_days_policy'] = company_id.accumulated_days_policy
            res['worker_policy'] = company_id.worker_policy
            res['company_activity_form'] = company_id.company_activity_form
            res['automatic_salary_reconciliation'] = company_id.automatic_salary_reconciliation
            res['uab_report_size'] = company_id.uab_report_size
            res['additional_analytic_actions'] = company_id.additional_analytic_actions
            res['required_du_analytic'] = company_id.required_du_analytic
            res['change_analytic_on_accountant_validated'] = company_id.change_analytic_on_accountant_validated
            res['politika_atostogu_suteikimas'] = company_id.politika_atostogu_suteikimas
            res['holiday_policy_inform_manager'] = company_id.holiday_policy_inform_manager
            res['inv_due_date_edit'] = company_id.inv_due_date_edit
            res['proforma_show_price_vat_incl'] = company_id.proforma_show_price_vat_incl
            res['advance_payment_day'] = company_id.advance_payment_day
            res['fiscalyear_last_month'] = company_id.fiscalyear_last_month
            res['fiscalyear_last_day'] = company_id.fiscalyear_last_day
        res['allow_use_ne_pvm_objektas'] = company_id.allow_use_ne_pvm_objektas
        res['apr_send_reminders'] = company_id.apr_send_reminders
        res['apr_enabled_by_default'] = company_id.apr_enabled_by_default
        res['apr_send_before'] = company_id.apr_send_before
        res['apr_send_before_ndays'] = company_id.apr_send_before_ndays
        res['apr_send_on_date'] = company_id.apr_send_on_date
        res['apr_send_after'] = company_id.apr_send_after
        res['apr_send_after_ndays'] = company_id.apr_send_after_ndays
        res['apr_min_amount_to_send'] = company_id.apr_min_amount_to_send
        res['apr_email_cc'] = company_id.apr_email_cc
        res['apr_email_reply_to'] = company_id.apr_email_reply_to
        res['analytic_account_id'] = company_id.analytic_account_id.id
        return res

    def _update_company_fields(self, field_list, field_map=None):
        """ Update the data on res.company as needed """
        self.ensure_one()
        if field_map is None:
            field_map = {}
        company = self.company_id
        update_vals = {}
        for f in field_list:
            m = field_map.get(f)
            if m is None:
                val = self[f]
            elif callable(m):
                val = m()
            else:
                val = m
            if isinstance(company._fields.get(f), fields.Many2one):
                if not isinstance(val, int):
                    val = val.id
                if company[f].id != val:
                    update_vals[f] = val
            else:
                if company[f] != val:
                    update_vals[f] = val
        if update_vals:
            company.write(update_vals)

    @api.model
    def _get_company_policy_field_list(self):
        """ Return the list of fields on res.company that should be updated on set_policy """
        return [
            'isaf_default_day',
            'politika_atostoginiai',
            'module_work_schedule',
            'module_work_schedule_analytics',
            'module_robo_analytic',
            'module_robo_api',
            'politika_neatvykimai',
            'accumulated_days_policy',
            'worker_policy',
            'company_activity_form',
            'automatic_salary_reconciliation',
            'uab_report_size',
            'additional_analytic_actions',
            'required_du_analytic',
            'change_analytic_on_accountant_validated',
            'politika_atostogu_suteikimas',
            'holiday_policy_inform_manager',
            'inv_due_date_edit',
            'proforma_show_price_vat_incl',
            'force_accounting_date',
            'longterm_assets_min_val',
            'longterm_non_material_assets_min_val',
            'e_documents_allow_historic_signing',
            'print_invoice_partner_balance',
            'force_need_action_repr',
            'e_documents_allow_historic_signing_spec',
            'sumine_apskaita_period_amount',
            'sumine_apskaita_period_start',
            'substitute_report_partner',
            'advance_payment_day',
            'invoice_print_discount_type',
            'enable_paypal_integration',
            'enable_revolut_integration',
            'enable_paysera_integration',
            'enable_e_banking_integration',
            'enable_braintree_integration',
            'enable_seb_integration',
            'fiscalyear_last_month',
            'fiscalyear_last_day',
            'holiday_accumulation_usage_policy',
            'holiday_accumulation_usage_start_date',
            'allow_use_ne_pvm_objektas',
        ]

    @api.multi
    def _get_company_policy_field_map(self):
        """
        Return a mapping for field on res.company when the robo.company.settings is not matching
        :returns: dict with res.company field name as key, and matching value for that field.
        """
        return {}

    @api.model
    def _get_company_info_field_list(self):
        """ Return the list of fields on res.company that should be updated on set_company_info"""
        return [
            'name',
            'vadovas',
            'vat',
            'street',
            'street2',
            'city',
            'invoice_vat_printing_method',
            'zip',
            'country_id',
            'draudejo_kodas',
            'savivaldybe',
            'evrk',
            'phone',
            'email',
            'fax',
            'website',
            'default_payment_term_id',
            'default_supplier_payment_term_id',
            'payroll_bank_journal_id',
            'advance_journal_id',
            'apr_send_reminders',
            'apr_enabled_by_default',
            'apr_send_before',
            'apr_send_before_ndays',
            'apr_send_on_date',
            'apr_send_after',
            'apr_send_after_ndays',
            'apr_min_amount_to_send',
            'apr_email_cc',
            'apr_email_reply_to',
            'prevent_duplicate_product_code',
            'prevent_empty_product_code',
            'allow_zero_allowance_business_trip',
            'form_business_trip_payments_immediately_after_signing',
            'enable_product_uom',
            'enable_periodic_invoices',
            'enable_periodic_front_statements',
            'enable_cash_registers',
            'enable_invoice_journal_selection',
            'enable_invoice_reconciliation_on_private_consumption',
            'enable_employment_requests_on_order_sign',
            'form_gpm_line_with_holiday_payout',
            'e_documents_enable_advance_setup',
            'auto_form_employee_advance_balance_document',
            'activate_threaded_front_reports',
            'company_invoice_text',
            'print_invoice_presenter',
            'company_proforma_invoice_text_different_from_regular',
            'company_invoice_proforma_text',
            'embed_einvoice_xml',
            'show_machine_readable',
            'custom_cash_receipt_header_enabled',
            'custom_invoice_color_text',
            'custom_invoice_color_details',
            'custom_invoice_footer_enabled',
            'custom_invoice_footer',
            'require_2fa',
            'analytic_account_id',
            'invoice_print_only_foreign_currency',
            'invoice_default_proforma_print',
            'invoice_cc_emails',
            'allow_accumulative_work_time_accounting_net_bonus_orders',
            'use_latest_product_price',
            'default_action_after_fixed_term_contract_end',
            'fixed_term_contract_extension_by_months',
            'automatically_send_business_trip_allowance_payments_to_bank',
            'use_children_records_for_parental_leave_documents',
            'show_paid_invoice_state_on_printing',
            'use_last_unit_price_of_account_invoice_line',
        ]

    @api.multi
    def _get_company_info_field_map(self):
        return {
            'name': self.company_name,
            'vadovas': self.company_ceo and self.company_ceo.id or False,
            'vat': self.company_vat,
            'street': self.company_street,
            'street2': self.company_street2,
            'city': self.company_city,
            'zip': self.company_zip,
            'country_id': self.company_country_id and self.company_country_id.id or False,
            'draudejo_kodas': self.company_draudejo_kodas,
            'savivaldybe': self.company_savivaldybe,
            'evrk': self.company_evrk.id if self.company_evrk else False,
            'phone': self.company_phone,
            'email': self.company_email,
            'fax': self.company_fax,
            'website': self.company_website,
            'default_payment_term_id': self.env['account.payment.term'].get_or_create_payment_term_by_days(
                self.default_payment_term_days),
            'default_supplier_payment_term_id': self.env['account.payment.term'].get_or_create_payment_term_by_days(
                self.default_supplier_payment_term_days),
            'advance_journal_id': self.payroll_bank_journal_id,
        }

    @api.multi
    def set_policy(self):
        self.ensure_one()
        if not self.env.user.is_accountant():
            return
        if self.politika_atostogu_suteikimas == 'department':
            department_ids = self.env['hr.employee'].search([('department_id.manager_id', '=', False)], count=True)
            if department_ids:
                raise exceptions.UserError(_('Ne visi padaliniai turi nustatytus vadovus.'))
        if self.set_default_payment_term_to_all_customers:
            term_id = self.env['account.payment.term'].get_or_create_payment_term_by_days(
                self.default_payment_term_days
            )
            customers = self.env['res.partner'].search([('customer', '=', True)])
            customers.write({'property_payment_term_id': term_id})
        self.sudo()._update_company_fields(self._get_company_policy_field_list(), self._get_company_policy_field_map())

    @api.multi
    def _set_invoice_number(self):
        self.ensure_one()
        if self.env.user.has_group('robo_basic.group_robo_select_invoice_journal'):
            if self.invoice_journal_id:
                customer_journal = self.invoice_journal_id.sudo()
            else:
                raise exceptions.ValidationError(_('Turite nurodyti žurnalą'))
        else:
            customer_journal = self.sudo().env['account.journal'].search([('type', '=', 'sale')], limit=1)
        sequence_id = customer_journal.sequence_id
        if sequence_id:
            example_number = str(self.invoice_series or '') + str(self.invoice_number or 1).zfill(self.invoice_length)
            if self.sudo().env['account.invoice'].search_count([('number', '=', example_number)]):
                raise exceptions.UserError(_('Sąskaita tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.invoice_series,
                'padding': self.invoice_length,
                'number_next_actual': self.invoice_number
            })

    @api.multi
    def _set_credit_invoice_number(self):
        self.ensure_one()
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', 'KR')])
        if sequence_id:
            if self.sudo().env['account.invoice'].search_count([('number', '=', self.credit_invoice_actual_number)]):
                raise exceptions.UserError(_('Kreditinė sąskaita tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.credit_invoice_series,
                'padding': self.credit_invoice_length,
                'number_next_actual': self.credit_invoice_number
            })

    @api.multi
    def _set_proforma_number(self):
        self.ensure_one()
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', 'account.invoice.proforma')])
        if sequence_id:
            if self.sudo().env['account.invoice'].search_count([('proforma_number', '=', self.proforma_actual_number)]):
                raise exceptions.UserError(_('Išankstinė sąskaita tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.proforma_series,
                'padding': self.proforma_length,
                'number_next_actual': self.proforma_number
            })

    @api.multi
    def _set_in_payment_number(self):
        self.ensure_one()
        sequence_code = 'account.payment.customer.invoice'
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', sequence_code)])
        if sequence_id:
            example_number = str(self.in_payment_series or '') + str(self.in_payment_number or 1).zfill(
                self.in_payment_length)
            if self.sudo().env['account.payment'].search_count(
                    [('payment_type', '=', 'inbound'), ('name', '=', example_number)]):
                raise exceptions.UserError(_('Kasos pajamų orderis tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.in_payment_series,
                'padding': self.in_payment_length,
                'number_next_actual': self.in_payment_number
            })

    @api.multi
    def _set_contract_number(self):
        self.ensure_one()
        sequence_code = 'DU'
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', sequence_code)])
        if sequence_id:
            example_number = str(self.contract_series or '') + str(self.contract_number or 1).zfill(
                self.contract_length)
            if self.sudo().env['hr.contract'].search_count([('name', '=', example_number)]) or \
                    self.sudo().env['hr.contract.appointment'].search_count([('name', '=', example_number)]):
                raise exceptions.UserError(_('Darbo sutartis/priedas su šiuo numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.contract_series,
                'padding': self.contract_length,
                'number_next_actual': self.contract_number
            })

    @api.multi
    def _set_out_payment_number(self):
        self.ensure_one()
        sequence_code = 'account.payment.supplier.invoice'
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', sequence_code)])
        if sequence_id:
            example_number = str(self.out_payment_series or '') + str(self.out_payment_number or 1).zfill(
                self.out_payment_length)
            if self.sudo().env['account.payment'].search_count(
                    [('payment_type', '=', 'outbound'), ('name', '=', example_number)]):
                raise exceptions.UserError(_('Kasos išlaidų orderis tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.out_payment_series,
                'padding': self.out_payment_length,
                'number_next_actual': self.out_payment_number
            })

    @api.multi
    def _set_in_cash_receipt_number(self):
        self.ensure_one()
        if self.in_cash_journal_id:
            customer_journal = self.in_cash_journal_id.sudo()
        else:
            customer_journal = self.sudo().env['account.journal'].search([('type', '=', 'cash')], limit=1)
        sequence_id = customer_journal.sequence_id
        if sequence_id:
            example_number = str(self.in_cash_receipt_series or '') + str(self.in_cash_receipt_number or 1).zfill(
                self.in_cash_receipt_length)
            if self.sudo().env['cash.receipt'].search_count(
                    [('payment_type', '=', 'inbound'), ('name', '=', example_number)]):
                raise exceptions.UserError(_('Grynųjų pinigų žurnalas tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.in_cash_receipt_series,
                'padding': self.in_cash_receipt_length,
                'number_next_actual': self.in_cash_receipt_number
            })

    @api.multi
    def _set_out_cash_receipt_number(self):
        self.ensure_one()
        sequence_code = 'cash.receipt.outbound'
        sequence_id = self.env['ir.sequence'].sudo().search([('code', '=', sequence_code)])
        if sequence_id:
            example_number = str(self.out_cash_receipt_series or '') + str(self.out_cash_receipt_number or 1).zfill(
                self.out_cash_receipt_length)
            if self.sudo().env['cash.receipt'].search_count(
                    [('payment_type', '=', 'outbound'), ('name', '=', example_number)]):
                raise exceptions.UserError(_('Pinigų išdavimo kvitas tokiu numeriu jau egzistuoja.'))
            sequence_id.write({
                'prefix': self.out_cash_receipt_series,
                'padding': self.out_cash_receipt_length,
                'number_next_actual': self.out_cash_receipt_number
            })

    @api.multi
    def _get_central_server_update_needs(self):
        self.ensure_one()
        company = self.company_id
        vals = {}
        if company.name != self.company_name:
            vals['company_name'] = self.company_name
        if company.vat != self.company_vat:
            vals['company_vat'] = self.company_vat
        if company.evrk != self.company_evrk:
            vals['evrk_code'] = self.company_evrk.code
        if company.draudejo_kodas != self.company_draudejo_kodas:
            vals['company_draudejo_kodas'] = self.company_draudejo_kodas
        return vals

    @api.multi
    def set_company_info(self):
        self.ensure_one()
        if not self.env.user.is_manager():
            return False
        company_id = self.sudo().company_id
        central_server_update_vals = self._get_central_server_update_needs()
        company_logo_md5_old = hashlib.md5(company_id.partner_id.image_medium.decode(
            'base64')).hexdigest() if company_id.partner_id.image_medium else None
        company_logo_md5_new = hashlib.md5(
            self.company_logo.decode('base64')).hexdigest() if self.company_logo else None
        if company_logo_md5_old != company_logo_md5_new:
            company_id.write({'logo': self.company_logo})
        if self.apr_send_reminders and self.apply_default_apr_settings_to_all:
            partners = self.sudo().env['res.partner'].search([('company_id', '=', company_id.id)])
            partners.write({'apr_send_reminders': True})
            partners.set_default_send_reminders_settings()
        self.sudo()._update_company_fields(self._get_company_info_field_list(), self._get_company_info_field_map())
        # Upload info to central server if needed
        if central_server_update_vals:
            try:
                internal = self.env['res.company']._get_odoorpc_object()
                central_server_update_vals['dbname'] = self.env.cr.dbname
                central_server_update_vals['company_kodas'] = self.company_kodas  #used only to find record
                internal.env['project.issue'].update_company_info(**central_server_update_vals)
            except:
                pass

    @api.one
    def inverse_set_default_bank_account(self):
        if not self.env.user.is_manager():
            return
        journal_obj = self.env['account.journal'].sudo()
        company_id = self.env.user.sudo().company_id
        default_bank_journal = self.env['account.journal'].sudo().search([
            ('type', '=', 'bank'),
            ('code', '=', 'BNK1')
        ]).filtered(lambda b: not b.bank_acc_number and not b.bank_id and b.company_id.id == company_id.id)
        default_being_used = False if default_bank_journal else True
        for bank in self.sudo().company_bank_accounts:
            bank_journal_id = False
            if not bank.company_bank_account:
                continue
            if bank.journal_id:
                bank_journal_id = bank.journal_id
            else:
                if bank.company_bank_account and bank.company_bank_currency:
                    if bank.company_bank_currency.id == company_id.currency_id.id:
                        currency_id = False
                    else:
                        currency_id = bank.company_bank_currency.id
                    bank_journal_id = journal_obj.search([('bank_acc_number', '=', bank.company_bank_account),
                                                          ('currency_id', '=', currency_id)], limit=1)
            if not bank_journal_id:
                if not default_being_used:
                    bank_journal_id = default_bank_journal
                    bank_journal_id.write({
                        'name': (bank.company_bank_name or 'Bankas') + ' (' + bank.company_bank_account[
                                                                              -4:] + ') ' + bank.company_bank_currency.name,
                        'display_on_footer': bank.show_footer,
                        'currency_id': bank.company_bank_currency.id,
                    })
                    default_being_used = True
                else:
                    code_base = 'BNK'
                    code = 'BNK'
                    for i in range(1, 100):
                        code = (code_base + str(i))[-5:]
                        journal = self.sudo().env['account.journal'].search([('code', '=', code)])
                        if not journal:
                            break
                    bank_journal_id = self.sudo().env['account.journal'].create({
                        'name': (bank.company_bank_name or 'Bankas') + ' (' + bank.company_bank_account[
                                                                              -4:] + ') ' + bank.company_bank_currency.name,
                        'code': code,
                        'bank_statements_source': 'file_import',
                        'display_on_footer': bank.show_footer,
                        'type': 'bank',
                        'company_id': company_id.id,
                        'currency_id': currency_id,
                    })
            # Search bank
            bank_id = False
            if bank.company_bank_bic:
                bank_id = self.sudo().env['res.bank'].search([('bic', '=', bank.company_bank_bic.strip())])
                if len(bank_id) > 1:
                    # Multiple banks with same BIC and different codes (kodas) might exist. If a bank with a code
                    # different from company_bank_code is found, setting company_bank_code on the bank violates the SQL
                    # constraint where the bank code (kodas) should be unique.
                    bank_with_code = bank_id.filtered(lambda b: b.kodas == bank.company_bank_code)
                    bank_id = bank_with_code[0] if bank_with_code else bank_id[0]
            if not bank_id and bank.company_bank_code:
                bank_id = self.sudo().env['res.bank'].search([('kodas', '=', bank.company_bank_code.strip())], limit=1)
            if not bank_id and bank.company_bank_name:
                bank_id = self.sudo().env['res.bank'].search([('name', '=', bank.company_bank_name.strip())], limit=1)
            if bank_id:
                bank_id.name = bank.company_bank_name or 'Bankas'
                bank_id.kodas = bank.company_bank_code or ''
                bank_id.bic = bank.company_bank_bic or ''
            else:
                bank_id = self.sudo().env['res.bank'].create({
                    'name': bank.company_bank_name or 'Bankas',
                    'kodas': bank.company_bank_code or '',
                    'bic': bank.company_bank_bic or '',
                })
            # Update info if needed
            update_vals = {}
            if bank_journal_id.bank_acc_number != bank.company_bank_account:
                update_vals['bank_acc_number'] = bank.company_bank_account
            if bank_journal_id.bank_id != bank_id:
                update_vals['bank_id'] = bank_id.id
            if bank_journal_id.display_on_footer != bank.show_footer:
                update_vals['display_on_footer'] = bank.show_footer
            if update_vals:
                bank_journal_id.write(update_vals)
                # Check for connectors to activate
                self.env['enable.banking.connector'].activate_connectors(journal=bank_journal_id)

    @api.multi
    def set_default_message_receivers(self):
        self.ensure_one()
        if self.env.user.is_accountant():
            company = self.company_id
            company.sudo().write({'default_msg_receivers': [(6, 0, self.sudo().default_msg_receivers.ids)], })

    @api.multi
    def set_default_import(self):
        """
        Calls threaded import preparation method
        on all possible front XLS files.
        :return: None
        """
        self.threaded_import_prep('import_partners')
        self.threaded_import_prep('import_customer_invoices')
        self.threaded_import_prep('import_supplier_invoices')
        self.threaded_import_prep('import_products')
        self.threaded_import_prep('import_financials')
        self.threaded_import_prep('import_employees')
        self.threaded_import_prep('import_aml')
        self.threaded_import_prep('import_du')

    @api.multi
    def threaded_import_prep(self, action, function=None, imported_file=None):
        """
        Prepares system for threaded XLS data import,
        checks whether any job of the same type is running,
        creates related job record and starts the thread.
        :param action: System-like name of the action that is to-be executed (str)
        :param function: function that will be used to process the file (function)
        :param imported_file: Imported, to-process file (str)
        :return: None
        """

        # Check whether file that we're trying to import is passed
        imported_file = getattr(self, action) if imported_file is None else imported_file
        if not imported_file:
            return
        now_dt = datetime.now(tz=timezone('Europe/Vilnius'))
        now_str = now_dt.strftime("%H:%M")
        if now_str > "20:20" and now_str < "20:40":
            raise exceptions.UserError(_("Negalima importuoti duomenų tarp {}:20 ir {}:40, prašome palaukti.").format(now_dt.hour, now_dt.hour))


        import_obj = self.env['robo.import.job']
        # Check whether there are any jobs of the same type that are being imported
        if import_obj.search_count([('state', '=', 'in_progress'), ('action', '=', action)]):
            raise exceptions.UserError(
                _('Šiuo metu ataskaita yra perkraunama, pabandykite po kelių minučių.'))

        # Get the function -- Either take it from local scope or args
        function = sys.modules[__name__].__dict__[action] if function is None else function

        vals = {
            'action': action,
            'execution_start_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
            'state': 'in_progress',
            'imported_file': imported_file,
            'user_id': self.env.user.id,
        }
        # Create import job record
        import_job = import_obj.create(vals)
        self.env.cr.commit()

        # Start the thread
        threaded_calculation = threading.Thread(
            target=import_obj.execute_threaded_import,
            args=(self.id, import_job.id, function, imported_file)
        )
        threaded_calculation.start()

    @api.multi
    def change_module_state(self, module_name, action):
        """
        Installs or uninstalls module based on the
        name and action that is passed to the method
        :return: refreshed self.env
        """
        self.ensure_one()
        module_obj = self.sudo().env['ir.module.module']

        # Check whether module should be installed or uninstalled
        if action == 'install':
            module_obj.update_list()
            module_obj.search([('name', '=', module_name)]).button_immediate_install()
        elif action == 'uninstall':
            module_obj.sudo().search([('name', '=', module_name)]).button_immediate_uninstall()
        self.env.reset()
        current_id = self.id

        # Need to re-browse, because this newly returned env is empty set,
        # and if two modules are being installed on the method, it will crash
        return self.env()[self._name].browse(current_id)

    @api.multi
    def execute(self):
        if not self.env.user.is_manager():
            return self.env.ref('robo.open_robo_vadovas_client_action').read()[0]
        self.ensure_one()
        module_obj = self.sudo().env['ir.module.module']
        self.set_company_info()
        self.inverse_set_default_bank_account()
        self.set_default_message_receivers()
        if self.env.user.is_accountant():
            self.set_policy()

            is_system = self.env.user.has_group('base.group_system')

            # Check whether modules that can be manipulated by accountant are installed or not
            robo_stock_installed = module_obj.search_count(
                [('name', '=', 'robo_stock'), ('state', '=', 'installed')])
            work_schedule_installed = module_obj.search_count(
                [('name', '=', 'work_schedule'), ('state', '=', 'installed')])
            work_schedule_analytics_installed = module_obj.search_count(
                [('name', '=', 'work_schedule_analytics'), ('state', '=', 'installed')])
            robo_mrp_installed = module_obj.search_count(
                [('name', '=', 'robo_mrp'), ('state', '=', 'installed')])
            robo_analytic_installed = module_obj.search_count(
                [('name', '=', 'robo_analytic'), ('state', '=', 'installed')])
            robo_api_installed = module_obj.search_count(
                [('name', '=', 'robo_api'), ('state', '=', 'installed')])
            mixed_vat_rate_installed = module_obj.search_count(
                [('name', '=', 'mixed_vat_rate'), ('state', '=', 'installed')])
            amazon_installed = module_obj.search_count(
                [('name', '=', 'amazon_integration'), ('state', '=', 'installed')])

            # Install / Uninstall robo_stock
            if self.module_robo_stock and not robo_stock_installed:
                self = self.change_module_state(module_name='robo_stock', action='install')
            elif not self.module_robo_stock and robo_stock_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='robo_stock', action='uninstall')

            # Install / Uninstall robo_analytic
            if self.module_robo_analytic and not robo_analytic_installed:
                self = self.change_module_state(module_name='robo_analytic', action='install')
            elif not self.module_robo_analytic and robo_analytic_installed:
                if is_system:
                    # TODO Issues occur when module gets uninstalled (most likely due to robo_analytic code moved to robo)
                    raise exceptions.UserError(_('Dėl kodo perkėlimo negalima išjungti analitikos šiuo būdu.'))
                    # if self.module_work_schedule_analytics:
                    # raise exceptions.UserError(_('Norėdami išjungti analitiką - pirmiausia išjungite grafikų analitiką.'))
                    # self = self.change_module_state(module_name='robo_analytic', action='uninstall')
                else:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))

            # Install work_schedule
            if self.module_work_schedule and not work_schedule_installed:
                self = self.change_module_state(module_name='work_schedule', action='install')

            # Check constraints / Install / Uninstall work_schedule_analytics
            if self.module_work_schedule_analytics and not self.module_work_schedule:
                raise exceptions.UserError(
                    _('Norint aktyvuoti analtiką pagal darbo grafikus - aktyvuokite darbo grafikus'))

            if self.module_work_schedule_analytics and not work_schedule_analytics_installed:
                if not self.module_robo_analytic:
                    raise exceptions.UserError(
                        _('Norint aktyvuoti analtiką pagal darbo grafikus - aktyvuokite analitiką'))
                self = self.change_module_state(module_name='work_schedule_analytics', action='install')
            elif not self.module_work_schedule_analytics and work_schedule_analytics_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='work_schedule_analytics', action='uninstall')

            # Install / Uninstall robo_mrp
            if self.module_robo_mrp and not robo_mrp_installed:
                self = self.change_module_state(module_name='robo_mrp', action='install')
            elif not self.module_robo_mrp and robo_mrp_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='robo_mrp', action='uninstall')

            # Install / Uninstall robo_api
            if self.module_robo_api and not robo_api_installed:
                self = self.change_module_state(module_name='robo_api', action='install')
            elif not self.module_robo_api and robo_api_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='robo_api', action='uninstall')

            # Install / Uninstall mixed_vat_rate
            if self.module_mixed_vat_rate and not mixed_vat_rate_installed:
                self = self.change_module_state(module_name='mixed_vat_rate', action='install')
            elif not self.module_mixed_vat_rate and mixed_vat_rate_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='mixed_vat_rate', action='uninstall')

            # Install / Uninstall amazon_integration
            if self.module_amazon_integration and not amazon_installed:
                self = self.change_module_state(module_name='amazon_integration', action='install')
            elif not self.module_amazon_integration and amazon_installed:
                if not is_system:
                    raise exceptions.UserError(_('Ištrinti modulius gali tik sistemos administratorius.'))
                self = self.change_module_state(module_name='amazon_integration', action='uninstall')

            self.set_default_import()

        # Install / Uninstall neopay
        neopay_installed = module_obj.search_count(
            [('name', '=', 'neopay'), ('state', 'in', ['installed', 'to upgrade'])])

        if self.module_neopay and not neopay_installed:
            self = self.change_module_state(module_name='neopay', action='install')
        elif not self.module_neopay and neopay_installed:
            if not self.env.user.is_premium_manager():
                raise exceptions.UserError(_('Ištrinti modulį gali tik įmonės vadovas.'))
            self = self.change_module_state(module_name='neopay', action='uninstall')

        # Return the action to refresh the view
        return self.env.ref('robo.open_robo_vadovas_client_action').read()[0]

    @api.multi
    def cancel(self):
        self.ensure_one()
        return self.env.ref('robo.open_robo_vadovas_client_action').read()[0]

    @api.multi
    def name_get(self):
        return [(rec.id, _('Kompanijos nustatymai')) for rec in self]

    @api.multi
    def subscribe_manager_to_mail_channels(self):
        channels = self.env['res.company'].get_manager_mail_channels()
        manager = self.company_ceo.user_id.partner_id
        if manager and channels:
            channels.write({'channel_partner_ids': [(4, manager.id)]})
