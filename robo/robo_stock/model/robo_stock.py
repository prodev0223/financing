# -*- coding: utf-8 -*-
from numpy import column_stack
from odoo import models, fields, api, _, tools, exceptions
from odoo.addons.queue_job.job import job, identity_exact
from datetime import datetime
import openpyxl as px
import openpyxl.utils.cell
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill
import io
import base64
from odoo.addons.robo.wizard.robo_company_settings import get_all_values, get_mapped, ImportRecord, RoboImportError
from ..wizard.stock_picking_import import import_pickings
from odoo.addons.robo.models.robo_tools import sanitize_account_number
import odoo.addons.decimal_precision as dp
from odoo.addons.robo_basic.models.utils import humanize_number
from dateutil.relativedelta import relativedelta
from odoo.tools.misc import xlwt
import StringIO
import logging

_logger = logging.getLogger(__name__)
XLS_EXT = 'xlsx'
class SaleOrder(models.Model):
    _inherit = 'sale.order'

    name = fields.Char(string='Užsakymo numeris', required=True, copy=False, readonly=True,
                       states={'draft': [('readonly', False)]}, index=True, default=lambda self: _('Naujas užsakymas'))
    state = fields.Selection([
        ('draft', 'Juodraštis'),
        ('sent', 'Išsiųsta'),
        ('sale', 'Vykdomas'),
        ('done', 'Užrakintas'),
        ('cancel', 'Atšauktas'),
    ], string='Būsena', readonly=True, copy=False, index=True, track_visibility='onchange', default='draft')
    invoice_status = fields.Selection([
        ('upselling', 'Papildomo pardavimo galimybė'),
        ('invoiced', 'Užpajamuota'),
        ('to invoice', 'Reikia pajamuoti'),
        ('no', 'Nėra ką pajamuoti')
    ], string='Pajamavimo statusas', compute='_get_invoiced', store=True, readonly=True, lt_string='Pajamavimo statusas')
    products_inside = fields.Html(compute='_compute_order_line_products', store=True, sequence=100,
                                  )

    partner_tag_ids = fields.Many2many(related='partner_id.category_id', string='Partnerio žymos', readonly=True,
                                       groups='robo_basic.group_robo_invoice_see_partner_tags',
                                       )
    is_invoiced = fields.Boolean(compute='_compute_is_invoiced')

    team_id = fields.Many2one(sequence=100)
    tag_ids = fields.Many2many(sequence=100)
    source_id = fields.Many2one(sequence=100)
    campaign_id = fields.Many2one(sequence=100)
    medium_id = fields.Many2one(sequence=100)
    related_project_id = fields.Many2one(sequence=100)
    project_id = fields.Many2one(sequence=100)
    procurement_group_id = fields.Many2one(sequence=100)
    pricelist_id = fields.Many2one(sequence=100)
    origin = fields.Char(sequence=100)
    order_line = fields.One2many(sequence=100)
    opportunity_id = fields.Many2one(sequence=100)
    note = fields.Text(sequence=100)
    message_is_follower = fields.Boolean(sequence=100)
    message_last_post = fields.Datetime(sequence=100)
    message_needaction = fields.Boolean(sequence=100)
    message_needaction_counter = fields.Integer(sequence=100)
    message_follower_ids = fields.One2many(sequence=100)
    message_partner_ids = fields.Many2many(sequence=100)
    message_channel_ids = fields.Many2many(sequence=100)
    invoice_shipping_on_delivery = fields.Boolean(sequence=100)
    client_order_ref = fields.Char(sequence=100)
    carrier_id = fields.Many2one(sequence=100)
    amount_untaxed = fields.Monetary(sequence=100)

    @api.depends('invoice_ids', 'invoice_status')
    def _compute_is_invoiced(self):
        for rec in self:
            rec.is_invoiced = any(x.state in ['open', 'paid'] for x in rec.sudo().mapped('invoice_ids'))

    @api.model
    def create_action_multi_lock(self):
        action = self.env.ref('robo_stock.action_multi_lock')
        if action:
            action.create_action()

    @api.multi
    def mass_lock(self):
        failed_to_lock = str()
        for rec in self.filtered(lambda x: x.state == 'sale'):
            try:
                rec.action_done()
                self.env.cr.commit()
            except Exception as exc:
                self.env.cr.rollback()
                failed_to_lock += '{}. Klaida - {}\n'.format(rec.name, exc.args[0])
                continue
        if failed_to_lock:
            failed_to_lock = 'Nepavyko užrakinti šių pardavimų:\n\n' + failed_to_lock + '\n'
            raise exceptions.UserError(failed_to_lock)

    @api.multi
    @job
    def job_action_done(self):
        """ Validate Sale Orders """
        self.action_done()

    @api.multi
    def action_done(self):
        for rec in self:
            rec._compute_is_invoiced()
            if not rec.is_invoiced:
                raise exceptions.Warning(_('Negalite atlikti šio veiksmo kol pardavimas neturi'
                                           ' bent vienos patvirtintos PVM sąskaitos faktūros.'))
        res = super(SaleOrder, self).action_done()
        if not self._context.get('skip_amount_constraints', False):
            # Execute constraint checking methods and store their potential warning messages in the list
            constraint_checks = [
                self.check_related_invoice_amount_diff(self.is_multi_set()),
                self.check_invoice_picking_prime_cost_diff(self.is_multi_set())
            ]
            warnings = str()
            for check in constraint_checks:
                warnings += '{}\n'.format(check) if check else str()
            if warnings:
                # If there are warnings call the warning wizard
                self.env.cr.rollback()
                message = 'Pardavimo užsakymo užrakinimo įspėjimas:\n{}\nAr tikrai norite tęsti?'.format(warnings)
                wizard = self.env['order.done.warning.wizard'].create(
                    {'sale_order_id': self.id, 'warning_message': message})
                return {
                    'name': _('Įspėjimas'),
                    'type': 'ir.actions.act_window',
                    'view_type': 'form',
                    'view_mode': 'form',
                    'res_model': 'order.done.warning.wizard',
                    'view_id': self.env.ref('robo_stock.form_order_done_warning_wizard').id,
                    'target': 'new',
                    'res_id': wizard.id,
                    'context': self.env.context,
                }
        return res

    @api.model
    def default_get(self, fields):
        res = super(SaleOrder, self).default_get(fields)
        if 'warehouse_id' in fields and not res['warehouse_id']:
            res['warehouse_id'] = self.env['stock.warehouse'].search([], limit=1).id
        return res

    @api.multi
    def sale_order_offer_print(self):
        self.ensure_one()
        return self.env['report'].get_action(self, 'robo_stock.report_sale_offer')

    @api.one
    @api.depends('order_line.product_id', 'order_line.product_uom_qty', 'order_line.product_uom')
    def _compute_order_line_products(self):
        if self.order_line:
            list_of_products = self.order_line.mapped(
                lambda
                    r: '<div class="sale-product-line">' + '<span class="sale-product">' + r.product_id.name + u' ' + (
                    (u'[' + r.product_id.default_code + u'] ') if r.product_id.default_code else u'') + '</span>'
                       + '<span class="sale-qty">' + unicode(humanize_number(r.product_uom_qty)) + '</span>' + '</div>'
            )
            if len(list_of_products) > 3:
                list_of_products = list_of_products[:3]
                list_of_products.append('<div class="sale-product-line"><span>...</span></div>')

            self.products_inside = ''.join(list_of_products)
        else:
            self.products_inside = ' '

    @api.multi
    def check_related_invoice_amount_diff(self, multi_set_operation):
        """
        Assert if summed amount total of related open or paid invoices matches sale order amount total.
        if error is encountered and record set is multi, we raise the error, otherwise we
        return the error message.
        :return: None or error_message (str)
        """
        for rec in self:
            related_invoices = rec.invoice_ids.filtered(lambda x: x.state in ['open', 'paid'])
            non_service_dis_included = all_dis_included = non_service_dis_excluded = all_dis_excluded = 0.0
            allowed_error_rate = 0.1  # Quantities * units can produce minor rounding errors
            for invoice_id in related_invoices:
                # Compute four different prices: All lines ignoring the discount
                # All lines with discount
                # Non service lines ignoring the discount
                # Non service lines with discount. If at least one price matches the SO end sum, we allow the closing
                non_service_dis_included += sum(
                    x.price_unit_tax_included_discount * x.quantity for x in invoice_id.invoice_line_ids.filtered(
                        lambda x: x.product_id.type != 'service'))
                all_dis_included += sum(
                    x.price_unit_tax_included_discount * x.quantity for x in invoice_id.invoice_line_ids)

                non_service_dis_excluded += sum(
                    x.price_unit_tax_included * x.quantity for x in invoice_id.invoice_line_ids.filtered(
                        lambda x: x.product_id.type != 'service'))
                all_dis_excluded += sum(
                    x.price_unit_tax_included * x.quantity for x in invoice_id.invoice_line_ids)

            sale_order_total = rec.amount_total
            if tools.float_compare(abs(non_service_dis_included - sale_order_total),
                                   allowed_error_rate, precision_digits=2) > 0 and \
                    tools.float_compare(abs(all_dis_included - sale_order_total),
                                        allowed_error_rate, precision_digits=2) > 0 and \
                    tools.float_compare(abs(non_service_dis_excluded - sale_order_total),
                                        allowed_error_rate, precision_digits=2) > 0 and \
                    tools.float_compare(abs(all_dis_excluded - sale_order_total),
                                        allowed_error_rate, precision_digits=2) > 0:
                error_message = _('Susijusių sąskaitų bendra suma %s nesutampa su pardavimo bendra suma %s. '
                                  'Parduodant didelį kiekį prekių galima susidurti su apvalinimo paklaida, '
                                  'jei paklaida nėra didelė, siūlome šį pranešimą ignoruoti.') \
                                % (all_dis_included, sale_order_total)
                if multi_set_operation:
                    raise exceptions.UserError(error_message)
                else:
                    return error_message

    @api.multi
    def check_invoice_picking_prime_cost_diff(self, multi_set_operation):
        """
        Collect and sum prime cost from invoice lines and stock quants, compare them
        Raise an error if they exceed allowed prime cost difference.
        if error is encountered and record set is multi, we raise the error, otherwise we
        return the error message.
        :param multi_set_operation: Indicates whether self is multi or single set.
        :return: None or error_message (str)
        """
        for rec in self:
            invoice_ids = rec.invoice_ids.filtered(lambda x: x.state in ['open', 'paid'])
            invoices_total_prime_cost = float(sum(invoice_ids.mapped('invoice_line_ids.cost')))
            quant_ids = rec.picking_ids.filtered(
                lambda x: x.state in ['done']).mapped('move_lines.non_error_quant_ids')
            picking_total_prime_cost = float(sum(q.cost * q.qty for q in quant_ids))
            prime_cost_diff_percentage = abs(((invoices_total_prime_cost / picking_total_prime_cost) - 1) * 100) \
                if picking_total_prime_cost else 100
            allowed_prime_cost_percentage_diff = 10.0  # !Change this part if we want to allow bigger difference
            if tools.float_compare(
                    prime_cost_diff_percentage, allowed_prime_cost_percentage_diff, precision_digits=2) > 0:
                error_message = _('Susijusių važtaraščių ir sąskaitų faktūrų savikaina nesutampa.')
                if multi_set_operation:
                    raise exceptions.UserError(error_message)
                else:
                    return error_message

    @api.multi
    def action_view_invoice_robo(self):
        invoices = self.mapped('invoice_ids')
        action = self.env.ref('robo.open_client_invoice').read()[0]
        action['context'] = {
            'robo_header': {},
            'robo_menu_name': self.env.ref('robo_stock.menu_robo_stock').id,
        }
        if len(invoices) > 1:
            action['domain'] = [('id', 'in', invoices.ids)]
        elif len(invoices) == 1:
            action['views'] = [(self.env.ref('robo.pajamos_form').id, 'form')]
            action['res_id'] = invoices.ids[0]
        return action

    @api.multi
    def action_view_delivery_robo(self):
        '''
        This function returns an action that display existing delivery orders
        of given sales order ids. It can either be a in a list or in a form
        view, if there is only one delivery order to show.
        '''
        action = self.env.ref('robo_stock.open_robo_stock_picking').read()[0]
        action['context'] = {
            'robo_header': {},
            'robo_menu_name': self.env.ref('robo_stock.menu_robo_stock_pickings').id,
        }
        pickings = self.mapped('picking_ids')
        if len(pickings) > 1:
            action['domain'] = [('id', 'in', pickings.ids)]
        elif pickings:
            action['views'] = [(self.env.ref('robo_stock.robo_stock_picking_form').id, 'form')]
            action['res_id'] = pickings.id
        return action

    @api.multi
    def action_invoice_create(self, grouped=False, final=False):
        invoice_ids = super(SaleOrder, self).action_invoice_create(grouped=grouped, final=final)
        for invoice in self.env['account.invoice'].browse(invoice_ids):
            invoice.partner_data_force()
        return invoice_ids

    @api.multi
    def action_unlock(self):
        self.write({'state': 'sale'})

    @api.multi
    def is_multi_set(self):
        """
        Checks if self is multi record-set on single-set
        :return: True if multi, else False
        """
        return True if len(self) > 1 else False


SaleOrder()


class SaleOrderLine(models.Model):
    _inherit = 'sale.order.line'

    def _warehouse_id(self):
        if self._context.get('warehouse_id', False):
            return self._context.get('warehouse_id', False)
        else:
            return self.env['stock.warehouse'].search([], limit=1)

    warehouse_id = fields.Many2one('stock.warehouse', string='Sandėlis', default=_warehouse_id)
    product_uom_categ_id = fields.Many2one('product.uom.categ', related='product_id.uom_id.category_id', readonly=True)
    price_below_cost = fields.Boolean(compute='_compute_sale_below_cost')

    @api.one
    @api.depends('price_unit', 'product_id.avg_cost')
    def _compute_sale_below_cost(self):
        self.price_below_cost = self.price_unit < self.product_id.avg_cost if self.price_unit and self.product_id else False

    @api.onchange('warehouse_id', 'product_id')
    def warehouse_id_changed(self):
        # Do not force the routes if user has advanced location group
        # The route is decided by the procurement rules
        if self.env.user.has_group('stock.group_adv_location'):
            return

        route = self.env['stock.location.route']
        if self.warehouse_id and self.product_id:
            route_id = self.env['stock.location.route'].search(
                [('categ_ids', 'in', self.product_id.categ_id.id), ('sale_selectable', '=', True),
                 ('warehouse_ids', 'in', self.warehouse_id.id)], limit=1)
            if route_id:
                route = route_id.id
        if self.warehouse_id and self.warehouse_id.default_route_id and not route:
            route = self.warehouse_id.default_route_id.id
        if route:
            self.route_id = route
        if self.product_id.product_tmpl_id.packaging_ids:
            self.product_packaging = self.product_id.packaging_ids[0].id

    @api.model
    def create(self, vals):  # added only for testing to get warehouse
        if 'warehouse_id' not in vals:
            order_id = vals['order_id']
            sale_order = self.env['sale.order'].browse(order_id)
            warehouse_id = sale_order.warehouse_id.id
            if warehouse_id:
                vals['warehouse_id'] = warehouse_id
        if 'tax_id' in vals:
            if len(vals['tax_id']) > 0:
                tax_ids_set = set()
                for line in vals['tax_id']:
                    if line[0] == 6:
                        tax_ids_set = set(self.env['account.tax'].browse(line[2]).mapped('id'))
                    elif line[0] == 4:
                        tax_ids_set.add(self.env['account.tax'].browse(line[1]).id)
                    else:
                        continue
                tax_ids_obj = self.env['account.tax'].browse(list(tax_ids_set))
                child_ids = tax_ids_obj.mapped('child_tax_ids.id')
                tax_ids = tax_ids_obj.mapped('id')
                all_ids = list(set(tax_ids + child_ids))
                new_vals = [(6, 0, all_ids)]
                vals['tax_id'] = new_vals
        return super(SaleOrderLine, self).create(vals)

    @api.multi
    def write(self, vals):
        if 'tax_id' in vals:
            if len(vals['tax_id']) > 0:
                tax_ids_set = set()
                for line in vals['tax_id']:
                    if line[0] == 6:
                        tax_ids_set = set(self.env['account.tax'].browse(line[2]).mapped('id'))
                    elif line[0] == 4:
                        tax_ids_set.add(self.env['account.tax'].browse(line[1]).id)
                    else:
                        continue
                tax_ids_obj = self.env['account.tax'].browse(list(tax_ids_set))
                child_ids = tax_ids_obj.mapped('child_tax_ids.id')
                tax_ids = tax_ids_obj.mapped('id')
                all_ids = list(set(tax_ids + child_ids))
                new_vals = [(6, 0, all_ids)]
                vals['tax_id'] = new_vals
        res = super(SaleOrderLine, self).write(vals)
        return res

    @api.onchange('product_uom_qty', 'product_uom', 'route_id', 'warehouse_id', 'product_id')
    def _onchange_product_id_check_availability(self):
        if not self.product_id or not self.product_uom_qty or not self.product_uom or not self.warehouse_id:
            self.product_packaging = False
            return {}
        if self.product_id.type == 'product':
            precision = self.env['decimal.precision'].precision_get('Product Unit of Measure')
            if self.product_uom != self.product_id.uom_id:
                # Sanitize the value so it does not contain 0.000000001 at the end
                # and it is rounded correctly afterwards
                max_rounding = max(self.env['decimal.precision'].precision_get('Product Price'), 8)
                product_uom_qty = tools.float_round(self.product_uom_qty, precision_rounding=max_rounding)
                product_qty = self.product_uom._compute_quantity(product_uom_qty, self.product_id.uom_id)
            else:
                product_qty = self.product_uom_qty
            qty_wh_virtual = self.product_id.with_context(warehouse=self.warehouse_id.id, show_only_internal=False).virtual_available
            qty_wh_on_hand = self.product_id.with_context(warehouse=self.warehouse_id.id, show_only_internal=False).qty_available
            if tools.float_compare(qty_wh_virtual, product_qty, precision_digits=precision) == -1:
                # forecast_schedule = self.get_quantity_forecast(self.product_id.id, product_qty, self.warehouse_id.id)
                forecast_schedule = self.get_quantity_forecast(self.product_id.id, self.warehouse_id.id)
                warning_mess = {
                    'title': _('Neužtenka atsargų!'),
                    'message': _(
                        'Planuojate parduoti %.2f %s, bet prognozuojamas kiekis yra %.2f %s sandėlyje %s!\nTurimos atsargos šiam momentui %.2f %s.\nPrognozuojami kiekiai %s sandėlyje: \n%s') % \
                               (self.product_uom_qty, self.product_uom.name, qty_wh_virtual,
                                self.product_id.uom_id.name, self.warehouse_id.name or 'Sandėlis',
                                qty_wh_on_hand, self.product_id.uom_id.name,
                                self.warehouse_id.name or 'Sandėlis',
                                forecast_schedule)
                }
                return {'warning': warning_mess}
            elif tools.float_compare(qty_wh_on_hand, product_qty, precision_digits=precision) == -1:
                # forecast_schedule = self.get_quantity_forecast(self.product_id.id, product_qty, self.warehouse_id.id)
                forecast_schedule = self.get_quantity_forecast(self.product_id.id, self.warehouse_id.id)

                warning_mess = {
                    'title': _('Neužtenka atsargų!'),
                    'message': _(
                        'Planuojate parduoti %.2f %s, bet turimos atsargos yra tik %.2f %s sandėlyje %s warehouse!\nPrognozuojamas kiekis %.2f %s.\nPrognozuojami kiekiai %s sandėlyje: \n%s') % \
                               (self.product_uom_qty, self.product_uom.name, qty_wh_on_hand,
                                self.product_id.uom_id.name, self.warehouse_id.name or 'Sandėlis',
                                qty_wh_virtual, self.product_id.uom_id.name,
                                self.warehouse_id.name or 'Sandėlis',
                                forecast_schedule)
                }
                return {'warning': warning_mess}
        return {}

    def get_quantity_forecast(self, product_id, warehouse_id):

        query = '''select date, SUM(quantity) as quantity, SUM(cumulative_quantity) as cumulative_quantity
                            from report_stock_forecast
                            WHERE product_id = %s ''' % product_id
        location_ids = self.sudo().env['stock.location'].search([('warehouse_id', '=', warehouse_id)]).mapped('id')
        if location_ids:
            query += 'and location_id IN (%s)' % (','.join(map(str, location_ids)))
        query += 'GROUP BY date ORDER BY date'

        self._cr.execute(query)
        result = self._cr.dictfetchall()
        forecast_lines = ''
        for i, row in enumerate(result):
            if i > 5:
                break
            if i > 0 and tools.float_compare(row['quantity'], 0, precision_digits=2) == 0:
                continue
            forecast_lines += row['date'] + ': ' + str(row['quantity']) + ' (' + str(row['cumulative_quantity']) + ')\n'
        if not forecast_lines:
            forecast_lines += datetime.now().strftime(tools.DEFAULT_SERVER_DATE_FORMAT) + ': 0.0'

        return forecast_lines

    @api.multi
    def _prepare_order_line_procurement(self, group_id=False):
        vals = super(SaleOrderLine, self)._prepare_order_line_procurement(group_id=group_id)
        vals.update({
            'warehouse_id': self.warehouse_id.id,
        })
        return vals


SaleOrderLine()


class PurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    def default_location_dest_id(self):
        return self.env['stock.location'].search([('usage', '=', 'internal')], order='id desc', limit=1).id

    name = fields.Char(string='Užsakymo numeris', default=lambda self: _('Naujas užsakymas'))
    state = fields.Selection([
        ('draft', 'Juodraštis'),
        ('sent', 'Išsiųsta'),
        ('to approve', 'Laukia patvirtinimo'),
        ('purchase', 'Vykdomas'),
        ('done', 'Užrakintas'),
        ('cancel', 'Atšauktas')
    ], string='Būsena')
    invoice_status = fields.Selection([
        ('no', 'Nėra ką pajamuoti'),
        ('to invoice', 'Reikia pajamuoti'),
        ('invoiced', 'Užpajamuota')
    ], string='Pajamavimo statusas', compute='_get_invoiced', lt_string='Pajamavimo statusas')
    products_inside = fields.Html(compute='_compute_purchase_line_products', store=True, sequence=100)
    location_dest_id = fields.Many2one('stock.location', default=default_location_dest_id)
    partner_ref = fields.Char(help='', sequence=100)

    picking_type_id = fields.Many2one(sequence=100)
    order_line = fields.One2many(sequence=100)
    notes = fields.Text(sequence=100)
    message_is_follower = fields.Boolean(sequence=100)
    message_last_post = fields.Datetime(sequence=100)
    message_needaction = fields.Boolean(sequence=100)
    message_needaction_counter = fields.Integer(sequence=100)
    message_follower_ids = fields.One2many(sequence=100)
    message_partner_ids = fields.Many2many(sequence=100)
    message_channel_ids = fields.Many2many(sequence=100)
    default_location_dest_id_usage = fields.Selection(sequence=100)

    @api.multi
    def button_done(self):
        if self.company_id.politika_sandelio_apskaita == 'extended' and not self._context.get(
                'skip_amount_constraints', False):
            # Execute constraint checking methods and store their potential warning messages in the list
            constraint_checks = [
                self.check_invoice_picking_prime_cost_diff(self.is_multi_set())
            ]
            warnings = str()
            for check in constraint_checks:
                warnings += '{}\n'.format(check) if check else str()
            if warnings:
                # If there are warnings call the warning wizard
                self.env.cr.rollback()
                message = 'Pirkimo užsakymo užrakinimo įspėjimas:\n{}\nAr tikrai norite tęsti?'.format(warnings)
                wizard = self.env['order.done.warning.wizard'].create(
                    {'purchase_order_id': self.id, 'warning_message': message})
                return {
                    'name': _('Įspėjimas'),
                    'type': 'ir.actions.act_window',
                    'view_type': 'form',
                    'view_mode': 'form',
                    'res_model': 'order.done.warning.wizard',
                    'view_id': self.env.ref('robo_stock.form_order_done_warning_wizard').id,
                    'target': 'new',
                    'res_id': wizard.id,
                    'context': self.env.context,
                }
        return super(PurchaseOrder, self).button_done()

    @api.one
    @api.depends('order_line.product_id', 'order_line.product_qty', 'order_line.product_uom')
    def _compute_purchase_line_products(self):
        if self.order_line:
            list_of_products = self.order_line.mapped(
                lambda
                    r: '<div class="purchase-product-line">' + '<span class="purchase-product">' + r.product_id.name + u' ' + (
                    (u'[' + r.product_id.default_code + u'] ') if r.product_id.default_code else u'') + '</span>'
                       + '<span class="purchase-qty">' + unicode(humanize_number(r.product_qty)) + '</span>' + '</div>'
            )
            if len(list_of_products) > 3:
                list_of_products = list_of_products[:3]
                list_of_products.append('<span>...</span>')

            self.products_inside = ''.join(list_of_products)
        else:
            self.products_inside = ''

    @api.depends('state', 'order_line.qty_invoiced', 'order_line.product_qty', 'order_line.qty_received')
    def _get_invoiced(self):
        precision = self.env['decimal.precision'].precision_get('Product Unit of Measure')
        for order in self:
            if order.state not in ('purchase', 'done'):
                order.invoice_status = 'no'
                continue

            purchase_lines = order.order_line.filtered(
                lambda r: r.product_id.purchase_method in ['purchase', 'received_with_adjustment'])
            receive_lines = order.order_line - purchase_lines

            if purchase_lines and any(
                    tools.float_compare(line.qty_invoiced, line.product_qty, precision_digits=precision) == -1
                    for line in
                    purchase_lines):
                order.invoice_status = 'to invoice'
            if receive_lines and any(
                    tools.float_compare(line.qty_invoiced, line.qty_received, precision_digits=precision) == -1
                    for line in
                    receive_lines):
                order.invoice_status = 'to invoice'
            elif purchase_lines and all(
                    line.qty_invoiced > 0 and tools.float_compare(line.qty_invoiced, line.product_qty,
                                                                  precision_digits=precision) >= 0 for
                    line in
                    purchase_lines):
                order.invoice_status = 'invoiced'
            elif receive_lines and all(
                    line.qty_invoiced > 0 and tools.float_compare(line.qty_invoiced, line.qty_received,
                                                                  precision_digits=precision) >= 0 for
                    line in
                    receive_lines):
                order.invoice_status = 'invoiced'
            else:
                order.invoice_status = 'no'

    @api.multi
    def action_view_invoice_robo(self):
        '''
        This function returns an action that display existing vendor bills of given purchase order ids.
        When only one found, show the vendor bill immediately.
        '''
        self.ensure_one()
        action = self.env.ref('robo.robo_expenses_action')
        result = action.read()[0]

        # override the context to get rid of the default filtering
        result['context'] = {
            'default_purchase_id': self.id,
            'robo_menu_name': self.env.ref('robo.menu_islaidos').id,
            'default_type': 'in_invoice',
            'type': 'in_invoice',
            'journal_type': 'purchase',
            'robo_template': 'RecentInvoices',
            'activeBoxDomain': [('state', '!=', 'cancel')],
            'search_add_custom': False,
            'robo_header': {},
            'default_currency_id': self.currency_id.id,
            'default_reference': self.partner_ref,
            'creation_from_purchase_order': True
        }

        if not self.invoice_ids:
            # Choose a default account journal in the same currency in case a new invoice is created
            journal_domain = [
                ('type', '=', 'purchase'),
                ('company_id', '=', self.company_id.id),
                ('currency_id', '=', self.currency_id.id),
            ]
            default_journal_id = self.env['account.journal'].search(journal_domain, limit=1)
            if default_journal_id:
                result['context']['default_journal_id'] = default_journal_id.id
        else:
            # Use the same account journal than a previous invoice
            result['context']['default_journal_id'] = self.invoice_ids[0].journal_id.id

        # choose the view_mode accordingly
        if len(self.invoice_ids) != 1:
            result['domain'] = "[('id', 'in', " + str(self.invoice_ids.ids) + ")]"
        elif len(self.invoice_ids) == 1:
            res = self.env.ref('robo.robo_expenses_form', False)
            result['views'] = [(res and res.id or False, 'form')]
            result['view_id'] = res.id if res else False
            result['res_id'] = self.invoice_ids.id
        result['context'] = unicode(result['context'])
        return result

    @api.multi
    def action_view_picking_robo(self):
        '''
        This function returns an action that display existing picking orders of given purchase order ids.
        When only one found, show the picking immediately.
        '''
        action = self.env.ref('robo_stock.open_robo_stock_picking')
        result = action.read()[0]

        # override the context to get rid of the default filtering on picking type
        result.pop('id', None)
        result['context'] = {'robo_header': {},
                             'robo_menu_name': self.env.ref('robo_stock.menu_robo_stock_pickings').id,
                             }
        pick_ids = sum([order.picking_ids.ids for order in self], [])
        # choose the view_mode accordingly
        if len(pick_ids) > 1:
            result['domain'] = "[('id','in',[" + ','.join(map(str, pick_ids)) + "])]"
        elif len(pick_ids) == 1:
            res = self.env.ref('robo_stock.robo_stock_picking_form', False)
            result['views'] = [(res and res.id or False, 'form')]
            result['res_id'] = pick_ids and pick_ids[0] or False
        return result

    @api.one
    def _check_for_zero_qty_line(self):
        for line in self.order_line:
            if tools.float_is_zero(line.product_qty, precision_rounding=line.product_id.uom_id.rounding or 0.001):
                raise exceptions.UserError(_('%s produktas neturi nustatyto kiekio.') % line.product_id.name)

    @api.multi
    def button_confirm(self):
        if self.env.user.has_group('robo_stock.group_purchase_user_all'):
            self = self.sudo()
        for order in self:
            if order.partner_ref and self.search([('partner_ref', '=', order.partner_ref),
                                                  ('partner_id', '=', order.partner_id.id)], count=True) > 1 \
                    and not self._context.get('skip', False):
                raise exceptions.Warning(_('Negali kartotis tiekėjo sąskaitos numeris.'))
            order._check_for_zero_qty_line()
        return super(PurchaseOrder, self).button_confirm()

    @api.multi
    def check_invoice_picking_prime_cost_diff(self, multi_set_operation):
        """
        Collect and sum prime cost from invoice lines and stock quants, compare them
        Raise an error if they exceed allowed prime cost difference.
        if error is encountered and record set is multi, we raise the error, otherwise we
        return the error message.
        :param multi_set_operation: Indicates whether self is multi or single set.
        :return: None or error_message (str)
        """
        for rec in self:
            quants_total = 0.0
            invoices_total = sum(
                line.price_subtotal_signed for line in
                rec.mapped('invoice_ids.invoice_line_ids').filtered(lambda r: r.product_id.type == 'product')
            )
            for picking_id in rec.picking_ids:
                if picking_id.state not in ['cancel', 'done']:
                    raise exceptions.Warning(_('Negalite užrakinti pirkimo užsakymo, turite vykdomų važtaraščių.'))
                quant_ids = picking_id.sudo().mapped('move_lines.non_error_quant_ids')
                landed_costs = sum(quant_ids.mapped('valuation_adjustment_ids.additional_landed_cost'))
                quants_total += sum(quant_ids.mapped('inventory_value')) - landed_costs
            diff = invoices_total - quants_total
            if tools.float_is_zero(diff, precision_digits=2):
                continue
            if tools.float_is_zero(quants_total, precision_digits=2) or \
                    tools.float_compare(abs(diff / quants_total), 0.01, precision_digits=2) > 0:
                error_message = _('Pirktų produktų savikaina nesutampa su sąskaitos faktūros suma.')
                if multi_set_operation:
                    raise exceptions.Warning(error_message)
                else:
                    return error_message

    @api.multi
    def is_multi_set(self):
        """
        Checks if self is multi record-set on single-set
        :return: True if multi, else False
        """
        return True if len(self) > 1 else False


PurchaseOrder()


class StockMove(models.Model):
    _inherit = 'stock.move'

    def get_date(self):
        return self._context.get('date', False) or datetime.utcnow()

    state = fields.Selection([
        ('draft', 'Nepatvirtinta'), ('cancel', 'Atšaukta'),
        ('waiting', 'Laukia susijusių operacijų'), ('confirmed', 'Trūksta atsargų'),
        ('assigned', 'Paruošta išsiuntimui'), ('done', 'Išsiųsta')],
        help="")

    date_expected = fields.Datetime(default=get_date)

    # Add sequence 1..2..3 to important fields in pivot
    location_id = fields.Many2one(sequence=1)
    location_dest_id = fields.Many2one(sequence=2)
    picking_id = fields.Many2one(sequence=3)

    # Add sequence 100 to some semi-systemic fields
    # so that they are excluded from robo front pivot view
    bom_line_id = fields.Many2one(sequence=100)
    consume_unbuild_id = fields.Many2one(sequence=100)
    split_from = fields.Many2one(sequence=100)
    operation_id = fields.Many2one(sequence=100)
    origin_returned_move_id = fields.Many2one(sequence=100)
    partially_available = fields.Boolean(sequence=100)
    product_packaging = fields.Many2one(sequence=100)
    rule_id = fields.Many2one(sequence=100)
    push_rule_id = fields.Many2one(sequence=100)
    to_refund_so = fields.Boolean(sequence=100)
    weight_uom_id = fields.Many2one(sequence=100)
    workorder_id = fields.Many2one(sequence=100)
    move_dest_id = fields.Many2one(sequence=100)
    non_error_qty = fields.Float(sequence=100)
    inventory_line_id = fields.Many2one(sequence=100)

    original_product_id = fields.Many2one('product.product', string='Original product before mapping', readonly=True)

    def get_price_unit(self):
        move = self
        picking = move.picking_id
        if move.product_id.purchase_method != 'received_with_adjustment':
            return super(StockMove, self).get_price_unit()

        if move.purchase_line_id:
            qty = 0.0
            for line in picking.move_lines.filtered(lambda r: r.product_id.id == move.product_id.id):
                if line.state in ['confirmed'] and line.split_from:
                    continue
                qty += line.product_uom_qty
            order = move.purchase_line_id.order_id
            if order.currency_id != move.company_id.currency_id:
                price_unit = move.purchase_line_id._get_stock_move_price_unit()
                purchase_qty = move.purchase_line_id.product_qty
                price_total = price_unit * purchase_qty
            else:
                price_total = 0.0
                for order_line in order.order_line.filtered(lambda r: r.product_id.id == move.product_id.id):
                    price_total += order_line.price_subtotal
                    price_total -= move.company_id.currency_id.round(order_line.price_unit * order_line.qty_received)
                    # line_subtotal = move.purchase_line_id.price_subtotal
            unit_cost = price_total / qty

            move.price_unit = unit_cost
            return unit_cost
        return super(StockMove, self).get_price_unit()

    @api.model
    def create_stock_move_unreserve_action(self):
        action = self.env.ref('robo_stock.stock_move_unreserve_action')
        if action:
            action.create_action()

    @api.onchange('product_id')
    def onchange_product_id(self):
        product = self.product_id.with_context(lang=self.partner_id.lang or self.env.user.lang)
        self.name = product.partner_ref
        self.product_uom = product.uom_id.id
        return {'domain': {'product_uom': [('category_id', '=', product.uom_id.category_id.id)]}}


StockMove()


class ProductProduct(models.Model):
    _inherit = 'product.product'

    @api.multi
    def action_open_quants_robo(self):
        self.ensure_one()
        action = self.env.ref('robo_stock.robo_open_current_inventory').read()[0]
        action['domain'] = [('product_id', '=', self.id)]
        action['context'] = {
            'search_default_locationgroup': 1,
            'search_default_internal_loc': 1,
            'robo_header': {'fit': True},
        }
        return action

    def action_open_forecast_robo(self):
        self.ensure_one()
        action = self.env.ref('stock.action_stock_level_forecast_report_template').read()[0]
        action['domain'] = [('product_id', '=', self.id)]
        action['context'] = {
            'search_default_pivot_by': 1,
            'search_default_graph_by': 1
        }
        return action

    @api.multi
    def open_done_moves(self):
        action = self.env.ref('robo_stock.action_robo_done_stock_move').read()[0]
        domain = [('product_id', 'in', self.ids)]
        action['domain'] = domain
        return action

    @api.multi
    def open_reserved_moves(self):
        action = self.env.ref('robo_stock.action_robo_reserved_stock_move').read()[0]
        domain = [
            ('state', 'in', ['confirmed', 'assigned']),
            ('product_id', 'in', self.ids),
            ('reserved_quant_ids', '!=', False),
        ]
        action['domain'] = domain
        return action

    @api.multi
    def get_mapped_moves(self, moves, partner):
        """
        Filters out and returns stock moves of current product mapping
        if it exists for current passed partner, or if there's
        a mapping with no associated partner.
        :param moves: stock.move recordset
        :param partner: res.partner record
        :return: stock.move recordset
        """
        self.ensure_zero_or_one()

        product_moves = self.env['stock.move']
        # Try to search for mapping by partner
        mapping = self.mapping_ids.filtered(
            lambda r: r.partner_id == partner
        )
        # If there's no mapping, try to search for mapping
        # without any partner (constraint - one per product)
        if not mapping:
            mapping = self.mapping_ids.filtered(
                lambda r: not r.partner_id
            )
        # If mapping was found, try to search for moves using the mapping
        if mapping:
            product_moves = moves.filtered(
                lambda r: r.product_id == mapping.mapped_product_id
            )
        return product_moves


ProductProduct()


class StockQuant(models.Model):
    _inherit = 'stock.quant'

    name = fields.Char(string='ID')
    product_id = fields.Many2one(string='Produktas')
    location_id = fields.Many2one(string='Vieta')
    qty = fields.Float(string='Kiekis', sequence=1,
                       help="Kiekis rodomas pagrindiniais matavimo vienetais")
    product_uom_id = fields.Many2one(string='Matavimo vienetai')
    package_id = fields.Many2one(string='Pakuotė',
                                 help="Pakuotė su šiuo produktu")
    packaging_type_id = fields.Many2one(string='Pakuotės tipas')
    reservation_id = fields.Many2one(string='Rezervuota perkėlimui', help="", lt_string='Rezervuota perkėlimui')
    lot_id = fields.Many2one(string='SN')
    cost = fields.Float(string='Vieneto kaina', group_operator='qty', sequence=2)
    owner_id = fields.Many2one(string='Savininkas',
                               help="Produkto savininkas")
    create_date = fields.Datetime(string='Sukūrimo data')
    in_date = fields.Datetime(string='Gavimo data', lt_string='Gavimo data')
    history_ids = fields.Many2many(string='Perkėlimai', help="")
    company_id = fields.Many2one(string='Kompanija', help="", sequence=100)
    inventory_value = fields.Float(string='Atsargų vertė')
    propagated_from_id = fields.Many2one(string='Susijęs produkto gavimas',
                                         help='Atsiranda dėl pristatymo neturint atsargų')
    negative_move_id = fields.Many2one(string='Neigiamo kiekio perkėlimas', help="", sequence=100)
    negative_dest_location_id = fields.Many2one(string="Neigiamo kiekio lokacija", help="", sequence=100)

    category_id = fields.Many2one('product.category', string='Kategorija', compute='_category_id', store=True)

    consumed_quant_ids = fields.Many2many(sequence=100)
    produced_quant_ids = fields.Many2many(sequence=100)
    valuation_adjustment_ids = fields.Many2many(sequence=100)

    @api.one
    @api.depends('product_id.product_tmpl_id.categ_id')
    def _category_id(self):
        self.category_id = self.product_id.categ_id.id


StockQuant()


class StockWarehouse(models.Model):
    _inherit = 'stock.warehouse'

    default_route_id = fields.Many2one('stock.location.route', string='Sandėlio taisyklė')
    is_consignative = fields.Boolean(string='Konsignacinis sandėlis')

    @api.model
    def create(self, vals):
        if not self.env.user.is_premium_manager():
            raise exceptions.AccessError(_('Only the company manager can perform this action'))
        route_obj = self.env['stock.location.route']
        proc_rule_obj = self.env['procurement.rule']
        # vals.update({'default_route_id': route_id})
        ret = super(StockWarehouse, self).create(vals)
        if ret.default_route_id:
            return ret

        cust_location_id = self.env['stock.location'].search([('usage', '=', 'customer'), ('active', '=', True)],
                                                             limit=1)
        proc_rule_vals = {'action': 'move',
                          'active': True,
                          'name': 'From ' + vals['name'],
                          'procure_method': 'make_to_stock',
                          'location_id': cust_location_id.id,
                          'warehouse_id': ret.id,
                          'location_src_id': ret.lot_stock_id.id,
                          'picking_type_id': ret.out_type_id.id,
                          'delay': 0
                          }
        proc_rule_id = proc_rule_obj.create(proc_rule_vals)
        name = ('Pardavimai - %s' % vals['name'])
        vals = {'name': name,
                'sale_selectable': True,
                'product_selectable': False,
                'pull_ids': [(6, 0, [proc_rule_id.id])]}
        route_id = route_obj.create(vals)
        ret.default_route_id = route_id.id
        return ret

    @api.model
    def create_routes_for_exist_wh(self):
        route_obj = self.env['stock.location.route']
        proc_rule_obj = self.env['procurement.rule']
        cust_location_id = self.env['stock.location'].search([('usage', '=', 'customer'), ('active', '=', True)],
                                                             limit=1)
        warehouses = self.env['stock.warehouse'].search([])
        for warehouse in warehouses:
            # vals.update({'default_route_id': route_id})
            if warehouse.default_route_id:
                continue

            proc_rule_vals = {'action': 'move',
                              'active': True,
                              'name': 'From ' + warehouse.name,
                              'procure_method': 'make_to_stock',
                              'location_id': cust_location_id.id,
                              'warehouse_id': warehouse.id,
                              'location_src_id': warehouse.lot_stock_id.id,
                              'picking_type_id': warehouse.out_type_id.id,
                              'delay': 0
                              }
            proc_rule_id = proc_rule_obj.create(proc_rule_vals)
            name = ('Pardavimai - %s' % warehouse.name)
            vals = {'name': name,
                    'sale_selectable': True,
                    'product_selectable': False,
                    'pull_ids': [(6, 0, [proc_rule_id.id])]}
            route_id = route_obj.create(vals)
            warehouse.default_route_id = route_id.id

    @api.model
    def activate_internal_transfers(self):
        self.env['stock.picking.type'].search([('code', '=', 'internal'), ('active', '=', False)]).write({
            'active': True
        })

    @api.multi
    def write(self, vals):
        if self.env.user.is_manager():
            return super(StockWarehouse, self.sudo()).write(vals=vals)
        else:
            return super(StockWarehouse, self).write(vals=vals)


StockWarehouse()


class InvoiceDeliveryWizard(models.TransientModel):
    _inherit = 'invoice.delivery.wizard'

    def _default_location_id(self):
        return self.env.user.employee_ids[0].mapped('department_id.default_stock_location_id') \
            if self.env.user.employee_ids else False

    location_id = fields.Many2one(default=_default_location_id)

    @api.multi
    def prepare_data_by_location(self):
        """
        Overridden method //
        Checks whether multi location mode is activated,
        if it is, groups invoice line IDs by location
        and returns the data
        :return: grouped data (dict)
        """
        if self.sudo().env.user.company_id.simplified_stock_multi_locations:
            lines_by_location = {}
            for line in self.invoice_id.invoice_line_ids:
                # Loop through lines and build dict of locations
                # If line does not have it's location set, keep the previous behaviour.
                # self.location_id is either the default value, or, if specified, one
                # from passed account invoice record
                loc = line.location_id or self.location_id
                if not loc.warehouse_id:
                    raise exceptions.UserError(_('Nepavyko nustatyti sandėlio atsargų vietai - %s.') % loc.name)
                lines_by_location.setdefault(loc, self.env['account.invoice.line'])
                lines_by_location[loc] |= line
        else:
            # If multi location mode is not activated, all lines will have the same location
            lines_by_location = {self.location_id: self.invoice_id.invoice_line_ids}
        return lines_by_location

    @api.multi
    def confirm_delivery(self, check_quantities=False):
        """
        Gather related pickings and confirm them
        :return: None
        """
        pickings = self.invoice_id.get_related_pickings()
        pickings.confirm_delivery(check_quantities=check_quantities, invoice=self.invoice_id)

    @api.multi
    def create_delivery(self):
        super(InvoiceDeliveryWizard, self).create_delivery()
        if self.invoice_id.picking_id:
            return self.invoice_id.open_picking()

    @api.multi
    def get_move_vals(self, invoice_line, picking):
        res = super(InvoiceDeliveryWizard, self).get_move_vals(invoice_line, picking)
        mapping = invoice_line.sudo().product_id.mapping_ids.get_mapping(self.invoice_id.partner_id)
        if mapping:
            res.update({
                'product_uom_qty': res.get('product_uom_qty') * mapping.ratio,
                'product_uom': mapping.mapped_product_id.uom_id.id,
                'original_product_id': invoice_line.product_id.id,
                'name': mapping.mapped_product_id.name,
                'product_id': mapping.mapped_product_id.id,
            })
        return res


InvoiceDeliveryWizard()


class ProductCategory(models.Model):
    _inherit = 'product.category'

    stock_surplus_account_categ_id = fields.Many2one('account.account', string='Atsargų pertekliaus sąskaita')


ProductCategory()


class ProductTemplate(models.Model):
    _name = 'product.template'
    _inherit = ['product.template', 'ir.attachment.drop']


    # tara_metalas = fields.Float(string='Tara (metalas)')
    # tara_plastmase = fields.Float(string='Tara (plastmasė)')
    # tara_stiklas = fields.Float(string='Tara (stiklas)')
    # tara_popierius = fields.Float(string='Tara (popierius)')
    # tara_medis = fields.Float(string='Tara (medis)')
    # tara_pet = fields.Float(string='Tara (PET)')
    # tara_kombinuota = fields.Float(string='Tara (kombinuota)')
    # tara_kita = fields.Float(string='Tara (kita)')
    attachment_drop_lock = fields.Boolean(compute='_compute_attachment_drop_lock')
    kilmes_salis = fields.Many2one('res.country', string='Kilmės šalis',
                                   default=lambda self: self.env.ref('base.lt').id)
    purchase_method = fields.Selection([
        ('purchase', 'Užsakytais kiekiais'),
        ('receive', 'Gautais kiekiais'),
        ('received_with_adjustment', 'Gautais kiekiais su kainos koregavimu'),
    ], default="purchase",
    )
    avg_cost = fields.Monetary(string='Vid. savikaina', currency_field='currency_id', compute='_avg_cost')
    avg_cost_float = fields.Float(
        string='Average cost', currency_field='currency_id', compute='_avg_cost',
        digits=dp.get_precision('Product Price')
    )

    latest_price = fields.Monetary(string='Naujausia pardavimo kaina', currency_field='currency_id', sequence=100)
    default_code = fields.Char(track_visibility='onchange')
    intrastat_id = fields.Many2one('report.intrastat.code', track_visibility='onchange')
    type = fields.Selection(track_visibility='onchange')
    list_price = fields.Float(track_visibility='onchange')
    weight = fields.Float(track_visibility='onchange')
    rel_product_variant_count = fields.Integer(compute='_compute_rel_product_variant_count')
    stock_surplus_account_id = fields.Many2one('account.account', string='Atsargų pertekliaus sąskaita', sequence=100)
    mapping_ids = fields.One2many('product.mapping', 'template_id', string='Product mapping for document processing',
                                  groups='base.group_system,robo_stock.group_robo_product_mapping',
                                  sequence=100,
                                  )

    @api.multi
    @api.constrains('mapping_ids')
    def constraint_mapping_ids(self):
        for rec in self:
            if len(rec.mapping_ids) != len(set(rec.mapping_ids.mapped('partner_id.id'))):
                raise exceptions.ValidationError(_('You can add partner only once.'))
            if rec.id in rec.mapping_ids.mapped('mapped_product_id.product_tmpl_id.id'):
                raise exceptions.ValidationError(_('You cannot map to the same product.'))

    @api.multi
    @api.depends('product_variant_ids')
    def _compute_rel_product_variant_count(self):
        """
        Compute //
        Calculate count of related product variants
        :return: None
        """
        for rec in self:
            rec.rel_product_variant_count = len(rec.product_variant_ids)

    @api.multi
    def toggle_active(self):
        for rec in self:
            if rec.active and rec.qty_available:
                raise exceptions.UserError(_('Negalite suarchyvuoti produkto %s, '
                                             'kurio turimas kiekis yra didesnis nei 0!' % rec.name))
        return super(ProductTemplate, self).toggle_active()

    @api.multi
    def _compute_attachment_drop_lock(self):
        user_has_group_robo_stock = self.env.user.has_group('stock_extend.group_robo_stock')
        for rec in self:
            if user_has_group_robo_stock:
                rec.attachment_drop_lock = False
            else:
                rec.attachment_drop_lock = True

    @api.multi
    @api.depends('type', 'standard_price')
    def _avg_cost(self):
        for rec in self:
            if rec.type == 'product':
                all_quants = self.sudo().env['stock.quant'].search(
                    [('product_id.product_tmpl_id', '=', rec.id), ('location_id.usage', '=', 'internal')])
                all_cost = sum(all_quants.mapped(lambda r: r.qty * r.cost))
                all_qty = sum(all_quants.mapped('qty'))
                if all_qty:
                    rec.avg_cost = rec.avg_cost_float = all_cost / all_qty
            else:
                rec.avg_cost = rec.avg_cost_float = rec.standard_price

    @api.multi
    def action_open_quants_robo(self):
        products = self.with_context(active_test=False).mapped('product_variant_ids')
        action = self.env.ref('robo_stock.robo_open_current_inventory').read()[0]
        action['domain'] = [('product_id', 'in', products.ids)]
        action['context'] = {
            'search_default_locationgroup': 1,
            'search_default_internal_loc': 1,
            'robo_header': {'fit': True},
        }
        return action

    def action_open_forecast_robo(self):
        products = self.with_context(active_test=False).mapped('product_variant_ids')
        action = self.env.ref('stock.action_stock_level_forecast_report_template').read()[0]
        action['domain'] = [('product_id', 'in', products.ids)]
        action['context'] = {
            'search_default_pivot_by': 1,
            'search_default_graph_by': 1
        }
        return action

    @api.multi
    def action_open_product_variants(self):
        """
        Reads and returns the action to open product.product tree
        Called from product.template, filters out the product.product
        records that are related to current template
        :return: None
        """
        self.ensure_one()
        products = self.with_context(active_test=False).mapped('product_variant_ids')
        action = self.env.ref('robo_stock.action_open_product_product').read()[0]
        action['domain'] = [('id', 'in', products.ids)]
        return action

    @api.onchange('landed_cost_ok')
    def onchange_landed_cost_ok(self):
        if self.landed_cost_ok:
            self.type = 'service'
        else:
            self.type = 'product'

    @api.multi
    def open_done_moves(self):
        action = self.env.ref('robo_stock.action_robo_done_stock_move').read()[0]
        domain = [('product_id', 'in', self.with_context(active_test=False).product_variant_ids.ids)]
        action['domain'] = domain
        return action

    @api.multi
    def open_reserved_moves(self):
        action = self.env.ref('robo_stock.action_robo_reserved_stock_move').read()[0]
        domain = [
            ('state', 'in', ['confirmed', 'assigned']),
            ('product_id', 'in', self.with_context(active_test=False).product_variant_ids.ids),
            ('reserved_quant_ids', '!=', False),
        ]
        action['domain'] = domain
        return action

    @api.model
    def create_multi_toggle_active_product_action(self):
        action = self.env.ref('robo_stock.multi_toggle_active_product_action')
        if action:
            action.create_action()


ProductTemplate()


class ProductMapping(models.Model):
    _name = 'product.mapping'

    template_id = fields.Many2one('product.template', string='Product Template', required=True, ondelete='cascade')
    partner_id = fields.Many2one('res.partner', string='Partner')
    mapped_product_id = fields.Many2one(
        'product.product', string='Map to',
        domain="[('type', '=', 'product')]",
        required=True, ondelete='cascade'
    )
    ratio = fields.Float(string='Conversion Ratio', default=1, digits=(16, 5))
    uom_from_id = fields.Many2one('product.uom', string='UOM', compute='_compute_uom')
    uom_to_id = fields.Many2one('product.uom', string='Mapped UOM', compute='_compute_uom')

    @api.multi
    @api.depends('mapped_product_id.uom_id', 'template_id.uom_id')
    def _compute_uom(self):
        for rec in self:
            rec.uom_from_id = rec.template_id.uom_id.id
            rec.uom_to_id = rec.mapped_product_id.uom_id.id

    @api.multi
    @api.constrains('ratio')
    def _check_ratio(self):
        for rec in self:
            # Ensure that ratio is higher than zero
            if tools.float_compare(rec.ratio, 0.0, precision_digits=2) <= 0:
                raise exceptions.ValidationError(_('Incorrect ratio'))

    @api.multi
    @api.constrains('partner_id')
    def _check_partner_id(self):
        """Check partner ID field constraints"""
        for rec in self:
            domain = [('template_id', '=', rec.template_id.id)]
            # Only one mapping without any partner is allowed in the system
            if not rec.partner_id and self.search_count(domain + [('partner_id', '=', False)]) > 1:
                raise exceptions.ValidationError(_('There can only be one mapping with no assigned partner'))

            # Only one mapping per partner is allowed in the system
            if rec.partner_id and self.search_count(domain + [('partner_id', '=', rec.partner_id.id)]) > 1:
                raise exceptions.ValidationError(_('There can only be one mapping assigned per partner'))

    @api.multi
    def get_mapping(self, partner):
        """
        Returns the mapping for passed partner,
        if such mapping does not exist,
        return the mapping with no partner
        :param partner: res.partner record
        :return: product.mapping record
        """
        # Constraints should ensure that it's only one record, but
        # limit to one just in case (for current data)
        mapping = self.filtered(lambda x: x.partner_id.id == partner.id)
        if not mapping:
            mapping = self.filtered(lambda x: not x.partner_id)
        return mapping


#
#
# class ReportIntrastat(models.Model):
#     _inherit = "report.intrastat"
#
#     tara_metalas = fields.Float(string='Tara (metalas)')
#     tara_plastmase = fields.Float(string='Tara (plastmasė)')
#     tara_stiklas = fields.Float(string='Tara (stiklas)')
#     tara_popierius = fields.Float(string='Tara (popierius)')
#     tara_medis = fields.Float(string='Tara (medis)')
#     tara_pet = fields.Float(string='Tara (PET)')
#     tara_kombinuota = fields.Float(string='Tara (kombinuota)')
#     tara_kita = fields.Float(string='Tara (kita)')
#     # product_id = fields.Many2one('product.template')
#
#     def init(self):
#         tools.drop_view_if_exists(self.env.cr, self._table)
#         self.env.cr.execute("""
#                create or replace view report_intrastat as (
#                        select name, month, id, intrastat_id, code, value, weight, supply_units, currency_id, ref, type, company_id,
#                         sum(tara_metalas*supply_units) as tara_metalas,
#                         sum(tara_plastmase*supply_units) as tara_plastmase,
#                         sum(tara_stiklas*supply_units) as tara_stiklas,
#                         sum(tara_popierius*supply_units) as tara_popierius,
#                         sum(tara_medis*supply_units) as tara_medis,
#                         sum(tara_pet*supply_units) as tara_pet,
#                         sum(tara_kombinuota*supply_units) as tara_kombinuota,
#                         sum(tara_kita*supply_units) as tara_kita
#
#                     from (
#                        select
#                            to_char(inv.date_invoice, 'YYYY') as name,
#                            to_char(inv.date_invoice, 'MM') as month,
#                            min(inv_line.id) as id,
#                            intrastat.id as intrastat_id,
#                            upper(inv_country.code) as code,
#                            sum(case when inv_line.price_unit is not null
#                                    then inv_line.price_unit * inv_line.quantity
#                                    else 0
#                                end) as value,
#                            sum(
#                                case when uom.category_id != puom.category_id then (pt.weight * inv_line.quantity)
#                                else (pt.weight * inv_line.quantity * uom.factor) end
#                            ) as weight,
#                            sum(
#                                case when uom.category_id != puom.category_id then inv_line.quantity
#                                else (inv_line.quantity * uom.factor) end
#                            ) as supply_units,
#
#                            inv.currency_id as currency_id,
#                            inv.number as ref,
#                            case when inv.type in ('out_invoice','in_refund')
#                                then 'export'
#                                else 'import'
#                                end as type,
#                            inv.company_id as company_id,
#                            pt.tara_metalas as tara_metalas,
#                            pt.tara_plastmase as tara_plastmase,
#                            pt.tara_stiklas as tara_stiklas,
#                            pt.tara_popierius as tara_popierius,
#                            pt.tara_medis as tara_medis,
#                            pt.tara_pet as tara_pet,
#                            pt.tara_kombinuota as tara_kombinuota,
#                            pt.tara_kita as tara_kita
#                        from
#                            account_invoice inv
#                            left join account_invoice_line inv_line on inv_line.invoice_id=inv.id
#                            left join (product_template pt
#                                left join product_product pp on (pp.product_tmpl_id = pt.id))
#                            on (inv_line.product_id = pp.id)
#                            left join product_uom uom on uom.id=inv_line.uom_id
#                            left join product_uom puom on puom.id = pt.uom_id
#                            left join report_intrastat_code intrastat on pt.intrastat_id = intrastat.id
#                            left join (res_partner inv_address
#                                left join res_country inv_country on (inv_country.id = inv_address.country_id))
#                            on (inv_address.id = inv.partner_id)
#                        where
#                            inv.state in ('open','paid')
#                            and inv_line.product_id is not null
#                            and inv_country.intrastat=true
#                        group by to_char(inv.date_invoice, 'YYYY'), to_char(inv.date_invoice, 'MM'),intrastat.id,inv.type,pt.intrastat_id, inv_country.code,inv.number,  inv.currency_id, inv.company_id, pt.id
#                       ) as Foo
#                       group by  name, month, id, intrastat_id, code, value, weight, supply_units, currency_id, ref, type, company_id
#                )""")
#
#
# ReportIntrastat()
#
#
# class ReportTara(models.Model):
#     _name = "report.tara"
#     _inherit = "report.intrastat"
#
#     tara_metalas = fields.Float(string='Tara (metalas)')
#     tara_plastmase = fields.Float(string='Tara (plastmasė)')
#     tara_stiklas = fields.Float(string='Tara (stiklas)')
#     tara_popierius = fields.Float(string='Tara (popierius)')
#     tara_medis = fields.Float(string='Tara (medis)')
#     tara_pet = fields.Float(string='Tara (PET)')
#     tara_kombinuota = fields.Float(string='Tara (kombinuota)')
#     tara_kita = fields.Float(string='Tara (kita)')
#
#     def init(self):
#         tools.drop_view_if_exists(self.env.cr, self._table)
#         self.env.cr.execute("""
#                create or replace view report_tara as (
#                        select name, month, id, intrastat_id, code, value, weight, supply_units, currency_id, ref, type, company_id,
#                         sum(tara_metalas*supply_units) as tara_metalas,
#                         sum(tara_plastmase*supply_units) as tara_plastmase,
#                         sum(tara_stiklas*supply_units) as tara_stiklas,
#                         sum(tara_popierius*supply_units) as tara_popierius,
#                         sum(tara_medis*supply_units) as tara_medis,
#                         sum(tara_pet*supply_units) as tara_pet,
#                         sum(tara_kombinuota*supply_units) as tara_kombinuota,
#                         sum(tara_kita*supply_units) as tara_kita
#
#                     from (
#                        select
#                            to_char(inv.date_invoice, 'YYYY') as name,
#                            to_char(inv.date_invoice, 'MM') as month,
#                            min(inv_line.id) as id,
#                            intrastat.id as intrastat_id,
#                            upper(inv_country.code) as code,
#                            sum(case when inv_line.price_unit is not null
#                                    then inv_line.price_unit * inv_line.quantity
#                                    else 0
#                                end) as value,
#                            sum(
#                                case when uom.category_id != puom.category_id then (pt.weight * inv_line.quantity)
#                                else (pt.weight * inv_line.quantity * uom.factor) end
#                            ) as weight,
#                            sum(
#                                case when uom.category_id != puom.category_id then inv_line.quantity
#                                else (inv_line.quantity * uom.factor) end
#                            ) as supply_units,
#
#                            inv.currency_id as currency_id,
#                            inv.number as ref,
#                            case when inv.type in ('out_invoice','in_refund')
#                                then 'export'
#                                else 'import'
#                                end as type,
#                            inv.company_id as company_id,
#                            pt.tara_metalas as tara_metalas,
#                            pt.tara_plastmase as tara_plastmase,
#                            pt.tara_stiklas as tara_stiklas,
#                            pt.tara_popierius as tara_popierius,
#                            pt.tara_medis as tara_medis,
#                            pt.tara_pet as tara_pet,
#                            pt.tara_kombinuota as tara_kombinuota,
#                            pt.tara_kita as tara_kita
#                        from
#                            account_invoice inv
#                            left join account_invoice_line inv_line on inv_line.invoice_id=inv.id
#                            left join (product_template pt
#                                left join product_product pp on (pp.product_tmpl_id = pt.id))
#                            on (inv_line.product_id = pp.id)
#                            left join product_uom uom on uom.id=inv_line.uom_id
#                            left join product_uom puom on puom.id = pt.uom_id
#                            left join report_intrastat_code intrastat on pt.intrastat_id = intrastat.id
#                            left join (res_partner inv_address
#                                left join res_country inv_country on (inv_country.id = inv_address.country_id))
#                            on (inv_address.id = inv.partner_id)
#                        where
#                            inv.state in ('open','paid')
#                            and inv_line.product_id is not null
#                        group by to_char(inv.date_invoice, 'YYYY'), to_char(inv.date_invoice, 'MM'),intrastat.id,inv.type,pt.intrastat_id, inv_country.code,inv.number,  inv.currency_id, inv.company_id, pt.id
#                       ) as Foo
#                       group by  name, month, id, intrastat_id, code, value, weight, supply_units, currency_id, ref, type, company_id
#                )""")
#
#
# ReportTara()

class InventorySummaryWizard(models.TransientModel):
    _inherit = 'inventory.summary.wizard'

    threaded_reports_are_enabled = fields.Boolean(compute='_compute_threaded_reports_are_enabled')

    @api.multi
    @api.depends('date_from')
    def _compute_threaded_reports_are_enabled(self):
        is_enabled = self.env.user.sudo().company_id.activate_threaded_front_reports
        for rec in self:
            rec.threaded_reports_are_enabled = is_enabled

    @api.multi
    def show_report(self):
        result = super(InventorySummaryWizard, self).show_report()
        if 'name' not in result:
            result['name'] = self._context.get('name', '')
        return result


InventorySummaryWizard()

class InventoryForecastWizard(models.TransientModel):
    _name = 'inventory.forecast.wizard'

    @api.multi
    def name_get(self):
        return [(rec.id, _('Atsargų prognozė')) for rec in self]

    @api.multi
    def show_forecast(self):
        return {
            'context': self._context,
            'view_type': 'form',
            'view_mode': 'pivot',
            'res_model': 'report.stock.forecast',
            'view_id': False,
            'type': 'ir.actions.act_window',
            'id': self.env.ref('robo_stock.robo_view_inventory_forecast').id,
        }

    def generate_excel(self, domain=None):
        if not domain:
            domain = list()
        workbook = px.Workbook()
        worksheet = workbook.active
        worksheet.title = _('Atsargų prognozė')
        headers = [_('Stock Location'), _('Product Name'), _('Quantity')]

        robo_background = ('{:X}{:X}{:X}').format(197, 217, 241)
        fill = PatternFill(patternType='solid', fill_type="solid", fgColor=robo_background)
        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_bold = NamedStyle(name="header bold", font=bold, border=border)

        # variables to adjust the width automatically
        column_widths = [0, 0, 0]
        row_num = 2
        group_list = [
            'warehouse_id', 'product_id', 'quantity'
        ]
        reports = self.env['report.stock.forecast'].read_group(
            domain=domain, fields=[], groupby=group_list, lazy=False
        )
        for report in reports:
            product = report['product_id']
            warehouse = report['warehouse_id']
            quantity = report['quantity']
            row_values = [warehouse[1], product[1], quantity]
            for col_index, cell_value in enumerate(row_values):
                col_num = col_index + 1
                cell = worksheet.cell(row_num, col_num)
                cell.value = cell_value
                cell.border = border

                cell_width = len(str(cell_value))
                if column_widths[col_index] < cell_width:
                    column_widths[col_index] = cell_width
            row_num += 1

        # make headers
        row_num = 1
        for col_index, header in enumerate(headers):
            col_num = col_index + 1
            cell = worksheet.cell(row_num, col_num)
            cell.value = header
            cell.style = header_bold
            cell.fill = fill

            cell_width = len(header)
            if column_widths[col_index] < cell_width:
                column_widths[col_index] = cell_width
            col_letter = openpyxl.utils.cell.get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = column_widths[col_index] + 0.5

        f = StringIO.StringIO()
        workbook.save(f)
        base64_file = f.getvalue().encode('base64')

        attach_id = self.env['ir.attachment'].create({
            'res_model': 'inventory.forecast.wizard',
            'res_id': self[0].id,
            'type': 'binary',
            'name': 'name.xls',
            'datas_fname': _('Atsargų_prognozė') + '.' + XLS_EXT,
            'datas': base64_file
        })
        return attach_id
        
    @api.multi
    def show_report(self):
        # Check if threading is enabled
        threaded = self.sudo().env.user.company_id.activate_threaded_front_reports
        if threaded:
            return self.action_background_report()
        else:
            return self.action_report(threaded=False)
    
    @api.multi
    def action_report(self, threaded=False):
        domain = []
        attach_id = self.generate_excel(domain)
        if threaded:
            exported_file_name = attach_id.datas_fname
            base64_file = attach_id.datas
            return {
                'exported_file_name': exported_file_name,
                'base64_file': base64_file
            }
        else:
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/binary/download?res_model=inventory.forecast.wizard&res_id=%s&attach_id=%s' % (
                    self[0].id, attach_id.id),
                'target': 'self',
            }

    @job
    @api.multi
    def perform_xlsx_export_job(self, import_job_id, additional_context=None):
        context = self._context.copy()
        if additional_context:
            context.update(additional_context)

        # Re-browse import object
        report_job = self.env['robo.report.job'].browse(import_job_id)
        if not report_job.exists():
            return

        try:
            if not context.get('active_ids'):
                context['active_ids'] = [self.id]
            res = self.with_context(context).action_report(threaded=True)

            base64_file = res.get('base64_file')
            exported_file_name = res.get('exported_file_name')
            exported_file_type = XLS_EXT
        except Exception as exc:
            report_job.write({
                'state': 'failed',
                'fail_message': str(exc.args[0]),
                'execution_end_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT)
            })
        else:
            report_job.write({
                'state': 'succeeded',
                'execution_end_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
                'exported_file': base64_file,
                'exported_file_name': exported_file_name,
                'exported_file_type': exported_file_type
            })
        report_job.post_message()

    @api.multi
    def action_background_report(self):
        user_id = self.env.user.id
        report_name = self.display_name
        now = datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT)

        report_job = self.env['robo.report.job'].create({
            'report_name': report_name,
            'execution_start_date': now,
            'state': 'in_progress',
            'user_id': user_id,
            'job_type': 'export'
        })
        context = self._context.copy()
        
        # Start export job
        self.with_delay(eta=5, channel='root', identity_key=identity_exact).perform_xlsx_export_job(
            report_job.id, additional_context=context
        )
        # Return the action which displays information on where to find the report
        action = self.env.ref('robo.action_open_robo_report_job').read()[0]
        action.update({
            'view_mode': 'form', 'res_id': report_job.id,
            'view_id': self.env.ref('robo.form_robo_report_job').id
        })  # Force form view of the created import job
        return action
    
InventoryForecastWizard()

class StockLandedCost(models.Model):
    _inherit = 'stock.landed.cost'

    state = fields.Selection([
        ('draft', 'Juodraštis'),
        ('done', 'Patvirtintas'),
        ('cancel', 'Atšauktas')], string='Būsena')


StockLandedCost()


class StockPackOperation(models.Model):
    _inherit = 'stock.pack.operation'

    @api.multi
    def show_details(self):
        view_id = self.env.ref('robo_stock.view_pack_operation_details_form_save').id
        return {
            'name': _('Operacijų detalizavimas'),
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'stock.pack.operation',
            'views': [(view_id, 'form')],
            'view_id': view_id,
            'target': 'new',
            'res_id': self.ids[0],
            'context': self.env.context}


StockPackOperation()

stock_mapping = {
    u'Pavadinimas': 'name',
    u'Kodas': 'code',
    u'Data': 'date',
    u'Kiekis': 'qty',
    u'Bendra vertė': 'value',
    u'Sandėlis': 'warehouse',
    u'Sandėlio kodas': 'warehouse_code',
    u'SN': 'sn',
    u'Pakuotė': 'package'
}


def import_stock_file(self, import_file):
    env = self.sudo().env
    location_obj = env['stock.location']
    warehouse_obj = env['stock.warehouse']
    product_obj = env['product.product']
    quant_import_obj = env['quant.import.table']
    imported_quants = env['quant.import.table']
    xls_file = io.BytesIO(base64.decodestring(import_file))
    workbook = px.load_workbook(xls_file)
    sheet = workbook.get_sheet_by_name(name='Atsargos')
    iter_rows = iter(sheet.iter_rows())
    header = []
    header_mapped = []
    errors_general = []
    errors_system = []
    for i, row in enumerate(iter_rows):
        try:
            if not header:
                header = get_all_values(row)
                header_mapped = get_mapped(header, stock_mapping)
                continue
            values = get_all_values(row)
            if len(set(values)) == 1:
                break
            record = ImportRecord(values, header_mapped)
            name = record.name
            code = record.code
            date = record.date
            qty = record.qty
            value = record.value
            warehouse = record.warehouse
            warehouse_code = record.warehouse_code
            location_id = False
            warehouse_id = False
            if warehouse:
                if warehouse_code:
                    warehouse_id = warehouse_obj.search([('code', '=', warehouse_code)], limit=1)
                if not warehouse_id:
                    warehouse_id = warehouse_obj.search([('name', '=', warehouse)], limit=1)
                if not warehouse_id and not warehouse_code:
                    warehouse_code = warehouse.lower()[:5]
                    for n in xrange(10):
                        if not warehouse_obj.search([('code', '=', warehouse_code)], count=True):
                            warehouse_id = warehouse_obj.create({
                                'name': warehouse,
                                'code': warehouse_code,
                                'reception_steps': 'one_step',
                                'delivery_steps': 'ship_only'
                            })
                            break
                        else:
                            warehouse_code = warehouse.lower()[:4] + str(n)
                elif not warehouse_id and warehouse_code:
                    warehouse_id = warehouse_obj.create({
                        'name': warehouse,
                        'code': warehouse_code,
                        'reception_steps': 'one_step',
                        'delivery_steps': 'ship_only'
                    })
                if not warehouse_id:
                    raise exceptions.UserError(_('Nepavyko sukurti naujo sandėlio %s') % warehouse)
                if warehouse_id:
                    location_obj.search([('usage', '=', 'internal'), ('warehouse_id', '=', False)])._warehouse_id()
                    location_id = location_obj.search([('usage', '=', 'internal'),
                                                       ('warehouse_id', '=', warehouse_id.id)], order='id', limit=1)
                    if not location_id:
                        location_id = location_obj.search([('usage', '=', 'transit'),
                                                           ('warehouse_id', '=', warehouse_id.id)], order='id', limit=1)
            if not location_id and (not warehouse and not warehouse_code):
                location_id = location_obj.search([('usage', '=', 'internal')], order='id', limit=1)
            elif not location_id:
                raise exceptions.UserError(_('Nenustatyta lokacija.'))
            if not code and not name:
                raise exceptions.UserError(_('Nenurodyta produkto informacija.'))
            product_id = product_obj
            if code:
                product_id = product_obj.search([('default_code', '=', code)])
            if name:
                if len(product_id) > 1 or not product_id:
                    product_id = product_obj.with_context(lang='lt_LT').search([('name', '=', name)])
                if not product_id:
                    product_id = product_obj.with_context(lang='en_US').search([('name', '=', name)])
            if not product_id:
                raise exceptions.UserError(_('Nerastas produktas.'))
            if len(product_id) > 1:
                raise exceptions.UserError(_('Rasti keli produkto %s atitikmenys.') % code or name)
            if not date or not qty:
                raise exceptions.UserError(_('Nenurodyta reikalaujama informacija.'))
            if not value and type(value) not in [int, float, long]:
                raise exceptions.UserError(_('Nenurodyta reikalaujama informacija.'))
            vals = {
                'product_id': product_id.id,
                'location_id': location_id.id,
                'qty': qty,
                'in_date': date,
                'total_cost_theoretical': value,
            }
            if record.sn:
                vals['serial_number'] = record.sn
            if record.package:
                vals['package'] = record.package
            imported_quants |= quant_import_obj.create(vals)
        except exceptions.UserError as exc:
            errors_general.append(_('%s eilutė %s') % (exc.name, i + 1))
        except exceptions.ValidationError as exc:
            errors_general.append(_('%s eilutė %s') % (exc.name, i + 1))
        except Exception as e:
            errors_system.append(_('%s eilutė %s') % (e, i + 1))
            env.cr.rollback()
    if errors_general:
        raise exceptions.UserError('\n'.join(errors_general))
    if errors_system:
        raise RoboImportError('\n'.join(errors_system))
    if imported_quants:
        try:
            for quant in imported_quants:
                force_date = quant.in_date
                quant.with_context(force_period_date=force_date).add_qtys()
            product_ids = self.env['product.product'].search([('type', '=', 'product')])
            quant_obj = self.env['stock.quant']
            for product_id in product_ids:
                quant_ids = quant_obj.search(
                    [('product_id', '=', product_id.id), ('location_id.usage', '=', 'internal')])
                if quant_ids:
                    total_cost = sum(quant_ids.mapped(lambda r: r.cost * r.qty))
                    total_qty = sum(quant_ids.mapped('qty'))
                    if total_qty:
                        standard_price = total_cost / total_qty
                        product_id.write({'standard_price': standard_price})
        except:
            env.cr.rollback()
            raise exceptions.UserError(_('Nepavyko importuoti kiekių.'))


class CompanySettings(models.TransientModel):
    _inherit = 'robo.company.settings'

    @api.model
    def _default_simplified_stock_default_location(self):
        """Get default location - earliest location with internal usage"""
        return self.env['stock.location'].get_default_location()

    enable_price_lists_front = fields.Boolean(string='Įgalinti kainoraščius',
                                              help='Įgalinti kainoraščius priekinėje vartotojo sąsajoje')
    enable_serial_numbers = fields.Boolean(string='Įgalinti serijos numerius')
    enable_landed_costs = fields.Boolean(string='Įgalinti savikainos koregavimą')
    enable_advanced_stock_routing = fields.Boolean(string='Įgalinti išplėstinį sandėlio maršrutizavimą')

    import_stock = fields.Binary()
    import_pickings = fields.Binary()
    politika_sandelio_apskaita = fields.Selection(
        [('simple', 'Supaprastintas'), ('extended', 'Išplėstinis')], string='Sandėlio apskaitos politika')
    auto_sale_picking_create = fields.Boolean(string='Auto sale picking create', default=True)
    scan_new_serial = fields.Boolean(string='Leisti skenuojant kurti naujus SN', default=True)

    simplified_stock_mode = fields.Selection(
        [('enabled', 'Leisti pardavimus be likučio'),
         ('disabled', 'Neleisti pardavimų be likučio')], string='Supaprastinto sandėlio tipas',
        default='disabled')
    simplified_stock_multi_locations = fields.Boolean(string='Įgalinti skirtingų atsargų vietų pasirinkimus sąskaitose')
    simplified_stock_default_location = fields.Many2one(
        'stock.location', string='Numatytoji atsargų vieta',
        domain="[('usage','=','internal')]",
        default=_default_simplified_stock_default_location
    )
    allow_zero_value_quant = fields.Boolean(string='Leisti produktus su nuline verte')
    default_api_invoice_picking_stock_warehouse = fields.Many2one('stock.warehouse',
                                                                  string='Per API pateiktų sąskaitų atsargas '
                                                                         'nurašyti nuo sandėlio')
    default_stock_surplus_account_id = fields.Many2one('account.account', string='Atsargų pertekliaus sąskaita')
    restore_vat_in_inventory_write_off = fields.Boolean(string='Kurti PVM atstatymo įrašus atsargų nurašymams')
    stock_inventory_require_committee = fields.Boolean(
        string='Reikalauti pasirinkti komisiją atsargų nurašymo aktuose'
    )
    stock_inventory_require_reason_line = fields.Boolean(
        string='Reikalauti pasirinkti nurašymo priežastį atsargų nurašymo aktuose'
    )

    @api.model
    def install_stock_policy(self):
        if self.env.user.sudo().company_id.politika_sandelio_apskaita == 'simple':
            group_extended_id = self.env.ref('robo_stock.robo_stock_extended')
            group_id = self.env.ref('stock_extend.group_robo_stock')
            group_id.sudo().write({
                'implied_ids': [(3, group_extended_id.id)]
            })
            group_extended_id.sudo().write({'users': [(5,)]})

    @api.model
    def default_get(self, field_list):
        res = super(CompanySettings, self).default_get(field_list)
        if self.env.user.is_accountant():
            company = self.sudo().env.user.company_id
            res.update({
                'enable_advanced_stock_routing': company.enable_advanced_stock_routing,
                'politika_sandelio_apskaita': company.politika_sandelio_apskaita,
                'simplified_stock_mode': company.simplified_stock_mode,
                'enable_price_lists_front': company.enable_price_lists_front,
                'enable_serial_numbers': company.enable_serial_numbers,
                'enable_landed_costs': company.enable_landed_costs,
                'scan_new_serial': company.scan_new_serial,
                'allow_zero_value_quant': company.allow_zero_value_quant,
                'default_api_invoice_picking_stock_warehouse': company.default_api_invoice_picking_stock_warehouse.id,
                'simplified_stock_multi_locations': company.simplified_stock_multi_locations,
                'simplified_stock_default_location': company.simplified_stock_default_location.id,
                'default_stock_surplus_account_id': company.default_stock_surplus_account_id.id,
                'restore_vat_in_inventory_write_off': company.restore_vat_in_inventory_write_off,
                'stock_inventory_require_committee': company.stock_inventory_require_committee,
                'stock_inventory_require_reason_line': company.stock_inventory_require_reason_line,
            })
        res['auto_sale_picking_create'] = self.env.user.company_id.auto_sale_picking_create
        return res

    @api.model
    def _get_company_policy_field_list(self):
        res = super(CompanySettings, self)._get_company_policy_field_list()
        res.extend((
            'politika_sandelio_apskaita',
            'enable_price_lists_front',
            'simplified_stock_mode',
            'auto_sale_picking_create',
            'scan_new_serial',
            'allow_zero_value_quant',
            'default_api_invoice_picking_stock_warehouse',
            'enable_serial_numbers',
            'enable_landed_costs',
            'enable_advanced_stock_routing',
            'simplified_stock_multi_locations',
            'simplified_stock_default_location',
            'default_stock_surplus_account_id',
            'restore_vat_in_inventory_write_off',
        ))
        return res

    @api.model
    def _get_company_info_field_list(self):
        res = super(CompanySettings, self)._get_company_info_field_list()
        res.extend([
            'auto_sale_picking_create',
            'stock_inventory_require_committee',
            'stock_inventory_require_reason_line',
        ])
        return res

    @api.multi
    def set_default_import(self):
        """
        Calls threaded import preparation method
        for stock data.
        :return: None
        """
        super(CompanySettings, self).set_default_import()
        self.threaded_import_prep(
            'import_stock',
            function=import_stock_file,
            imported_file=self.import_stock
        )
        self.threaded_import_prep(
            'import_pickings',
            function=import_pickings,
            imported_file=self.import_pickings
        )


CompanySettings()


class ResCompany(models.Model):
    _inherit = 'res.company'

    enable_price_lists_front = fields.Boolean(string='Įgalinti kainoraščius',
                                              inverse='_set_enable_price_lists_front',
                                              help='Įgalinti kainoraščius priekinėje vartotojo sąsajoje')

    enable_serial_numbers = fields.Boolean(string='Įgalinti serijos numerius', inverse='_enable_serial_numbers')

    enable_landed_costs = fields.Boolean(string='Įgalinti savikainos koregavimą', inverse='_enable_landed_costs')
    enable_advanced_stock_routing = fields.Boolean(
        string='Įgalinti išplėstinį sandėlio maršrutizavimą',
        inverse='_set_enable_extended_stock_routing'
    )

    politika_sandelio_apskaita = fields.Selection(
        [('simple', 'Supaprastintas'), ('extended', 'Išplėstinis')], string='Sandėlio apskaitos politika',
        inverse='_set_politika_sandelio_apskaita',
        default='simple')
    scan_new_serial = fields.Boolean(string='Leisti skenuojant kurti naujus SN', default=True)

    simplified_stock_mode = fields.Selection(
        [('enabled', 'Leisti pardavimus be likučio'),
         ('disabled', 'Neleisti pardavimų be likučio')], string='Supaprastinto sandėlio tipas',
        default='disabled')

    default_api_invoice_picking_stock_warehouse = fields.Many2one(
        'stock.warehouse', string='Per API pateiktų sąskaitų atsargas nurašyti nuo sandėlio'
    )
    simplified_stock_multi_locations = fields.Boolean(
        string='Įgalinti papildomus lokacijų pasirinkimus',
        inverse='_set_simplified_stock_multi_locations'
    )
    simplified_stock_default_location = fields.Many2one(
        'stock.location', string='Numatytoji lokacija',
    )
    default_stock_surplus_account_id = fields.Many2one('account.account', string='Atsargų pertekliaus sąskaita')
    restore_vat_in_inventory_write_off = fields.Boolean(string='Kurti PVM atstatymo įrašus atsargų nurašymams')
    stock_inventory_require_committee = fields.Boolean(compute='_compute_stock_inventory_require_committee',
                                                       inverse='_set_stock_inventory_require_committee')
    stock_inventory_require_reason_line = fields.Boolean(compute='_compute_stock_inventory_require_reason_line',
                                                         inverse='_set_stock_inventory_require_reason_line')

    @api.model_cr
    def init(self):
        """
        Set default company settings values when robo_stock is installed
        :return: None
        """
        companies = self.env['res.company'].search([])
        for company in companies:
            if not company.politika_sandelio_apskaita:
                company.politika_sandelio_apskaita = 'simple'
            if not company.simplified_stock_mode:
                company.simplified_stock_mode = 'disabled'

    @api.multi
    def _compute_stock_inventory_require_committee(self):
        self.stock_inventory_require_committee = self.env['ir.config_parameter'].sudo().get_param('stock_inventory_require_committee') == 'True'

    @api.multi
    def _set_stock_inventory_require_committee(self):
        self.env['ir.config_parameter'].sudo().set_param('stock_inventory_require_committee', str(self.stock_inventory_require_committee))

    @api.multi
    def _compute_stock_inventory_require_reason_line(self):
        self.stock_inventory_require_reason_line = self.env['ir.config_parameter'].sudo().get_param('stock_inventory_require_reason_line') == 'True'

    @api.multi
    def _set_stock_inventory_require_reason_line(self):
        self.env['ir.config_parameter'].sudo().set_param('stock_inventory_require_reason_line', str(self.stock_inventory_require_reason_line))

    @api.multi
    def _set_enable_price_lists_front(self):
        """ Add / remove user from group on setting the field """
        self.ensure_one()
        if self.enable_price_lists_front:
            group_pricelist_id = self.env.ref('robo_stock.group_robo_front_pricelist')
            group_id = self.env.ref('stock_extend.group_robo_stock')
            group_id.sudo().write({
                'implied_ids': [(4, group_pricelist_id.id)]
            })
        else:
            group_pricelist_id = self.env.ref('robo_stock.group_robo_front_pricelist')
            group_id = self.env.ref('stock_extend.group_robo_stock')
            group_id.sudo().write({
                'implied_ids': [(3, group_pricelist_id.id)]
            })
            group_pricelist_id.sudo().write({'users': [(5,)]})

    @api.one
    def _enable_serial_numbers(self):
        serial_numbers_group = self.env.ref('robo_stock.group_robo_serial_numbers')
        production_lot_group = self.env.ref('stock.group_production_lot')
        groups = self.env.ref('stock_extend.group_robo_stock')
        if self.enable_serial_numbers:
            groups.sudo().write({'implied_ids': [(4, serial_numbers_group.id),
                                                 (4, production_lot_group.id)]})
        else:
            groups.sudo().write({ 'implied_ids': [(3, serial_numbers_group.id),
                                                  (3, production_lot_group.id)]})
            production_lot_group.write({'users': [(5,)]})
            serial_numbers_group.write({'users': [(5,)]})

    @api.one
    def _enable_landed_costs(self):
        landed_costs_group = self.env.ref('robo_stock.group_robo_landed_costs')
        groups = self.env.ref('stock.group_stock_manager')
        if self.enable_landed_costs:
            groups.sudo().write({'implied_ids': [(4, landed_costs_group.id)]})
        else:
            groups.sudo().write({'implied_ids': [(3, landed_costs_group.id)]})
            landed_costs_group.write({'users': [(5,)]})

    @api.multi
    def _set_simplified_stock_multi_locations(self):
        """
        Inverse //
        Add simplified stock multi location group
        to all users on activation
        :return: None
        """
        # Reference needed groups
        multi_loc_group = self.sudo().env.ref('robo_stock.group_simplified_stock_multi_locations')
        user_group = self.sudo().env.ref('base.group_user')

        for rec in self:
            if rec.simplified_stock_multi_locations:
                user_group.write({'implied_ids': [(4, multi_loc_group.id)]})
            else:
                # On deactivation, remove the inheritance, and clear the users
                user_group.write({'implied_ids': [(3, multi_loc_group.id)]})
                multi_loc_group.write({'users': [(5, )]})
                # Clear the default location
                rec.write({'simplified_stock_default_location': False})

    @api.multi
    def _set_enable_extended_stock_routing(self):
        """
        Inverse //
        Add the advanced routing group to the stock
        manager on inverse.
        :return: None
        """
        # Reference needed groups
        adv_routing_group = self.sudo().env.ref('stock.group_adv_location')
        stock_manager_group = self.sudo().env.ref('stock.group_stock_manager')

        for rec in self:
            if rec.enable_advanced_stock_routing:
                stock_manager_group.write({'implied_ids': [(4, adv_routing_group.id)]})
            else:
                # On deactivation, remove the inheritance, and clear the users
                stock_manager_group.write({'implied_ids': [(3, adv_routing_group.id)]})
                adv_routing_group.write({'users': [(5, )]})

    @api.multi
    def _set_politika_sandelio_apskaita(self):
        """ Add / remove user to groups depending on company warehouse accounting mode """
        self.ensure_one()
        if self.politika_sandelio_apskaita == 'extended':
            group_extended_id = self.env.ref('robo_stock.robo_stock_extended')
            group_id = self.env.ref('stock_extend.group_robo_stock')
            group_id.sudo().write({
                'implied_ids': [(4, group_extended_id.id)]
            })
        else:
            group_extended_id = self.env.ref('robo_stock.robo_stock_extended')
            group_id = self.env.ref('stock_extend.group_robo_stock')
            group_id.sudo().write({
                'implied_ids': [(3, group_extended_id.id)]
            })
            premium_accountant_group_id = self.env.ref('robo_basic.group_robo_premium_accountant')
            user_ids = self.env['res.users'].search([('groups_id', 'in', premium_accountant_group_id.ids)]).ids
            group_extended_id.sudo().write({'users': [(5,), (6, 0, user_ids)]})

    @api.constrains('politika_sandelio_apskaita', 'simplified_stock_multi_locations')
    def _check_location_number_for_simple_accounting(self):
        for rec in self:
            if rec.politika_sandelio_apskaita != 'extended' and not rec.simplified_stock_multi_locations:
                if self.env['stock.warehouse'].search([('company_id', '=', rec.id)], count=True) > 1:
                    raise exceptions.UserError(_('Supaprastinta sandėlio versija suderinama tik su vienu sandėliu.'))

    @api.multi
    @api.constrains('simplified_stock_default_location')
    def _check_simplified_stock_default_location(self):
        """
        Constraints //
        Ensure that simplified_stock_default_location
        is of 'internal' type
        :return: None
        """
        for rec in self:
            location = rec.simplified_stock_default_location
            if location and rec.simplified_stock_multi_locations:
                if location.usage != 'internal':
                    raise exceptions.ValidationError(
                        _('Numatytosios atsargų vietos privalo būti vidinis'))
                if not location.warehouse_id:
                    raise exceptions.ValidationError(
                        _('Nunustatytas numatytosios atsargų vietos sandėlis'))


ResCompany()


class StockPackOpExt(models.Model):
    _name = 'stock.pack.operation'
    _inherit = ['stock.pack.operation', 'barcodes.barcode_events_mixin']

    scan_status_text = fields.Char(string='Skenavimo statusas', groups='stock.group_production_lot', store=False)
    scan_status = fields.Integer(string='Indikatorius', groups='stock.group_production_lot', store=False)
    barcode = fields.Char(string='Barkodas', groups='stock.group_production_lot', store=False)

    @api.onchange('barcode')
    def onchange_barcode(self):
        if self.barcode:
            self.on_barcode_scanned(self.barcode)
            self.barcode = ''

    @api.one
    def on_barcode_scanned(self, barcode):
        if not barcode:
            return
        if not self.env.user.has_group('stock.group_production_lot'):
            return
        if self.picking_id.state == 'done':
            return
        barcode = barcode.strip()
        product_serial = self.env['stock.production.lot'].search([('name', '=', barcode),
                                                                  ('product_id', '=', self.product_id.id)], limit=1)
        if product_serial:
            serial_id = product_serial.id
            if 'default_picking_id' in self._context and serial_id in self.env['stock.picking'].browse(
                    self._context['default_picking_id']).pack_operation_pack_ids.filtered(
                    lambda r: r.processed_boolean).mapped('package_id.quant_ids.lot_id.id'):
                self.scan_status_text = _('Nuskenuotas <%s> SN jau pridėtas pakuotėje.') % barcode
                self.scan_status = 0
                return
            elif 'default_picking_id' in self._context and serial_id in self.env['stock.picking'].browse(
                    self._context['default_picking_id']).pack_operation_product_ids.mapped('pack_lot_ids').filtered(
                    lambda r: r.qty > 0).mapped('lot_id.id'):
                self.scan_status_text = _('Nuskenuotas <%s> SN jau pridėtas.') % barcode
                self.scan_status = 0
                return
            for line in self.pack_lot_ids:
                if line.lot_id.id == serial_id:
                    if line.qty >= line.qty_todo:
                        self.scan_status_text = _('SN <%s> jau nuskenuotas.') % product_serial.name
                        self.scan_status = 0
                        return
                    else:
                        line.qty += 1
                        self.scan_status_text = _('Sėkmingai pridėtas <%s> SN numeris.') % product_serial.name
                        self.scan_status = 1
                        return
            quants = self.env['stock.quant'].search([('lot_id', '=', serial_id)])
            if quants:
                lot = quants.filtered(lambda r: r.product_id.id == self.product_id.id)
                if not lot:
                    self.scan_status_text = (_('Skenuotas <%s> SN priklauso kitam produktui (%s).')
                                             % (barcode, quants[0].product_id.name))
                    self.scan_status = 0
                    return
                lot = quants.filtered(lambda r: r.location_id.id == self.location_id.id)
                if not lot and self.location_id:
                    scan_status_text = (_('Skenuotas <%s> SN yra kitoje lokacijoje (%s).')
                                        % (barcode, quants[0].location_id.display_name))
                    if self.env.user.is_accountant():
                        scan_status_text += '\n%s (ID: %s) != %s (ID: %s)' % (quants[0].location_id.display_name,
                                                                              quants[0].location_id.id,
                                                                              self.location_id.display_name,
                                                                              self.location_id.id)
                    self.scan_status_text = scan_status_text
                    self.scan_status = 0
                    return
                # we do not care about reservation
                #
                # lot = quants.filtered(lambda r: r.reservation_id)
                # if lot:
                #     move_name = lot.reservation_id.name
                #     if lot.reservation_id.picking_id:
                #         move_name += ' - ' + lot.reservation_id.picking_id.name
                #     self.scan_status_text = ('Scanned <%s> serial is reserved for move (%s).' % (barcode, move_name))
                #     self.scan_status = 0
                #     return
                vals = {'lot_id': product_serial.id,
                        'qty_todo': 1,
                        'qty': 1,
                        'lot_name': product_serial.name}
                self.pack_lot_ids |= self.pack_lot_ids.new(vals)
                self.scan_status_text = ('Pridėtas naujas SN <%s>.' % product_serial.name)
                self.scan_status = 2
                return
            else:
                if product_serial.product_id.id != self.product_id.id:
                    self.scan_status_text = (_('Skenuotas SN <%s> priskirtas kitam produktui (%s).')
                                             % (barcode, product_serial.product_id.name))
                    self.scan_status = 0
                    return
                vals = {'lot_id': product_serial.id,
                        'qty_todo': 1,
                        'qty': 1,
                        'lot_name': product_serial.name}
                self.pack_lot_ids |= self.pack_lot_ids.new(vals)
                self.scan_status_text = _('Pridėtas naujas SN <%s>.') % product_serial.name
                self.scan_status = 2
                return
        else:
            products = self.env['product.product'].search([('barcode', '=', barcode)], limit=1)
            if products:
                self.scan_status_text = _('Nuskenuotas produkto barkodas <%s>, turėtumėte nuskenuoti SN.') % barcode
                self.scan_status = 0
                return
            if not self.env.user.company_id.scan_new_serial:
                self.scan_status_text = _('Nuskenuotas SN <%s> nerastas.') % barcode
                self.scan_status = 0
                return
            if self._context.get('no_create', False):
                return
            self.create_lot(barcode)

    def create_lot(self, barcode):
        cdate = datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT)
        self.sudo()._cr.execute(
            u'''INSERT INTO stock_production_lot (create_date, write_date, create_uid, write_uid, name, product_id) VALUES ('%s', '%s', %s, %s, '%s', %s)''' %
            (cdate, cdate, self._uid, self._uid, barcode, self.product_id.id))
        new_serial = self.env['stock.production.lot'].search([('name', '=', barcode),
                                                              ('product_id', '=', self.product_id.id)
                                                              ], limit=1)
        vals = {'lot_id': new_serial.id,
                'qty_todo': 1,
                'qty': 1,
                'lot_name': new_serial.name}
        self.pack_lot_ids |= self.pack_lot_ids.new(vals)
        self.scan_status_text = _('Naujas SN <%s> pridėtas.') % new_serial.name
        self.scan_status = 2


StockPackOpExt()


class ReturnPicking(models.TransientModel):
    _inherit = 'stock.return.picking'

    def create_returns(self):
        res = super(ReturnPicking, self).create_returns()
        res['view_id'] = self.env.ref('robo_stock.robo_stock_picking_form').id
        res['view_mode'] = 'form'
        return res


ReturnPicking()


class PurchaseOrderLine(models.Model):
    _inherit = 'purchase.order.line'

    product_uom_categ_id = fields.Many2one('product.uom.categ', related='product_id.uom_id.category_id', readonly=True)

    @api.model
    def create(self, vals):
        if 'taxes_id' in vals:
            if len(vals['taxes_id']) > 0:
                tax_ids_set = set()
                for line in vals['taxes_id']:
                    if line[0] == 6:
                        tax_ids_set = set(self.env['account.tax'].browse(line[2]).mapped('id'))
                    elif line[0] == 4:
                        tax_ids_set.add(self.env['account.tax'].browse(line[1]).id)
                    else:
                        continue
                tax_ids_obj = self.env['account.tax'].browse(list(tax_ids_set))
                child_ids = tax_ids_obj.mapped('child_tax_ids.id')
                tax_ids = tax_ids_obj.mapped('id')
                all_ids = list(set(tax_ids + child_ids))
                new_vals = [(6, 0, all_ids)]
                vals['taxes_id'] = new_vals
        return super(PurchaseOrderLine, self).create(vals)

    @api.multi
    def write(self, vals):
        if 'taxes_id' in vals:
            if len(vals['taxes_id']) > 0:
                tax_ids_set = set()
                for line in vals['taxes_id']:
                    if line[0] == 6:
                        tax_ids_set = set(self.env['account.tax'].browse(line[2]).mapped('id'))
                    elif line[0] == 4:
                        tax_ids_set.add(self.env['account.tax'].browse(line[1]).id)
                    else:
                        continue
                tax_ids_obj = self.env['account.tax'].browse(list(tax_ids_set))
                child_ids = tax_ids_obj.mapped('child_tax_ids.id')
                tax_ids = tax_ids_obj.mapped('id')
                all_ids = list(set(tax_ids + child_ids))
                new_vals = [(6, 0, all_ids)]
                vals['taxes_id'] = new_vals
        res = super(PurchaseOrderLine, self).write(vals)
        return res


PurchaseOrderLine()


class WizardValuationHistory(models.TransientModel):
    _inherit = 'wizard.valuation.history'

    @api.multi
    def open_table(self):
        self.ensure_one()
        ctx = dict(
            self._context,
            history_date=self.date,
            search_default_product_wh=True,
            search_default_group_by_product=True,
            search_default_group_by_location=True)

        action = self.env['ir.model.data'].xmlid_to_object('stock_account.action_stock_history')
        if not action:
            action = {
                'view_type': 'form',
                'view_mode': 'tree,graph,pivot',
                'res_model': 'stock.history',
                'type': 'ir.actions.act_window',
            }
        else:
            action = action[0].read()[0]

        action['domain'] = "[('date', '<=', '" + self.date + "')]"
        action['display_name'] = action['name'] = _('Stock Value At Date') + ' -- {}'.format(self.date)
        action['context'] = ctx
        return action


WizardValuationHistory()


class StockAccountingCompareReport(models.TransientModel):
    _name = 'stock.accounting.compare.report'

    def _default_date_to(self):
        return datetime.utcnow() + relativedelta(day=31)

    def _default_date_from(self):
        return datetime.utcnow() - relativedelta(day=1)

    date_from = fields.Date(string='Data nuo', default=_default_date_from, required=True)
    date_to = fields.Date(string='Data iki', default=_default_date_to, required=True)

    @api.multi
    def generate_report(self):
        header = ['Sąskaitos data', 'Sąskaitos Nr.', 'Važtaraščio numeris', 'Sandėlio vertė',
                  'Koregavimų vertė', 'Apskaitos vertė', 'Skirtumas', 'Galutinis skirtumas']

        stock_moves = self.env['stock.move'].search([('date', '>=', self.date_from),
                                                     ('date', '<=', self.date_to),
                                                     '|',
                                                     ('location_id.usage', '!=', 'internal'),
                                                     ('location_dest_id.usage', '!=', 'internal')])
        stock_moves = stock_moves.filtered(lambda r: r.location_id.usage == 'internal' or r.location_dest_id.usage == 'internal')
        picking_ids = stock_moves.mapped('picking_id')
        used_invoice_ids = []
        used_aml_ids = []
        landed_cost_ids = []
        data = []
        for picking_id in picking_ids:
            if picking_id.invoice_ids:
                corresponding_ids = picking_id.invoice_ids
            elif picking_id.sale_id:
                corresponding_ids = picking_id.sale_id.invoice_ids
            elif picking_id.purchase_id:
                corresponding_ids = picking_id.purchase_id.invoice_ids
            elif picking_id.origin:
                corresponding_ids = self.env['account.invoice'].search([('number', '=', picking_id.origin)])
            else:
                corresponding_ids = self.env['account.invoice']
            corresponding_ids = corresponding_ids.filtered(lambda x: x.state in ['open', 'paid'])
            inventory_value = abs(tools.float_round(
                sum(x.inventory_value for x in picking_id.quant_ids), precision_digits=2))
            adjusted_value = 0.0
            landed_cost_ids += picking_id.quant_ids.mapped('valuation_adjustment_ids.cost_id.id')
            account_ids = corresponding_ids.mapped('invoice_line_ids.product_id.categ_id.property_stock_valuation_account_id')
            if not account_ids:
                account_ids = self.env['account.account'].search([('code', '=', '2040')])

            lc_aml_ids = picking_id.mapped('quant_ids.valuation_adjustment_ids.cost_id.account_move_id.line_ids').filtered(lambda x: x.account_id in account_ids)
            for adjust_id in picking_id.mapped('quant_ids.valuation_adjustment_ids'):
                in_pick = picking_id.location_dest_id.usage == 'internal'
                accounting_line = adjust_id.cost_id.account_move_id.line_ids.filtered(
                    lambda l: adjust_id.name in l.name
                              and l.product_id.id == adjust_id.product_id.id
                              and tools.float_compare(abs(l.debit - l.credit), abs( adjust_id.additional_landed_cost * (l.quantity / adjust_id.quantity)), precision_digits=2) == 0
                              and l in lc_aml_ids
                )
                if in_pick:
                    accounting_line = accounting_line.filtered(lambda l: 'jau nebėra' not in l.name)
                else:
                    accounting_line = accounting_line.filtered(lambda l: 'jau nebėra' in l.name)

                if accounting_line:
                    adjusted_value += (accounting_line[0].debit - accounting_line[0].credit) * (1 if in_pick else -1) * sum(q.qty for q in picking_id.quant_ids if adjust_id in q.valuation_adjustment_ids) / accounting_line[0].quantity
                    lc_aml_ids -= accounting_line[0]

            aml = self.env['account.move.line'].search([('ref', '=', picking_id.name), ('account_id', 'in', account_ids.ids)])
            used_aml_ids += aml.ids
            accounting_value = abs(sum(x.balance for x in aml))
            display_name = str()
            invoice_date = False
            for ind, x in enumerate(corresponding_ids):
                if ind:
                    display_name += ', '
                display_name += x.computed_number
                if not invoice_date:
                    invoice_date = x.date_invoice
            diff = tools.float_round(inventory_value - accounting_value - adjusted_value, precision_digits=2)
            accounting_wo_adjustments = tools.float_round(diff + adjusted_value, precision_digits=2)
            if abs(accounting_wo_adjustments) > abs(diff):
                final_diff = diff
            else:
                final_diff = accounting_wo_adjustments
            if abs(final_diff) < 0.03:
                continue
            data.append(
                [invoice_date or '-', display_name, picking_id.display_name, inventory_value, adjusted_value, accounting_value, diff, final_diff]
            )
            used_invoice_ids += corresponding_ids.ids
        # invoices = self.env['account.invoice'].search([('date_invoice', '>=', self.date_from),
        #                                                ('date_invoice', '<=', self.date_to),
        #                                                ('id', 'not in', used_invoice_ids),
        #                                                ('invoice_line_ids.product_id.type', '=', 'product'),
        #                                                ('state', 'in', ['open', 'paid'])])
        # for invoice_id in invoices:
        #     data.append([invoice_id.date_invoice, invoice_id.computed_number])
        landed_cost_ids = self.env['stock.landed.cost'].browse(list(set(landed_cost_ids)))
        skip_move_ids = landed_cost_ids.mapped('account_move_id.id')
        stock_account_ids = self.env['product.category'].search([
            ('property_stock_valuation_account_id', '!=', False)]).mapped('property_stock_valuation_account_id')
        aml_ids = self.env['account.move.line'].search([
            ('id', 'not in', used_aml_ids),
            ('move_id', 'not in', skip_move_ids),
            ('account_id', 'in', stock_account_ids.ids),
            ('date', '>=', self.date_from),
            ('date', '<=', self.date_to)], order='date')
        for aml_id in aml_ids:
            data.append([aml_id.date, aml_id.ref, aml_id.name, 0, 0, aml_id.balance])

        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet(_('Sandėlio/Apskaitos palyginimo ataskaita'))

        # Write header
        header_bold_brd = xlwt.easyxf(
            "font: bold on; borders: left thin, right thin, bottom thin ")

        col = 0
        for val in header:
            worksheet.write(0, col, val, header_bold_brd)
            worksheet.col(col).width = 256 * 20
            col += 1
        for row, line in enumerate(data, 1):
            for col, val in enumerate(line):
                worksheet.write(row, col, val)

        worksheet.set_panes_frozen(True)
        worksheet.set_horz_split_pos(1)
        f = StringIO.StringIO()
        workbook.save(f)
        base64_file = f.getvalue().encode('base64')

        attach_id = self.env['ir.attachment'].create({
            'res_model': 'stock.accounting.compare.report',
            'res_id': self[0].id,
            'type': 'binary',
            'name': 'name.xls',
            'datas_fname': _('Apskaitos_sandėlio_palyginimas_' + self.date_from + '_' + self.date_to + '.xls'),
            'datas': base64_file
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/download?res_model=stock.accounting.compare.report&res_id=%s&attach_id=%s' % (
                self[0].id, attach_id.id),
            'target': 'self',
        }


StockAccountingCompareReport()
