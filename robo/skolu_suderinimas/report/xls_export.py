# -*- coding: utf-8 -*-
import StringIO
import cStringIO as StringIO
import math
import os
from datetime import timedelta
from sys import platform
from six import iteritems, itervalues

from odoo import _, models, api, exceptions, fields
from odoo.tools import float_compare, float_is_zero, float_round
from odoo.tools.misc import formatLang
import openpyxl as px
from openpyxl.styles import Font, Border, Side, NamedStyle, colors, Alignment
from debt_reconciliation_base import get_data_by_account_code

kwd_mark = object()
cache_styles = {}

STATIC_FILE_DIR = '\\static\\src\\xlsx\\' if platform == 'win32' else '/static/src/xlsx/'
XLS_EXT = '.xlsx'

def getRowHeightNeeded(str, num):
    """ HACK: Based on how many chars fit into single cell, return needed height of cell for all the text to fit in """
    STD_HEIGHT = 12
    if isinstance(str, int):
        if str <= num:
            return STD_HEIGHT
        else:
            return int(STD_HEIGHT * (math.ceil(str / float(num))))
    else:
        if len(str) <= num:
            return STD_HEIGHT
        else:
            return int(STD_HEIGHT * (math.ceil(len(str) / float(num))))

def _format_number(number):
    return float_round(number, precision_digits=2)
class DebtActWizardExcel:
    def __init__(self):
        self.row_num = 1
        self.main_template = False
        self.main_sheet = False
        self.main_template_name = ''
        self.lang = False
        self.PARTNER_WIDTH = 45
        self.DOC_NUM_WIDTH = 28
        
        self.init_cell_styles()

    def init_cell_styles(self):
        """ initialize style values of 4 differents of cells """
        font = Font(size=7, color=colors.BLACK)
        thin = Side(style="thin", color=colors.BLACK)
        medium = Side(style="medium", color=colors.BLACK)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.sched_cell_style_thin_left = NamedStyle(
            name="Sched Cell Style Thin Left", font=font, border=border,
            alignment=Alignment(horizontal="left", vertical="center", wrapText=True)
        )
        self.sched_cell_style_thin_right = NamedStyle(
            name="Sched Cell Style Thin Right", font=font, border=border,
            alignment=Alignment(horizontal="right", vertical="center", wrapText=True)
        )

        font = font = Font(size=7, color=colors.BLACK, bold=True)
        border = Border(left=thin, right=thin, top=medium, bottom=thin)
        self.sched_cell_style_medium_left = NamedStyle(
            name="Sched Cell Style Medium Left", font=font, border=border,
            alignment=Alignment(horizontal="left", vertical="center", wrapText=True)
        )
        self.sched_cell_style_medium_right = NamedStyle(
            name="Sched Cell Style Medium Right", font=font, border=border,
            alignment=Alignment(horizontal="right", vertical="center", wrapText=True)
        )

    def load_main_template(self, template_name):
        xls_location = STATIC_FILE_DIR + template_name + XLS_EXT
        file_loc = os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + xls_location
        self.main_template = px.load_workbook(file_loc)
        self.main_sheet = self.main_template.active

    def _set_template_names(self, detail_level, show_accounts, show_original_amounts):
        """ Set names for main and line templates """
        if detail_level == 'sum':
            main_template_name = 'Header_Sums_Accounts' if show_accounts else 'Header_Sums_No_Accounts'
        else:
            if show_original_amounts:
                main_template_name = 'Header_Details_Original_Accounts' if show_accounts \
                    else 'Header_Details_Original_No_Accounts'
            else:
                main_template_name = 'Header_Details_No_Original_Accounts' if show_accounts \
                    else 'Header_Details_No_Original_No_Accounts'
        
        if self.lang == 'en_US':
            main_template_name += '_en'
        self.main_template_name = main_template_name
        
    def load_top(self, date, date_from, date_to, type, show_original_amounts, show_accounts, detail_level, lang=None):
        self._set_template_names(detail_level, show_accounts, show_original_amounts)
        self.lang = lang
        self.load_main_template(self.main_template_name)
        main_sheet = self.main_sheet
        self.first_loop = False
        self.row_num += 1

        if not type == 'all':
            main_sheet.cell(self.row_num, 1).value = date
        else:
            main_sheet.cell(self.row_num, 1).value = date_from
            main_sheet.cell(self.row_num, 2).value = date_to

        self.row_num += 2

    def write_partner_line_sums(self, partner_name, amount_sum, cur_str, account_name=False):
        main_sheet = self.main_sheet
        sched_cell_style_left = self.sched_cell_style_thin_left
        sched_cell_style_right = self.sched_cell_style_thin_right

        col_num= 1
        if account_name:
            cell = main_sheet.cell(self.row_num, col_num)
            cell.value = unicode(account_name)
            cell.style = sched_cell_style_left
            col_num+= 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode(partner_name)
        cell.style = sched_cell_style_left
        col_num+= 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = amount_sum
        cell.style = sched_cell_style_right
        col_num+= 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode(cur_str)
        cell.style = sched_cell_style_right

        main_sheet.row_dimensions[self.row_num].height = getRowHeightNeeded(len(unicode(account_name)), self.PARTNER_WIDTH)
        self.row_num += 1

    def write_partner_line_details(self, partner_name, debit, credit, debt, original_amounts=False, accounts=False):
        main_sheet = self.main_sheet
       
        sched_cell_style_left = self.sched_cell_style_medium_left
        sched_cell_style_right = self.sched_cell_style_medium_right

        col_num = 1
        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode(partner_name)
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        if accounts:
            cell = main_sheet.cell(self.row_num, col_num)
            cell.value = unicode('')
            cell.style = sched_cell_style_left
            col_num += 1

        if original_amounts:
            cell = main_sheet.cell(self.row_num, col_num)
            cell.value = unicode('')
            cell.style = sched_cell_style_left
            col_num += 1

            cell = main_sheet.cell(self.row_num, col_num)
            cell.value = unicode('')
            cell.style = sched_cell_style_left
            col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = debit
        cell.style = sched_cell_style_right
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = credit
        cell.style = sched_cell_style_right
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = debt
        cell.style = sched_cell_style_right
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left

        main_sheet.row_dimensions[self.row_num].height = getRowHeightNeeded(len(unicode(partner_name)), self.PARTNER_WIDTH)
        self.row_num += 1

    def write_account_line_details(self, account_name):
        main_sheet = self.main_sheet
        
        sched_cell_style_left = self.sched_cell_style_medium_left
        
        col_num = 1
        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode(account_name)
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        cell = main_sheet.cell(self.row_num, col_num)
        cell.value = unicode('')
        cell.style = sched_cell_style_left
        col_num += 1

        main_sheet.row_dimensions[self.row_num].height = getRowHeightNeeded(len(unicode(account_name)), self.PARTNER_WIDTH)
        self.row_num += 1

    def write_info_line_details(self, data):
        main_sheet = self.main_sheet
        sched_cell_style_left = self.sched_cell_style_thin_left
        sched_cell_style_right = self.sched_cell_style_thin_right
        
        col = 1
        for item in data:
            if col < 3:
                sched_cell_style = sched_cell_style_left
            else:
                sched_cell_style = sched_cell_style_right
            if type(item) is not float:
                item = unicode(item)
            cell = main_sheet.cell(self.row_num, col)
            cell.value = item
            cell.style = sched_cell_style
            col += 1
        main_sheet.row_dimensions[self.row_num].height = getRowHeightNeeded(max(len(unicode(data[0])), len(unicode(data[4]))), self.DOC_NUM_WIDTH)
        self.row_num += 1

    def export(self):
        f = StringIO.StringIO()
        self.main_template.save(f)
        return f.getvalue().encode('base64')

class AccReportGeneralLedgerSL(models.TransientModel):
    _inherit = "debt.act.wizard"

    @api.multi
    def export_excel(self, data):
        Partner = self.env['res.partner']
        date = self.date
        type = self.type
        date_from = self.date_from
        date_to = self.date_to
        detail_level = self.detail_level
        show_original_amounts = self.show_original_amounts
        show_accounts = self.show_accounts
        lang = self._context.get('lang') or 'lt_LT'

        data = self.render_html(doc_ids=False, data=data)

        excel = DebtActWizardExcel()

        excel.load_top(date, date_from, date_to, type, show_original_amounts, show_accounts, detail_level, lang)

        self = self.with_context(lang=excel.lang)

        if detail_level == 'sum' and show_accounts:  # Special case when data is grouped by account at first
            for account_key, account_value in iteritems(data['data_by_account_code']):
                excel.write_account_line_details(self._get_account_name(account_key))
                for partner_key, partner_value in iteritems(account_value):
                    partner = Partner.browse(partner_key)
                    for currency_key, currency_value in iteritems(partner_value):
                        currency_data = self._get_bare_currency_data(currency_key)
                        balance = currency_value['debit'] - currency_value['credit']
                        excel.write_partner_line_sums(partner.display_name, _format_number(balance),
                                                      currency_data['display_name'],
                                                      self._get_account_name(account_key))
        else:
            for partner in data['docs']:
                partner_data = data['data_by_partner_id'][partner.id]
                for currency in partner_data:
                    currency_data = self._get_bare_currency_data(currency)
                    debit = partner_data[currency]['debit']
                    credit = partner_data[currency]['credit']
                    skola = debit-credit
                    if detail_level == 'sum':
                        excel.write_partner_line_sums(partner.display_name, _format_number(skola),
                                                      currency_data['display_name'])
                    else:
                        excel.write_partner_line_details(partner.display_name, _format_number(debit),
                                                         _format_number(credit), _format_number(skola),
                                                         show_original_amounts, show_accounts)
                        for line in partner_data[currency]['lines']:
                            original_value_currency_data = self._get_bare_currency_data(line['orig_currency'].id)
                            skolos_likutis = line['debit']-line['credit']
                            if show_original_amounts:
                                if show_accounts:
                                    write_data = [line['date'], line['date_operation'], line['date_maturity'],
                                                  partner.get_doc_type(line['doc_type']), line['ref'],
                                                  self._get_account_name(line['account_code']),
                                                  _format_number(line['orig_amount']),
                                                  original_value_currency_data['display_name'],
                                                  _format_number(line['debit']), _format_number(line['credit']),
                                                  _format_number(skolos_likutis), currency_data['display_name']]
                                else:
                                    write_data = [line['date'], line['date_operation'], line['date_maturity'],
                                                  partner.get_doc_type(line['doc_type']), line['ref'],
                                                  _format_number(line['orig_amount']),
                                                  original_value_currency_data['display_name'],
                                                  _format_number(line['debit']), _format_number(line['credit']),
                                                  _format_number(skolos_likutis), currency_data['display_name']]
                            else:
                                if show_accounts:
                                    write_data = [line['date'], line['date_operation'], line['date_maturity'],
                                                  partner.get_doc_type(line['doc_type']), line['ref'],
                                                  self._get_account_name(line['account_code']),
                                                  _format_number(line['debit']), _format_number(line['credit']),
                                                  _format_number(skolos_likutis), currency_data['display_name']]
                                else:
                                    write_data = [line['date'], line['date_operation'], line['date_maturity'],
                                                  partner.get_doc_type(line['doc_type']), line['ref'],
                                                  _format_number(line['debit']), _format_number(line['credit']),
                                                  _format_number(skolos_likutis), currency_data['display_name']]

                            excel.write_info_line_details(write_data)

        base64_file = excel.export()
        if type == 'all':
            filename = _('Skolu_ataskaita(') + date_from + '_' + date_to + ')' + XLS_EXT
        else:
            filename = _('Skolu_ataskaita(') + date + ')' + XLS_EXT
        attach_id = self.env['ir.attachment'].create({
            'res_model': 'debt.act.wizard',
            'res_id': self[0].id,
            'type': 'binary',
            'name': filename,
            'datas_fname': filename,
            'datas': base64_file
        })
        if self._context.get('archive', False):
            return base64_file
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/download?res_model=debt.act.wizard&res_id=%s&attach_id=%s' % (self[0].id, attach_id.id),
            'target': 'self',
        }

    def _get_bare_currency_data(self, currency_id):
        currency = self.env['res.currency'].browse(currency_id)
        show_name = True if (self.env['res.currency'].search_count([('symbol', '=', currency.symbol)]) > 1) else False
        if not currency:
            data = {
                'display_name': 'Valiuta',
                'position': 'after',
                'symbol': 'Valiuta',
                'show_name': False,
            }
        else:
            data = {
                'display_name': currency.display_name,
                'position': currency.position,
                'symbol': currency.symbol,
                'show_name': show_name
            }
        return data

    def get_raw_residual_move_line_data(self, partner_id, account_ids, date):
        '''returns  {aml_id: (amount_cmp, amount_cur, cur)}  '''
        company_currency = self.env.user.company_id.currency_id
        all_account_move_line_domain = [('partner_id', '=', partner_id),
                                        ('date', '<=', date),
                                        ('move_id.state', '=', 'posted'),
                                        ('account_id.id', 'in', account_ids)]
        all_account_move_line_ids = self.env['account.move.line'].search(all_account_move_line_domain).ids
        later_reconciled = self.env['account.partial.reconcile'].search(['|',
                                                                             '&',
                                                                                 ('credit_move_id', 'in', all_account_move_line_ids),
                                                                                 ('debit_move_id.date', '>', date),
                                                                             '&',
                                                                                 ('debit_move_id', 'in', all_account_move_line_ids),
                                                                                 ('credit_move_id.date', '>', date),
                                                                         ])
        later_reconciled_move_line_ids = later_reconciled.mapped('credit_move_id.id') + later_reconciled.mapped(
            'debit_move_id.id')
        filtered_move_lines = self.env['account.move.line'].search([('id', 'in', all_account_move_line_ids),
                                                                    '|',
                                                                        ('reconciled', '=', False),
                                                                        ('id', 'in', later_reconciled_move_line_ids)])
        amount_residual_company_curr_by_id = dict([(aml.id, aml.amount_residual) for aml in filtered_move_lines])
        amount_residual_currency_by_id = dict([(aml.id, aml.amount_residual_currency) for aml in filtered_move_lines])
        for apr in later_reconciled:
            if apr.credit_move_id.id in amount_residual_company_curr_by_id:
                amount_residual_company_curr_by_id[apr.credit_move_id.id] -= apr.amount
            if apr.debit_move_id.id in amount_residual_company_curr_by_id:
                amount_residual_company_curr_by_id[apr.debit_move_id.id] += apr.amount
            if apr.credit_move_id.id in amount_residual_currency_by_id:
                amount_residual_currency_by_id[apr.credit_move_id.id] -= apr.amount_currency
            if apr.debit_move_id.id in amount_residual_currency_by_id:
                amount_residual_currency_by_id[apr.debit_move_id.id] += apr.amount_currency
        res = {}
        for aml in filtered_move_lines:
            amount_company = amount_residual_company_curr_by_id.get(aml.id, 0.0)
            if aml.currency_id and aml.currency_id != company_currency:
                currency = aml.currency_id
                amount_currency = amount_residual_currency_by_id.get(aml.id, 0.0)
            else:
                currency = company_currency
                amount_currency = 0
            if float_compare(amount_company, 0, precision_rounding=company_currency.rounding) == 0 and \
                    float_compare(amount_currency, 0, precision_rounding=currency.rounding) == 0:
                continue
            res[aml.id] = (amount_company, amount_currency, currency.id)
        return res

    def get_raw_turnover_move_line_data(self, partner_id, account_ids, date_from, date_to,
                                        include_full_reconcile_ids=None):
        # full reconcile cancels out to zero
        company_currency = self.env.user.company_id.currency_id
        all_account_move_line_domain = [('partner_id', '=', partner_id),
                                        ('move_id.state', '=', 'posted'),
                                        ('account_id.id', 'in', account_ids),
                                        ('date', '<=', date_to)]
        if date_from:
            all_account_move_line_domain.append(('date', '>=', date_from))
        if include_full_reconcile_ids:
            all_account_move_line_domain.extend(['|',
                                                     ('full_reconcile_id', '=', False),
                                                     ('full_reconcile_id', 'in', include_full_reconcile_ids)])
        account_move_lines = self.env['account.move.line'].search(all_account_move_line_domain, order='date asc')
        res = {}
        for aml in account_move_lines:
            amount_company = aml.balance
            if aml.currency_id and aml.currency_id != company_currency:
                currency = aml.currency_id
                amount_currency = aml.amount_currency
            else:
                currency = company_currency
                amount_currency = 0
            if float_compare(amount_company, 0, precision_rounding=company_currency.rounding) == 0 and float_compare(
                    amount_currency, 0, precision_rounding=currency.rounding) == 0:
                continue
            res[aml.id] = (amount_company, amount_currency, currency.id)
        return res

    def convert_to_report_data(self, raw_data):
        '''returns {aml_id: (amount, report_currency)}'''
        company_currency = self.env.user.company_id.currency_id
        res = {}
        for aml_id, (amount_company, amount_currency, currency_id) in iteritems(raw_data):
            aml = self.env['account.move.line'].browse(aml_id)
            report_currency = aml.get_cluster_currency_id()
            if report_currency == company_currency.id:
                res[aml_id] = (amount_company, report_currency)
            elif currency_id == report_currency:
                res[aml_id] = (amount_currency, report_currency)
            else:
                res[aml_id] = (0.0, report_currency)
        return res

    def get_all_account_move_line_data(self, partner_id, account_ids, report_type, date, date_from, date_to):
        company_currency = self.env.user.company_id.currency_id
        company_currency_id = company_currency.id
        starting_balances = {}
        report_data = {}
        raw_data = []
        irreconcilable_data = []
        if report_type == 'unreconciled':
            if date is None:
                raise exceptions.UserError('Report date is not set')
            accounts = self.env['account.account'].browse(account_ids)
            reconcilable_account_ids = accounts.filtered(lambda x: x.reconcile).ids
            # Irreconcilable accounts may be selected when generating report
            # Their total amounts should be additionally included in the report
            irreconcilable_account_ids = accounts.filtered(lambda x: not x.reconcile).ids
            if reconcilable_account_ids:
                raw_data = self.get_raw_residual_move_line_data(partner_id, reconcilable_account_ids, date)
                report_data = self.convert_to_report_data(raw_data)
            if irreconcilable_account_ids:
                irreconcilable_data = self.env['report.debt.reconciliation.base'].sudo().\
                    get_irreconcilable_account_amount(partner_id, irreconcilable_account_ids, date)
        else:
            if date_from is None or date_to is None:
                raise exceptions.UserError('Report date range is not set')
            self._cr.execute('''SELECT DISTINCT(full_reconcile_id) from account_move_line where date >= %s''',
                             (date_from,))
            full_reconcile_ids = [row[0] for row in self._cr.fetchall() if row[0]]
            date_to_historical = fields.Date.to_string(fields.Date.from_string(date_from) - timedelta(days=1))
            historical_raw_data = self.get_raw_turnover_move_line_data(partner_id, account_ids, False,
                                                                       date_to_historical,
                                                                       include_full_reconcile_ids=full_reconcile_ids)
            raw_data = self.get_raw_turnover_move_line_data(partner_id, account_ids, date_from, date_to)
            report_data = self.convert_to_report_data(raw_data)
            historical_report_data = self.convert_to_report_data(historical_raw_data)
            for (value, currency_id) in itervalues(historical_report_data):
                starting_balances.setdefault(currency_id, 0)
                starting_balances[currency_id] += value
        full_data = {}
        for aml in self.env['account.move.line'].browse(report_data.keys()).sorted(lambda r: r.date):
            aml_id = aml.id
            amount = report_data[aml.id][0]
            report_currency = report_data[aml_id][1]
            currency_data = full_data.setdefault(report_currency, {})
            currency_data.setdefault('credit', 0)
            currency_data.setdefault('debit', 0)
            # if float_compare(amount, 0, precision_digits=2) == 0:
            #     continue
            credit = -amount if amount < 0 else 0.0
            debit = amount if amount > 0 else 0.0
            ref = aml.name
            date_operation = ''
            if aml.move_id.statement_line_id:  # todo advance payment has invoice_id but shouldnt be there
                doc_type = 'payment'
            elif aml.invoice_id:
                doc_type = 'invoice'
                date_operation = aml.invoice_id.operacijos_data or ''
                if aml.invoice_id.type in ['out_invoice', 'out_refund']:
                    ref = aml.invoice_id.number
            else:
                doc_type = ''
            orig_currency = raw_data[aml.id][2]
            if orig_currency == company_currency_id:
                orig_amount = raw_data[aml.id][0]
            else:
                orig_amount = raw_data[aml.id][1]
            lines = currency_data.setdefault('lines', [])
            line = {'ref': ref,
                    'credit': credit,
                    'debit': debit,
                    'date': aml.date,
                    'date_operation': date_operation,
                    'date_maturity': aml.date_maturity,
                    'doc_type': doc_type,
                    'orig_amount': orig_amount,
                    'orig_currency': self.env['res.currency'].browse(raw_data[aml.id][2]),
                    'account_code': aml.account_id.code
                    }
            currency_data['debit'] += debit
            currency_data['credit'] += credit
            lines.append(line)

        # Additionally include total amounts of irreconcilable accounts
        for data_line in irreconcilable_data:
            account = self.env['account.account'].browse(data_line['account_id'])
            currency_data = full_data.setdefault(company_currency_id, {})
            currency_data.setdefault('credit', 0)
            currency_data.setdefault('debit', 0)
            lines = currency_data.setdefault('lines', [])
            balance = data_line['debit'] - data_line['credit']
            debit = balance if balance > 0 else 0
            credit = -balance if balance < 0 else 0
            line = {'ref': account.name,
                    'credit': credit,
                    'debit': debit,
                    'date': date,
                    'date_operation': '',
                    'date_maturity': '',
                    'doc_type': 'account',
                    'orig_amount': debit - credit,
                    'orig_currency': company_currency,
                    'display_original': False,
                    'account_code': account.code
                    }
            currency_data['debit'] += debit
            currency_data['credit'] += credit
            lines.append(line)

        for currency_id in full_data:
            full_data[currency_id]['start_balance'] = starting_balances.get(currency_id, 0.0)
            full_data[currency_id]['end_balance'] = starting_balances.get(currency_id, 0.0) + \
                                                    full_data[currency_id]['debit'] - full_data[currency_id]['credit']
        for currency_id in starting_balances:
            if currency_id not in full_data:
                full_data[currency_id] = {'lines': [],
                                          'start_balance': starting_balances[currency_id],
                                          'end_balance': starting_balances[currency_id],
                                          'debit': 0.0,
                                          'credit': 0.0}

        if not report_data and not irreconcilable_data:
            full_data[company_currency_id] = {'lines': [],
                                              'start_balance': 0.0,
                                              'end_balance': 0.0,
                                              'debit': 0.0,
                                              'credit': 0.0}
        return full_data

    def _get_account_name(self, account_code):
        return '%s %s' % (account_code,
                          self.env['account.account'].search([('code', '=', account_code)], limit=1).name or '')

    def _get_partner_data(self, partner_id):
        return self.env['res.partner'].browse(partner_id)

    @api.multi
    def render_html(self, doc_ids, data=None):
        report_obj = self.env['report']
        report = report_obj._get_report_from_name('skolu_suderinimas.report_aktas_multi_minimal')
        partner_ids = data['partner_ids']
        partners = self.env['res.partner'].browse(partner_ids)
        date = data['date']
        date_from = data['date_from']
        date_to = data['date_to']
        default_account_domain = data.get('account_ids') if data.get('account_ids') \
            else self.env['report.debt.reconciliation.base'].get_default_payable_receivable(data['account_type_filter'])
        report_type = data['type']
        data_by_partner_id = {}
        for partner_id in partner_ids:
            data_by_partner_id[partner_id] = self.get_all_account_move_line_data(partner_id, default_account_domain,
                                                                                 report_type, date, date_from, date_to)

        if data['dont_show_zero_debts']:
            partners_not_to_show = list()
            for partner in partners:
                False if any(not float_is_zero(data_by_partner_id[partner.id][currency]['debit']
                                               - data_by_partner_id[partner.id][currency]['credit'], precision_digits=2)
                             for currency in data_by_partner_id[partner.id]) \
                    else partners_not_to_show.append(partner.id)
            if len(partners_not_to_show) > 0:
                partner_ids = list(set(partner_ids).difference(set(partners_not_to_show)))
                partners = self.env['res.partner'].browse(partner_ids)
                data_by_partner_id = {}
                for partner_id in partner_ids:
                    data_by_partner_id[partner_id] = self.get_all_account_move_line_data(partner_id,
                                                                                         default_account_domain,
                                                                                         report_type, date, date_from,
                                                                                         date_to)

        if data['dont_show_zero_values']:
            partners_not_to_show = list()
            for partner in partners:
                partner_data = data_by_partner_id[partner.id]
                credit_debit_value = 0
                for currency in partner_data:
                    credit_debit_value += partner_data[currency]['credit']
                    credit_debit_value += partner_data[currency]['debit']
                if float_is_zero(credit_debit_value, precision_digits=2):
                    partners_not_to_show.append(partner.id)
            if len(partners_not_to_show) > 0:
                partner_ids = list(set(partner_ids).difference(set(partners_not_to_show)))
                partners = self.env['res.partner'].browse(partner_ids)
                data_by_partner_id = {}
                for partner_id in partner_ids:
                    data_by_partner_id[partner_id] = self.get_all_account_move_line_data(partner_id,
                                                                                         default_account_domain,
                                                                                         report_type, date, date_from,
                                                                                         date_to)
        data_by_account_code = {}
        if data['detail_level'] == 'sum' and data['show_accounts']:
            data_by_account_code = get_data_by_account_code(data_by_partner_id)

        docargs = {
            'doc_ids': partner_ids,
            'doc_model': report.model,
            'docs': partners,
            'date': data['date'],
            'date_from': data['date_from'],
            'date_to': data['date_to'],
            'data_by_partner_id': data_by_partner_id,
            'data_by_account_code': data_by_account_code,
            'company': self.env.user.company_id,
            'get_bare_currency_data': self._get_bare_currency_data,
            'formatLang': lambda *a, **kw: formatLang(self.env, *a, **kw),
            'get_partner_data': self._get_partner_data,
            'get_account_name': self._get_account_name,
            'type': data['type'],
            'detail_level': data['detail_level'],
            'account_type_filter': data['account_type_filter'],
            'show_original_amounts': data['show_original_amounts'],
            'show_accounts': data['show_accounts'],
            'payment_reminder': data.get('payment_reminder', False),
        }
        # TODO DUPLICATE CODE, IMPORT FROM aktas.py, clean up excel code, styles are defined on every method, etc.
        return docargs


AccReportGeneralLedgerSL()