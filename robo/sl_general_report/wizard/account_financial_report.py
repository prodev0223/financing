# -*- coding: utf-8 -*-
import datetime
from dateutil.relativedelta import relativedelta
from odoo import api, fields, models, tools, http, _, exceptions
from xlutils.filter import process, XLRDReader, XLWTWriter
import openpyxl as px
import os
from sys import platform
import StringIO
import re
import unicodedata
from odoo.tools import ustr

XLS_EXT = 'xlsx'

try:
    import slugify as slugify_lib
except ImportError:
    slugify_lib = None


def slugify(s, max_length=None):
    """ Transform a string to a slug that can be used in a url path.
        This method will first try to do the job with python-slugify if present.
        Otherwise it will process string by stripping leading and ending spaces,
        converting unicode chars to ascii, lowering all chars and replacing spaces
        and underscore with hyphen "-".
        :param s: str
        :param max_length: int
        :rtype: str
    """
    s = ustr(s)
    if slugify_lib:
        # There are 2 different libraries only python-slugify is supported
        try:
            return slugify_lib.slugify(s, max_length=max_length)
        except TypeError:
            pass
    uni = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    slug_str = re.sub('[\W_]', ' ', uni).strip().lower()
    slug_str = re.sub('[-\s]+', '-', slug_str)

    return slug_str[:max_length]


def slug(value):
    if isinstance(value, models.BaseModel):
        if isinstance(value.id, models.NewId):
            raise ValueError("Cannot slug non-existent record %s" % value)
        # [(id, name)] = value.name_get()
        identifier, name = value.id, value.display_name
    else:
        # assume name_search result tuple
        identifier, name = value
    slugname = slugify(name or '').strip().strip('-')
    if not slugname:
        return str(identifier)
    return "%s-%d" % (slugname, identifier)

# NOTE: as the pattern is used as it for the ModelConverter (ir_http.py), do not use any flags
_UNSLUG_RE = re.compile(r'(?:(\w{1,2}|\w[A-Za-z0-9-_]+?\w)-)?(-?\d+)(?=$|/)')


def unslug(s):
    """Extract slug and id from a string.
        Always return un 2-tuple (str|None, int|None)
    """
    m = _UNSLUG_RE.match(s)
    if not m:
        return None, None
    return m.group(1), int(m.group(2))

class GeneralReportExcel:
    def __init__(self):
        self.wb = None
        self.worksheet = None

    def load_template(self, template_name):
        if platform == 'win32':
            xls_flocation = '\\data\\xls\\%s.%s' % (template_name, XLS_EXT)
        else:
            xls_flocation = '/data/xls/%s.%s' % (template_name, XLS_EXT)
        file_loc = os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + xls_flocation
        self.wb = px.load_workbook(file_loc)
        self.worksheet = self.wb.active

    def write_lines(self, data, col):
        base_sheet = self.worksheet
        for line in data:
            if line.get('row', False) and line.get('col', False) and not line.get('formula', False):
                cell = base_sheet.cell(line['row'] + 1, line['col'] + col + 1)
                cell.value = line['balance']

            elif line.get('row', False) and line.get('col', False) and \
                    line.get('formula', False) and line.get('formula_prev', False):
                if col:
                    cell = base_sheet.cell(line['row'] + 1, line['col'] + col + 1)
                    cell.value = '=' + line['formula_prev']
                else:
                    cell = base_sheet.cell(line['row'] + 1, line['col'] + 1)
                    cell.value = '=' + line['formula']

    def write_header(self, res, currency, address, name, template, bottom_text, ceo, ceo_job, date_from,
                     preliminary=False, lang='lt_LT'):
        base_sheet = self.worksheet
        set_mark = False
        mark_text = '(Tvirtinimo žyma)' if lang == 'lt_LT' else '(Approval mark)'
        if preliminary:
            set_mark = True
            mark_text = 'Preliminari ataskaita' if lang == 'lt_LT' else 'Preliminary report'

        date_to = datetime.datetime.strptime(res['data']['form']['date_to'], tools.DEFAULT_SERVER_DATE_FORMAT)
        from_to = str(date_from)[:10] + '-' + str(date_to)[:10]
        today_str = datetime.datetime.today().strftime('%Y-%m-%d')
        if template == 'Balansas0':
            base_sheet.cell(3, 3, name)
            base_sheet.cell(6, 3, address)
            base_sheet.cell(21, 5, currency)
            base_sheet.cell(21, 1, from_to)
            base_sheet.cell(18, 1, today_str)
            base_sheet.cell(132, 1, bottom_text)
            base_sheet.cell(130, 1, ceo_job)
            base_sheet.cell(130, 5, ceo)
            if set_mark:
                base_sheet.cell(13, 6, mark_text)
        elif template == 'Balansas00':
            base_sheet.cell(3, 3, name)
            base_sheet.cell(6, 3, address)
            base_sheet.cell(19, 5, currency)
            base_sheet.cell(19, 1, from_to)
            base_sheet.cell(16, 1, today_str)
            base_sheet.cell(45, 1, bottom_text)
            base_sheet.cell(43, 1, ceo_job)
            base_sheet.cell(43, 5, ceo)
            if set_mark:
                base_sheet.cell(12, 6, mark_text)
        elif template == 'Balansas000':
            base_sheet.cell(3, 3, name)
            base_sheet.cell(6, 3, address)
            base_sheet.cell(19, 5, currency)
            base_sheet.cell(19, 1, from_to)
            base_sheet.cell(16, 1, today_str)
            base_sheet.cell(54, 1, bottom_text)
            base_sheet.cell(51, 1, ceo_job)
            base_sheet.cell(51, 5, ceo)
            if set_mark:
                base_sheet.cell(12, 6, mark_text)
        elif template == 'Balansas1':
            base_sheet.cell(5, 1, name)
            base_sheet.cell(8, 1, address)
            base_sheet.cell(19, 5, currency)
            base_sheet.cell(19, 1, from_to)
            base_sheet.cell(16, 1, today_str)
            base_sheet.cell(92, 1, bottom_text)
            base_sheet.cell(90, 1, ceo_job)
            base_sheet.cell(90, 5, ceo)
            if set_mark:
                base_sheet.cell(12, 6, mark_text)
        elif template == 'Balansas11':
            base_sheet.cell(6, 4, name)
            base_sheet.cell(8, 4, address)
            base_sheet.cell(22, 6, currency)
            base_sheet.cell(18, 2, 'PAGAL %s DUOMENIS' % str(date_to)[:10])
            base_sheet.cell(19, 2, today_str)
            base_sheet.cell(50, 2, bottom_text)
            base_sheet.cell(48, 2, ceo_job)
            base_sheet.cell(48, 5, ceo)
            if set_mark:
                base_sheet.cell(14, 6, mark_text)
        elif template == 'Balansas2':
            base_sheet.cell(3, 2, name)
            base_sheet.cell(6, 2, address)
            base_sheet.cell(20, 5, currency)
            base_sheet.cell(20, 2, from_to)
            base_sheet.cell(17, 3, today_str)
            base_sheet.cell(50, 1, bottom_text)
            base_sheet.cell(48, 1, ceo_job)
            base_sheet.cell(48, 5, ceo)
            if set_mark:
                base_sheet.cell(12, 6, mark_text)
        elif template == 'Pelnas0':
            base_sheet.cell(5, 2, name)
            base_sheet.cell(8, 2, address)
            base_sheet.cell(22, 4, currency)
            base_sheet.cell(22, 2, from_to)
            base_sheet.cell(19, 1, today_str)
            base_sheet.cell(45, 1, bottom_text)
            base_sheet.cell(43, 1, ceo_job)
            base_sheet.cell(43, 5, ceo)
            if set_mark:
                base_sheet.cell(15, 4, mark_text)
        elif template == 'Pelnas00':
            base_sheet.cell(5, 2, name)
            base_sheet.cell(8, 2, address)
            base_sheet.cell(22, 4, currency)
            base_sheet.cell(22, 2, from_to)
            base_sheet.cell(19, 1, today_str)
            base_sheet.cell(39, 1, bottom_text)
            base_sheet.cell(37, 1, ceo_job)
            base_sheet.cell(37, 5, ceo)
            if set_mark:
                base_sheet.cell(15, 4, mark_text)
        elif template == 'Pelnas1':
            base_sheet.cell(5, 1, name)
            base_sheet.cell(8, 1, address)
            base_sheet.cell(21, 3, currency)
            base_sheet.cell(21, 2, from_to)
            base_sheet.cell(18, 1, today_str)
            base_sheet.cell(52, 1, bottom_text)
            base_sheet.cell(50, 1, ceo_job)
            base_sheet.cell(50, 6, ceo)
            if set_mark:
                base_sheet.cell(12, 5, mark_text)
        elif template == 'Pelnas11':
            base_sheet.cell(8, 1, name)
            base_sheet.cell(10, 1, address)
            base_sheet.cell(26, 3, currency)
            base_sheet.cell(22, 1, 'PAGAL %s DUOMENIS' % str(date_to)[:10])
            base_sheet.cell(23, 1, today_str)
            base_sheet.cell(42, 1, bottom_text)
            base_sheet.cell(40, 1, ceo_job)
            base_sheet.cell(40, 5, ceo)
            if set_mark:
                base_sheet.cell(17, 5, mark_text)
        elif template == 'Pelnas2':
            base_sheet.cell(6, 2, name)
            base_sheet.cell(9, 2, address)
            base_sheet.cell(23, 4, currency)
            base_sheet.cell(23, 2, from_to)
            base_sheet.cell(20, 1, today_str)
            base_sheet.cell(40, 1, bottom_text)
            base_sheet.cell(38, 1, ceo_job)
            base_sheet.cell(38, 5, ceo)
            if set_mark:
                base_sheet.cell(15, 5, mark_text)

    def export(self):
        f = StringIO.StringIO()
        self.wb.save(f)
        return f.getvalue().encode('base64')


class AccReport(models.TransientModel):
    _inherit = "accounting.report"
    # account.common.report

    def get_default_rep(self):
        return self.env.user.company_id.uab_report_size

    def _date_to_default(self):
        return datetime.datetime.utcnow()

    def _date_from_default(self):
        return self.env.user.company_id.compute_fiscalyear_dates()['date_from']

    def _default_force_lang(self):
        return 'en_US' if self.env.user.partner_id.lang == 'en_US' else 'lt_LT'

    @api.model
    def _get_account_report(self):
        rep_type = self._context.get('report_type', False)
        if self.env.user.company_id.company_activity_form == 'mb':
            activity_form = 'mb'
        elif self.env.user.company_id.company_activity_form == 'vsi':
            activity_form = 'vsi'
        else:
            activity_form = 'uab'
        reports = self.env['account.financial.report'].search([('report_type', '=', rep_type),
                                                               ('parent_id', '=', False),
                                                               ('activity_form', '=', activity_form)])
        return reports and reports[0] or False

    account_report_id = fields.Many2one('account.financial.report', string='Finansinė ataskaita', required=True, default=_get_account_report)

    date_from = fields.Date(required=True, default=_date_from_default)
    date_to = fields.Date(required=True, default=_date_to_default)
    display_account = fields.Selection([('all', 'Visas'), ('movement', 'Su įrašais'),],  # ('not_zero', 'With balance is not equal to 0'),
                                       string='Rodyti sąskaitas', required=True, default='movement')
    hierarchy_level = fields.Selection([('1', '1'), ('2', '2'), ('3', '3'), ('4', '4'), ('5', '5'),
                                        ('6', '6'), ('7', '7')],
                                       # ('not_zero', 'With balance is not equal to 0'),
                                       string='Hierarchy Level', required=True, default='1')

    reduced_uab_report = fields.Boolean(default=False, string='Sutrumpinta ataskaita')

    uab_report_size = fields.Selection([('max', 'Išplėstinis balansas'),
                                        ('mid', 'Sutrumpintas balansas'),
                                        ('min', 'Trumpas balansas')], string='Balanso dydis', default=get_default_rep)

    vsi_report_version = fields.Selection([('1', 'Iki 2018-12-31'),
                                           ('11', 'Nuo 2019-01-01')], string='Variantas', default='11')

    do_show_reduced = fields.Boolean(compute='_do_show_reduced', default=False)
    enable_filter = fields.Boolean(string='Enable Comparison')
    activity_form = fields.Char(compute='_activity_form')
    report_type = fields.Char(compute='_report_type')
    balance_hierarchy_warning = fields.Boolean(compute='_compute_balance_hierarchy_warning')
    force_lang = fields.Selection(default=_default_force_lang)
    ignore_balance_leveling_error = fields.Boolean(string='Ignoruoti balanso nesutapimo klaidas')

    @api.multi
    @api.depends('hierarchy_level', 'account_report_id.report_type')
    def _compute_balance_hierarchy_warning(self):
        """
        Compute //
        Determine whether warning should be presented in the form:
            If report to be exported is of 'BL' type, and it's
            hierarchy level is higher than 1, display the warning
        :return: None
        """
        for rec in self:
            if rec.account_report_id.report_type == 'BL' and rec.hierarchy_level != '1':
                rec.balance_hierarchy_warning = True

    @api.model
    def level_out_balance_lines(self, data):
        """
        Level out balance report data. Round balance lines, then compare upper and lower balance sides,
        if there is a difference, level out the upper and lower lines by adding or subtracting the difference
        :param data: XLS report data (list-of-dicts)
        :return: None
        """
        upper_amount = lower_amount = 0.0
        upper_leveling_line = lower_leveling_line = None
        company_currency = self.sudo().env.user.company_id.currency_id

        for line in data:
            # Determine the balance side
            balance_side = line.get('balance_line_side')
            if not balance_side:
                raise exceptions.ValidationError(
                    _('Balanso eilutės privalo turėti pažymėtą XLS poziciją (viršutinė/apatinė)'))

            # Determine balance leveling line
            if line.get('balance_leveling_line'):
                if balance_side == 'upper' and not upper_leveling_line:
                    upper_leveling_line = line
                elif balance_side == 'lower' and not lower_leveling_line:
                    lower_leveling_line = line

            # If line does not contain formula, round the balance to the whole number
            if not line.get('formula') and line.get('account_type') != 'sum':
                line_balance = round(line.get('balance', 0.0))
                line['balance'] = line_balance

                # Calculate upper and lower balance
                if balance_side == 'upper':
                    upper_amount += line_balance
                else:
                    lower_amount += line_balance

        side_diff = tools.float_round(abs(upper_amount) - abs(lower_amount), precision_digits=2)
        # Check whether side difference exists
        if not tools.float_is_zero(side_diff, precision_digits=2):

            # Determine which side is to be leveled
            if tools.float_compare(side_diff, 0.0, precision_digits=2) > 0:
                leveling_line = lower_leveling_line
            else:
                leveling_line = upper_leveling_line
            # After determining the leveling side, ABS the difference
            side_diff = abs(side_diff)

            # If it does, check whether leveling line exists
            if not leveling_line:
                raise exceptions.ValidationError(
                    _('Klaida balanso spausdinime. Nerasta išlyginamoji eilutė'))

            # Check whether balance error should be ignored or not
            if self.ignore_balance_leveling_error:
                return

            # If difference between upper and lower side is bigger than 5
            if tools.float_compare(abs(side_diff), 5.0, precision_digits=2) > 0:
                error = _('Klaida balanso spausdinime. Viršutinės ir apatinės dalies skirtumas didesnis nei 5 {}. '
                          'Patikrinkite ar praeitų metų įrašai yra uždaryti arba susisiekite '
                          'su administratoriais').format(company_currency.name or 'EUR')
                raise exceptions.ValidationError(error)

            # Level out the balance line
            leveling_line['balance'] += side_diff

    @api.onchange('activity_form', 'reduced_uab_report', 'uab_report_size', 'vsi_report_version')
    def get_report_id(self):
        if self._context.get('balansas', False):
            if self.activity_form == 'uab' and self.uab_report_size == 'max':
                self.account_report_id = self.env.ref('sl_general_report.balansas0').id
            elif self.activity_form == 'uab' and self.uab_report_size == 'min':
                self.account_report_id = self.env.ref('sl_general_report.balansas00').id
            elif self.activity_form == 'uab' and self.uab_report_size == 'mid':
                self.account_report_id = self.env.ref('sl_general_report.balansas000').id
            elif self.activity_form == 'vsi' and self.vsi_report_version == '11':
                self.account_report_id = self.env.ref('sl_general_report.balansas11').id
            elif self.activity_form == 'vsi':
                self.account_report_id = self.env.ref('sl_general_report.balansas1').id
            elif self.activity_form == 'mb':
                self.account_report_id = self.env.ref('sl_general_report.balansas2').id
        else:
            if self.activity_form == 'uab' and not self.reduced_uab_report:
                self.account_report_id = self.env.ref('sl_general_report.pelnas0').id
            elif self.activity_form == 'uab' and self.reduced_uab_report:
                self.account_report_id = self.env.ref('sl_general_report.pelnas00').id
            elif self.activity_form == 'vsi' and self.vsi_report_version == '11':
                self.account_report_id = self.env.ref('sl_general_report.pelnas11').id
            elif self.activity_form == 'vsi':
                self.account_report_id = self.env.ref('sl_general_report.pelnas1').id
            elif self.activity_form == 'mb':
                self.account_report_id = self.env.ref('sl_general_report.pelnas2').id

    @api.one
    @api.depends('activity_form')
    def _do_show_reduced(self):
        if self.activity_form == 'uab':
            self.do_show_reduced = True
        else:
            self.do_show_reduced = False

    @api.one
    @api.depends('account_report_id', 'reduced_uab_report')
    def _activity_form(self):
        if self.env.user.company_id.company_activity_form == 'mb':
            self.activity_form = 'mb'
        elif self.env.user.company_id.company_activity_form == 'vsi':
            self.activity_form = 'vsi'
        else:
            self.activity_form = 'uab'

    @api.one
    @api.depends('account_report_id')
    def _report_type(self):
        self.report_type = self._context.get('report_type', False)

    @api.multi
    def _print_report(self, data):
        ctx = self._context.copy()
        user = self.env.user
        company = user.company_id
        lang = company.partner_id.lang if company.partner_id.lang else ctx.get('lang')
        ctx.update({'lang': self.force_lang or lang})
        self = self.with_context(ctx)
        data['form'].update(self.read(
            ['date_from_cmp', 'debit_credit', 'date_to_cmp', 'filter_cmp', 'account_report_id', 'enable_filter',
             'label_filter', 'target_move', 'display_account', 'hierarchy_level'])[0])
        data['form']['used_context'].update({'active_id': self.id})
        return self.env['report'].get_action(self, 'sl_general_report.report_financial_sl', data=data)

    @api.multi
    def check_report(self):
        res = super(AccReport, self).check_report()
        if 'report_type' in res:
            if self._context.get('force_pdf'):
                res['report_type'] = 'qweb-pdf'
            if self._context.get('force_html'):
                res['report_type'] = 'qweb-html'
        return res

    @api.multi
    def _get_xls_template_name(self):
        """
        Return the XLS template name that matches the generated report
        :return: XLS template filename
        :rtype: str
        """
        self.ensure_one()
        template = ''
        if self.account_report_id.report_type == 'BL':
            if self.account_report_id.id == self.env.ref('sl_general_report.balansas0').id:
                template = 'Balansas0'
            elif self.account_report_id.id == self.env.ref('sl_general_report.balansas00').id:
                template = 'Balansas00'
            elif self.account_report_id.id == self.env.ref('sl_general_report.balansas000').id:
                template = 'Balansas000'
            elif self.account_report_id.id == self.env.ref('sl_general_report.balansas1').id:
                template = 'Balansas1'
            elif self.account_report_id.id == self.env.ref('sl_general_report.balansas11').id:
                template = 'Balansas11'
            elif self.account_report_id.id == self.env.ref('sl_general_report.balansas2').id:
                template = 'Balansas2'
        else:
            if self.account_report_id.id == self.env.ref('sl_general_report.pelnas0').id:
                template = 'Pelnas0'
            elif self.account_report_id.id == self.env.ref('sl_general_report.pelnas00').id:
                template = 'Pelnas00'
            elif self.account_report_id.id == self.env.ref('sl_general_report.pelnas1').id:
                template = 'Pelnas1'
            elif self.account_report_id.id == self.env.ref('sl_general_report.pelnas11').id:
                template = 'Pelnas11'
            elif self.account_report_id.id == self.env.ref('sl_general_report.pelnas2').id:
                template = 'Pelnas2'
        return template

    @api.multi
    def _get_xls_out_filename(self):
        """
        Get report filename style depending on the report
        :return: filename base (to be completed with dates)
        :rtype: str
        """
        filename_end = '_Balansas.xlsx' if self.account_report_id.report_type == 'BL' else '_Pelnas(nuostolis).xlsx'
        return '_' + slugify(self.env.user.company_id.name) + filename_end

    @api.multi
    def xls_export(self):
        self.ensure_one()
        company = self.env.user.company_id

        # Raise an error on XLS export if balance hierarchy warning is True
        if self.balance_hierarchy_warning:
            raise exceptions.UserError(_('Excel eksportas galimas tik tada, kai pasirinktas 1 hierarchijos lygmuo'))

        template = self._get_xls_template_name()
        if not template:
            raise exceptions.UserError(_('Nepavyko sugeneruoti ataskaitos'))
        filename = self._get_xls_out_filename()

        res = self.check_report()
        lines = self.get_account_lines(res['data']['form'])

        date_to = datetime.datetime.strptime(res['data']['form']['date_to'], tools.DEFAULT_SERVER_DATE_FORMAT)
        date_from_calc = datetime.datetime.strptime(res['data']['form']['date_from'], tools.DEFAULT_SERVER_DATE_FORMAT)
        current_fiscal_year_date = company.compute_fiscalyear_dates(date=date_to)
        previous_fiscal_year_date = company.compute_fiscalyear_dates(date=date_to + relativedelta(years=-1))
        date_from = current_fiscal_year_date['date_from']
        if date_from_calc > date_from:
            date_from = date_from_calc
        date_to_previous = previous_fiscal_year_date['date_to']
        date_from_previous = previous_fiscal_year_date['date_from']

        filename = str(date_from)[:10] + '-' + str(date_to)[:10] + filename
        lines.sort(key=lambda l: (l.get('sequence', 0), l.get('code', '0')))
        excel = GeneralReportExcel()
        print_lang = 'lt_LT'
        if self.force_lang == 'en_US':
            try:
                excel.load_template(template + '_en')
                print_lang = 'en_US'
            except IOError:
                excel.load_template(template)
        else:
            excel.load_template(template)

        address = u'{0} {1} {2} {3}'.format(company.street or u'', company.city or u'', company.zip or u'',
                                            company.country_id.display_name or u'')

        static_text = 'ROBOLABS UAB accounting services' if print_lang == 'en_US' \
            else 'Apskaitos paslaugas teikianti UAB "ROBOLABS"'
        company_name = (company.name or u'') + u' ' + (company.company_registry or u'')
        company_ceo = company.vadovas.name_related
        ceo_job = self.with_context(lang=print_lang).env.user.company_id.vadovas.job_id.name

        preliminary = bool(self._context.get('robo_front'))

        excel.write_header(
            res,
            company.currency_id.display_name or u'',
            address,
            company_name, template, static_text, company_ceo, ceo_job, date_from,
            preliminary, lang=print_lang
        )

        res['data']['form'].update({'date_to': date_to_previous})
        res['data']['form']['used_context'].update({'date_to': date_to_previous})
        if self.account_report_id.report_type == 'PL':
            res['data']['form'].update({'date_from': date_from_previous})
            res['data']['form']['used_context'].update({'date_from': date_from_previous})
        prev_lines = self.get_account_lines(res['data']['form'])

        # Round and level out balance lines
        if self.account_report_id.report_type == 'BL':
            self.level_out_balance_lines(lines)
            self.level_out_balance_lines(prev_lines)

        # Write excel lines, col signifies XLS spreadsheet column
        prev_period_col = 3 if template == 'Pelnas1' else 1
        excel.write_lines(lines, col=0)
        excel.write_lines(prev_lines, col=prev_period_col)
        base64_file = excel.export()

        attach_id = self.env['ir.attachment'].create({
            'res_model': 'accounting.report',
            'res_id': self[0].id,
            'type': 'binary',
            'name': filename,
            'datas_fname': filename,
            'datas': base64_file
        })
        if self._context.get('archive', False):
            return base64_file
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/download?res_model=accounting.report&res_id=%s&attach_id=%s' % (
            self[0].id, attach_id.id),
            'target': 'self',
        }

    @api.onchange('date_to')
    def onchange_date_to(self):
        if self.date_to and self._context.get('balansas', False):
            self.date_from = '1900-01-01'

    def get_account_lines(self, data):
        lines = []
        account_report = self.env['account.financial.report'].search([('id', '=', data['account_report_id'][0])])
        get_all_movement = True if data['display_account'] == 'all' else False
        child_reports = account_report._get_children_by_order()
        child_reports = child_reports.sorted(key=lambda r: r.sequence)
        ReportObj = self.env['report.sl_general_report.report_financial_sl']
        res = ReportObj.with_context(data.get('used_context'))._compute_report_balance(child_reports)
        self = self.with_context({'show_views': True})
        if data['enable_filter']:
            comparison_res = ReportObj.with_context(data.get('comparison_context'))._compute_report_balance(child_reports)
            for report_id, value in comparison_res.items():
                res[report_id]['comp_bal'] = value['balance']
                report_acc = res[report_id].get('account')
                if report_acc:
                    for account_id, val in comparison_res[report_id].get('account').items():
                        report_acc[account_id]['comp_bal'] = val['balance']

        acc_views = self.env['account.account'].with_context(show_views=True).search([('is_view', '=', True)],
                                                                                     limit=1,
                                                                                     order='hierarchy_level DESC')
        if acc_views:
            account_hierarchy = acc_views[0].hierarchy_level
        else:
            account_hierarchy = 11
        for i in range(account_hierarchy, 0, -1):
            res = ReportObj.calculate_view_balance(res, i)
        report_sequence = 0
        for report in child_reports:
            vals = {
                'account_id': 0,
                'name': report.name,
                'balance': res[report.id]['balance'] * report.sign,
                'type': 'report',
                'level': bool(report.style_overwrite) and report.style_overwrite or report.level,
                'label': report.label,
                'hierarchy_level': 1,
                'account_type': report.type or False,
                'sequence': report_sequence,
                'code': report.code,
                'col': report.template_col,
                'row': report.template_row,
                'formula': report.formula,
                'formula_prev': report.formula_prev,
                'balance_leveling_line': report.balance_leveling_line,
                'balance_line_side': report.balance_line_side,
                'report_type': report.report_type,
            }
            if data['debit_credit']:
                vals['debit'] = res[report.id]['debit']
                vals['credit'] = res[report.id]['credit']

            if data['enable_filter']:
                vals['balance_cmp'] = res[report.id]['comp_bal'] * report.sign

            lines.append(vals)
            if report.display_detail == 'no_detail':
                continue
            if res[report.id].get('account'):
                report_sequence += 1
                for account_id, value in res[report.id]['account'].items():
                    flag = False
                    account = self.env['account.account'].browse(account_id)
                    name = ''
                    vals = {
                        'account_id': account.id,
                        'code': account.code,
                        'name': name + account.name,
                        'balance': value['balance'] * report.sign or 0.0,
                        'type': 'account',
                        'level': report.display_detail == 'detail_with_hierarchy' and 4,
                        'hierarchy_level': account.hierarchy_level,  # self.get_hierarchy_level(account.code)
                        'account_type': account.internal_type,
                        'sequence': report_sequence,
                    }
                    if data['debit_credit']:
                        vals['debit'] = value['debit']
                        vals['credit'] = value['credit']
                        if not account.company_id.currency_id.is_zero(
                                vals['debit']) or not account.company_id.currency_id.is_zero(vals['credit']):
                            flag = True
                    if not account.company_id.currency_id.is_zero(vals['balance']):
                        flag = True
                    if data['enable_filter']:
                        vals['balance_cmp'] = value['comp_bal'] * report.sign
                        if not account.company_id.currency_id.is_zero(vals['balance_cmp']):
                            flag = True
                    if (flag or get_all_movement) and int(vals['hierarchy_level']) <= int(data['hierarchy_level']):
                        lines.append(vals)
            report_sequence += 1
        return lines

AccReport()
