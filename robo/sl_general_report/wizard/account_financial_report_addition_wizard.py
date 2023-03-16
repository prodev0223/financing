# -*- coding: utf-8 -*-
import openpyxl as px
import os
from StringIO import StringIO
from sys import platform
from datetime import datetime
from dateutil.relativedelta import relativedelta
from odoo import fields, models, api, tools, _
from six import iteritems, itervalues
from . account_financial_report import slugify

REPORT_MAPPING = {
    'assets': {
        'line': 169,
        'reports': ['report_2'],
    },
    'inventory_and_prepayments': {
        'line': 174,
        'reports': ['report_39'],
    },
    'buyers_debt': {
        'line': 180,
        'reports': ['report_51'],
    },
    'other_receivable': {
        'line': 181,
        'reports': ['report_52', 'report_53', 'report_54'],
    },
    'cash': {
        'line': 188,
        'reports': ['report_58'],
    },
    'financial_debt': {
        'line': 197,
        'reports': ['report_92', 'report_101'],
    },
    'debt_to_suppliers': {
        'line': 198,
        'reports': ['report_94', 'report_103'],
    },
    'received_pepayment': {
        'line': 199,
        'reports': ['report_93', 'report_102'],
    },
    'income_tax_liabilities': {
        'line': 200,
        'reports': ['report_107'],
    },
    'employment_liabilities': {
        'line': 201,
        'reports': ['report_108'],
    },
    'other_payable_and_liabilities': {
        'line': 202,
        'reports': ['report_98', 'report_109'],
    },
    'debt_liabilities': {
        'line': 203,
        'reports': ['report_91', 'report_100'],
    },
}

DATE_ROWS = ['167', '172', '177', '186', '195']
SHAREHOLDER_ROW_START = 13

class ReportExcel:
    """
    Extend the GeneralReportExcel class to fit requirements for this report
    """

    def __init__(self, filename=None):
        """
        Instance constructor
        :param filename: filename in the sl_general_report/data/xls/ directory
        """
        self.wb = None
        if filename:
            self.load_workbook(filename)

    def load_workbook(self, filename):
        """
        Load a file from sl_general_report/data/xls/ directory
        :param filename: file name with extension
        :return: None
        """
        if platform == 'win32':
            xls_flocation = '\\data\\xls\\' + filename
        else:
            xls_flocation = '/data/xls/' + filename
        file_loc = os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + xls_flocation
        self.wb = px.load_workbook(file_loc)

    def write_first_page(self, data):
        """
        Write the first page 'Rekvizitai' info
        :param data: dictionary containing the data to write
        :return: None
        """
        sheet = self.wb['Rekvizitai']
        sheet['B1'] = data.get('activity_form', '')
        sheet['B2'] = data.get('name', '')
        sheet['B3'] = data.get('company_code', '')
        sheet['B4'] = data.get('address', '')
        sheet['B5'] = data.get('year', '')
        sheet['B6'] = data.get('ceo_name', '')
        sheet['B7'] = data.get('ceo_code', '')
        sheet['B8'] = data.get('avg_employee_nb', '')
        sheet['B9'] = data.get('company_evrk', '')
        sheet['B10'] = data.get('material_assets', '')
        sheet['B11'] = data.get('non_material_assets', '')
        shareholders = data.get('shareholders') or []
        for idx, shareholder in enumerate(shareholders, SHAREHOLDER_ROW_START):
            sheet['B' + str(idx)] = shareholder

    def write_financial_data(self, data, col):
        """
        Writes financial data to the 'Aiškinamasis raštas' sheet
        :param data: a dictionary where items contains row and amount keys
        :param col: a letter or a  number. If number, converts 0 -> A, 1 -> B, etc
        :return: None
        """
        if isinstance(col, int):
            col = chr(65 + col)

        ws = self.wb['Aiškinamasis raštas']
        for line in itervalues(data):
            ws[col + line['row']] = line['amount']

    def export(self):
        """ Save Workbook to base64 encoded object """
        f = StringIO()
        self.wb.save(f)
        return f.getvalue().encode('base64')


class AccountFinancialReportAdditionWizard(models.TransientModel):
    _name = 'account.financial.report.addition.wizard'

    def _default_financial_year(self):
        return str(datetime.now().year - 1)

    financial_year = fields.Selection(selection='_selection_financial_year', default=_default_financial_year,
                                      string='Finansiniai metai', required=True)
    date_from = fields.Date(string='Nuo', compute='_compute_dates')
    date_to = fields.Date(string='Iki', compute='_compute_dates')

    @api.depends('financial_year')
    def _compute_dates(self):
        for rec in self:
            if not rec.financial_year:
                continue
            self.date_from = datetime(int(rec.financial_year), 1, 1).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)
            self.date_to = datetime(int(rec.financial_year), 12, 31).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)

    @api.model
    def _selection_financial_year(self):
        return [(y, y) for y in map(str, range(2017, datetime.today().year + 1))]

    @api.multi
    def _get_xls_template_name(self):
        """ Return the Excel template filename """
        self.ensure_one()
        return 'priedai_prie_finansines.xlsx'

    @api.multi
    def _get_xls_out_filename(self):
        """
        Get report filename style depending on the report
        :return: filename base (to be completed with dates)
        :rtype: str
        """
        date_str = self.date_from[:10] + '_' + self.date_to[:10]
        return date_str + '_' + slugify(self.env.user.company_id.name) + '_Priedai_prie_finansines.xlsx'

    @api.multi
    def prepare_first_page_data(self):
        """
        Prepare data to write on the first page of report
        :return: dictionary of company info
        """
        self.ensure_one()

        address = u'{0} {1} {2} {3}'.format(self.env.user.company_id.street or u'',
                                            self.env.user.company_id.city or u'',
                                            self.env.user.company_id.zip or u'',
                                            self.env.user.company_id.country_id.display_name or u'')
        activity_form_mapping = {
            'vsi': u'VšĮ',
            'uab': u'UAB',
            'mb': u'MB',
        }

        return {
            'name': self.env.user.company_id.name,
            'company_code': self.env.user.company_id.company_registry or u'',
            'activity_form': activity_form_mapping[self.env.user.company_id.company_activity_form],
            'company_evrk': self.env.user.company_id.evrk.name or u'',
            'ceo_name': self.env.user.company_id.vadovas.name_related or u'',
            'ceo_code': self.env.user.company_id.vadovas.identification_id or u'',  #TODO
            'address': address,
            'year': self.financial_year,
            'avg_employee_nb': 0,  # Extended in robo module
        }

    @api.multi
    def get_report_data(self, date):
        """
        Computes the financial data until a given date
        :param date: end date for report
        :return: dictionary with key as in MAPPING, and value another dictionary with key 'row':str and 'amount':float
        """
        report_data = {}
        ReportObj = self.env['report.sl_general_report.report_financial_sl']
        compute_balance = ReportObj.with_context(date_to=date, state='posted')._compute_report_balance
        for name, data in iteritems(REPORT_MAPPING):
            row = data['line']
            reports = [self.env.ref('sl_general_report.' + report) for report in data['reports']]
            amount = round(sum(compute_balance(report)[report.id]['balance'] * report.sign for report in reports))
            report_data[name] = {
                'row': str(row),
                'amount': amount,
            }
        return report_data

    @api.multi
    def create_attachment(self, filename, base64_file):
        """
        Create the attachment to download the exported file
        :param filename: str
        :param base64_file: bqse64 encoded file data.
        :return: ir.attachment record
        """
        self.ensure_one()
        attachment = self.env['ir.attachment'].create({
            'res_model': self._name,
            'res_id': self[0].id,
            'type': 'binary',
            'name': filename,
            'datas_fname': filename,
            'datas': base64_file
        })
        return attachment

    @api.multi
    def download_excel(self):
        """ Download the excel file for the selected financial year """
        template = self._get_xls_template_name()
        filename = self._get_xls_out_filename()
        excel = ReportExcel()
        excel.load_workbook(template)

        front_page_data = self.prepare_first_page_data()
        excel.write_first_page(front_page_data)

        data = self.get_report_data(self.date_to)
        excel.write_financial_data(data, 'I')

        date_to = datetime.strptime(self.date_to, tools.DEFAULT_SERVER_DATE_FORMAT)
        date_to_previous = (date_to  - relativedelta(years=1)).strftime(tools.DEFAULT_SERVER_DATE_FORMAT)

        data = self.get_report_data(date_to_previous)
        excel.write_financial_data(data, 'K')

        base64_file = excel.export()

        attachment = self.create_attachment(filename, base64_file)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/download?res_model=account.financial.report.addition.wizard&res_id=%s&attach_id=%s' % (
            self[0].id, attachment.id),
            'target': 'self',
        }


AccountFinancialReportAdditionWizard()
