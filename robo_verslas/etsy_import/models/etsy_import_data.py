# -*- coding: utf-8 -*-
import base64
import io
import threading
import openpyxl as px
import types
import odoo
from datetime import datetime
from six import iteritems
from odoo import models, fields, api, _, exceptions, tools
from odoo.api import Environment
from odoo.addons.queue_job.job import job, identity_exact

FIELD_MAPPING = {
    'Sale Date': 'sale_date',  # A column
    'Item Name': 'item_name',  # B column
    'Quantity': 'quantity',  # D column
    'Discount Amount': 'discount_amount',  # H column
    'Shipping Discount': 'shipping_discount',  # I column
    'Order Shipping': 'order_shipping',  # J column
    'Item Total': 'item_total',  # L column
    'Currency': 'currency',  # M column
    'Date Paid': 'date_paid',  # P column
    'Full Name': 'ship_name',  # R column
    'Street 1': 'ship_address',  # S column
    'Street 2': 'ship_address_2',  # T column
    'Ship City': 'ship_city',  # U column
    'Ship State': 'ship_state',  # V column
    'Ship Zipcode': 'zip',  # W column
    'Ship Country': 'ship_country',  # X column
    'Order ID': 'order_id',  # Y column
    'Order Total': 'order_total'
}

DATE_FORMATS = ['%m/%d/%y', '%m/%d/%Y']


def get_all_values(row):
    return [cell.value for cell in row]


def get_mapped(values, mapping=None):
    if not mapping:
        mapping = FIELD_MAPPING
    result = []
    for val in values:
        mapped = mapping[val] if val in mapping else 'unknown'
        result.append(mapped)
    return result


def convert_to_string(a):
    res = a
    if isinstance(a, types.FloatType):
        a_rounded = tools.float_round(a, precision_digits=2)
        if tools.float_compare(a, a_rounded, precision_digits=2) != 0:
            raise exceptions.UserError(_('%s is not an integer and cannot be converted to text'))
        res = str(int(round(a_rounded)))
    if isinstance(a, (types.IntType, types.LongType)):
        res = str(a)
    if not isinstance(res, basestring):
        if not a:
            return ''
        else:
            return a
    else:
        res = res.strip()
    return res


class ImportRecord(object):
    def __init__(self, vals=None, header=None):
        if vals is None:
            vals = []
        if header is None:
            header = []
        self.vals = vals
        self.header = header

    def __getattr__(self, attr):
        res = False
        if not self.vals or not self.header:
            return False
        try:
            index = self.header.index(attr)
        except ValueError:
            return False
        if len(self.vals) >= index:
            res = self.vals[index]
            try:
                res = res.strip()
            except:
                pass
            # Don't try to convert everything to date
            if 'date' in attr or 'data' in attr:
                for fmt in DATE_FORMATS:
                    try:
                        date = datetime.strptime(res, fmt)
                    except (ValueError, TypeError):
                        pass
                    else:
                        res = date
                        break
        return res

    def get_dict(self, attrs=None):
        if attrs is None:
            attrs = []
        result = {}
        if attrs:
            for attr in attrs:
                result[attr] = self.__getattr__(attr)
        else:
            for attr in self.header:
                result[attr] = self.__getattr__(attr)
        return result


class EtsyDataImport(models.TransientModel):
    _name = 'etsy.data.import'

    @api.model
    def default_get(self, fields_list):
        res = super(EtsyDataImport, self).default_get(fields_list)
        ProductProduct = self.env['product.product'].with_context(lang='lt_LT')
        product = ProductProduct.search([('name', '=', 'Parduotos prekės'),], limit=1)
        service = ProductProduct.search([('name', '=', 'Parduodamos paslaugos'),], limit=1)
        if product:
            res['default_product_id'] = product.id
        if service:
            res['default_service_id'] = service.id
        return res

    xlsx_data = fields.Binary(string='Excel file', required=True)
    xlsx_name = fields.Char(string='Excel file name', size=128, required=False)
    job_id = fields.Many2one('etsy.data.import.job')
    default_product_id = fields.Many2one('product.product', string='Product', required=True)
    default_service_id = fields.Many2one('product.product', string='Service', required=True)

    @api.multi
    def data_import(self):
        """
        Read data from XLSX file and prepare it for further account.invoice creation
        :return: None
        """
        self.ensure_one()
        active_jobs = self.env['etsy.data.import.job'].search([('state', '=', 'in_progress')])
        if active_jobs:
            raise exceptions.UserError(_("You can't perform this action. The XLSX file is currently being imported!"))

        job = self.env['etsy.data.import.job'].create({
            'execution_start_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
            'state': 'in_progress',
            'imported_file_name': self.xlsx_name
        })
        self.write({'job_id': job.id})
        self.with_delay(eta=10).data_import_thread()
        action = self.env.ref('etsy_import.action_etsy_data_import_job')
        return action.read()[0]

    @api.model
    def process_xlsx_file(self, xlsx_file):
        file_to_import = io.BytesIO(base64.decodestring(xlsx_file))
        try:
            wb = px.load_workbook(file_to_import)
            ws = wb.worksheets
        except:
            raise exceptions.UserError(_('Invalid file format'))

        non_taxable_tax_id = self.env['account.tax'].search([
            ('code', '=', 'Ne PVM'),
            ('type_tax_use', '=', 'sale'),
            ('price_include', '=', False),
        ], limit=1)
        pvm_12_tax_id = self.env['account.tax'].search([
            ('code', '=', 'PVM12'),
            ('type_tax_use', '=', 'sale'),
            ('price_include', '=', False),
        ], limit=1)

        journal = self.env['account.journal'].search([('type', '=', 'sale')], limit=1)
        europe_countries = self.with_context(lang='en_US').env.ref('base.europe').country_ids.mapped('name')
        grouped_by_order_id = dict()

        for sheet in ws:
            iter_rows = iter(sheet.iter_rows())
            header = []
            header_mapped = []
            for i, row in enumerate(iter_rows):
                if not header:
                    header = get_all_values(row)
                    if len(set(val for val in header if val)) == 1:
                        # Sheet has title row
                        header = None
                        continue
                    header_mapped = get_mapped(header)
                    continue

                processed_row = self.process_import_row(row, header_mapped, non_taxable_tax_id, europe_countries,
                                                        pvm_12_tax_id, journal)
                if not processed_row:
                    continue

                order_id = processed_row.get('order_id')
                processed_invoice_line_vals = processed_row.get('invoice_line_values', {})
                if not grouped_by_order_id.get(order_id):
                    grouped_by_order_id[order_id] = processed_row
                elif processed_invoice_line_vals:
                    grouped_by_order_id[order_id]['invoice_line_values'] += processed_invoice_line_vals

        return grouped_by_order_id

    @api.model
    def process_import_row(self, row, header_mapped, non_taxable_tax_id, europe_countries, pvm_12_tax_id, journal):
        values = get_all_values(row)
        if len(set(values)) == 1:
            return

        record = ImportRecord(values, header_mapped)
        ship_name = convert_to_string(record.ship_name)
        sale_date_dt = record.sale_date
        ship_country = convert_to_string(record.ship_country)
        item_name = convert_to_string(record.item_name)
        order_total = float(record.order_total)
        item_total = float(record.item_total)
        discount_amount = float(record.discount_amount)
        order_shipping = float(record.order_shipping)
        shipping_discount = float(record.shipping_discount)
        order_id = convert_to_string(record.order_id)
        quantity = 1
        currency = convert_to_string(record.currency)
        currency_id = self.env['res.currency'].search([('name', '=ilike', currency)], limit=1)
        if not currency_id:
            raise exceptions.UserError(_('Wrong currency: "%s", where Order ID: "%s"') % (currency, order_id))

        date_paid_dt = record.date_paid
        partner = self.env['res.partner'].search([('name', 'ilike', ship_name)], limit=1)
        if not partner:
            partner = self.create_partner(record)

        total_price = float(order_total)
        line_tax = non_taxable_tax_id if ship_country in europe_countries else pvm_12_tax_id
        product = self.default_product_id if item_name != 'Return Shipping' else self.default_service_id

        order_values = {
            'partner_id': partner.id,
            'journal_id': journal.id,
            'sale_date': sale_date_dt,
            'currency': currency_id.id,
            'date_paid': date_paid_dt,
            'price_include': line_tax.price_include,
            'order_id': order_id,
            'invoice_line_values': [{
                'product_id': product.id,
                'name': item_name or product.name,
                'price_unit': total_price / quantity,
                'quantity': quantity,
                'account_id': product.get_product_income_account(return_default=True).id,
                'invoice_line_tax_ids': [(6, 0, [line_tax.id])],
            }]
        }
        return order_values

    @api.model
    def create_partner(self, record):
        """
        Create new partner
        :param record: imported record values
        """
        ship_country = convert_to_string(record.ship_country)
        country = self.env['res.country'].with_context(lang='en_US').search([('name', '=ilike', ship_country)], limit=1)
        partner = self.env['res.partner'].create({
            'name': convert_to_string(record.ship_name).upper(),
            'street': convert_to_string(record.ship_address),
            'street2': convert_to_string(record.ship_address_2),
            'city': convert_to_string(record.ship_city),
            'zip': convert_to_string(record.zip),
            'country_id': country.id,
        })
        return partner

    @api.multi
    def create_invoices(self, grouped_by_order_id):
        """
        Create account.invoice record from passed XLSX data
        :param grouped_by_order_id: XLSX data. Format - [{}, {}...]
        """
        company = self.env.user.company_id
        order_ids = grouped_by_order_id.keys()
        existing_order_ids = self.env['account.invoice'].search([
            ('reference', 'in', order_ids)
        ]).mapped('reference')
        invoices = self.env['account.invoice']
        for order_id, data in iteritems(grouped_by_order_id):
            if order_id in existing_order_ids:
                continue
            total_price = 0.0
            for invoice_line_val in data.get('invoice_line_values', {}):
                total_price += invoice_line_val.get('total_price', 0.0)
            if tools.float_compare(total_price, 0, precision_digits=2) < 0:
                invoice_type = 'out_refund'
            else:
                invoice_type = 'out_invoice'
            invoice_values = {
                'company_id': company.id,
                'currency_id': data['currency'],
                'type': invoice_type,
                'journal_id': data['journal_id'],
                'date_invoice': data['sale_date'],
                'operacijos_data': data['sale_date'],
                'partner_id': data['partner_id'],
                'external_invoice': True,
                'imported_api': True,
                'invoice_line_ids': [(0, 0, line_vals) for line_vals in data.get('invoice_line_values', [])],
                'reference': order_id,
                'price_include_selection': 'inc' if data['price_include'] else 'exc',
            }
            try:
                invoices |= self.env['account.invoice'].sudo().create(invoice_values)
            except Exception as e:
                raise exceptions.UserError(
                    _("""Invoice creation error:
                    Where order id: "%s" row in Excel file
                     Error message: %s""") % (invoice_values['reference'], e.args[0]))

        return invoices

    @api.multi
    @job
    def data_import_thread(self):
        self.ensure_one()
        invoices = self.env['account.invoice']
        try:
            processed_data = self.process_xlsx_file(self.xlsx_data)
            invoices |= self.create_invoices(processed_data)
            invoices.action_invoice_open_multi_preprocess()
        except Exception as exc:
            self.env.cr.rollback()
            self.job_id.write({
                'state': 'failed',
                'fail_message': str(exc.args and exc.args[0] or exc),
                'execution_end_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
            })
        else:
            self.job_id.write({
                'state': 'finished',
                'execution_end_date': datetime.utcnow().strftime(tools.DEFAULT_SERVER_DATETIME_FORMAT),
                'created_ids': [(6, 0, invoices.ids)]
            })
